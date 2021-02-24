#Работа с базами данных
import pandas as pd
import glob,warnings,datetime,os,xlrd,csv,openpyxl,shutil,threading,pyodbc,time,numpy
from sending import send_all,send_me,send_epid,send_admin
from sqlalchemy import *
from sqlalchemy.orm import sessionmaker
from openpyxl.utils.dataframe import dataframe_to_rows
import win32com.client
from reports import short_report
from multiprocessing import Process
from multiprocessing.pool import ThreadPool
from reports import short_report

warnings.filterwarnings('ignore')

con = create_engine(os.getenv('sql_engine'),convert_unicode=False)
conn = pyodbc.connect(os.getenv('sql_conn'))

class my_except(Exception):
    pass

def get_dir(name):
    sql = f"SELECT Directory FROM [robo].[directions_for_bot] where NameDir = '{name}'"
    df = pd.read_sql(sql,con)
    return df.iloc[0,0] 

def check_table(name):
    sql=f"""
    SELECT distinct case when (cast([datecreated] as date) = cast(getdate() as date))
        then 1
        else 0
        end AS 'Check'
            FROM [dbo].[cv_{name}]
    """
    return pd.read_sql(sql,conn).iat[0,0]

def search_file(category):
    path = os.getenv('path_robot') + '\\' + datetime.datetime.now().strftime("%Y_%m_%d")
    if category == 'fr':
        pattern = path + r'\Федеральный регистр лиц, больных *[!ИАЦ].xlsx'
    if category == 'fr_death':
        pattern = path + r'\*Умершие пациенты*.xlsx'
    if category == 'fr_lab':
        pattern = path + r'\*Отчёт по лабораторным*.xlsx'
    if category == 'UMSRS':
        pattern = path + r'\*УМСРС*.xlsx'
    file_excel = glob.glob(pattern)
    file_csv = glob.glob(pattern[:-4] + 'csv')
    if len(file_csv):
        return True, True, file_csv[0]
    else:
        if len(file_excel):
            return True, False, file_excel[0]
        else:
            return False, False, None

def check_file(file,category):
    if category == 'fr':
        names = [
    'п/н','Дата создания РЗ','УНРЗ','Дата изменения РЗ','СНИЛС','ФИО','Пол','Дата рождения','Диагноз','Диагноз установлен','Осложнение основного диагноза','Субъект РФ','Медицинская организация','Ведомственная принадлежность','Вид лечения','Дата исхода заболевания','Исход заболевания','Степень тяжести','Посмертный диагноз','ИВЛ','ОРИТ','МО прикрепления','Медицинский работник'
                ]
    if category == 'fr_death':
        names = [
    '1', '2', '3', '4', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22','23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '34', '35', '36', '37', '38', '39', '40', '41', '42','43', '44', '45', '46', '47', '48', '49', '50', '51', '52', '53'
                ]
    if category == 'fr_lab':
        names = [
    'Субъект', 'УНРЗ', 'Мед. организация', 'Основной диагноз', 'Наименование лаборатории', 'Дата лабораторного теста','Тип лабораторного теста', """Результат теста
(положительный/
отрицательный)""", 'Этиология пневмония','Дата первого лабораторного подтверждения COVID-19','Дата последнего лабораторного подтверждения COVID-19'
                ]
    if category == 'UMSRS':
        names = [
    '№ п/п', 'Номер свидетельства о смерти', 'Дата выдачи', 'Категория МС', 'Фамилия', 'Имя', 'Отчество', 'Пол', 'Дата рождения','Дата смерти', 'Возраст', 'Страна', 'Республика', 'Субъект', 'Область', 'Район', 'Город', 'Населенный пункт', 'Элемент планировочной структуры','Район СПБ', 'Улица', 'Дом', 'Корпус', 'Строение', 'Квартира', 'Страна смерти', 'Республика смерти','Субъект смерти', 'Область смерти','Район смерти', 'Город смерти', 'Населенный пункт смерти', 'Элемент планировочной структуры смерти', 'Район СПБ смерти', 'Улица смерти','Дом смерти', 'Корпус смерти', 'Строение смерти', 'Квартира смерти', 'Место смерти', 'Код МКБ-10 а', 'Болезнь или состояние, непосред приведшее к смерти','Код МКБ-10 б', 'Патол. состояние, кот. привело к указанной причине', 'Код МКБ-10 в', 'Первоначальная причина смерти', 'Код МКБ-10 г','Внешняя причина при травмах и отравлениях','Код II', 'Прочие важние состояния', 'Код МКБ-10 а(д)', 'Основное заболевание плода или ребенка','Код МКБ-10 б(д)', 'Другие заболевания плода или ребенка', 'Код МКБ-10 в(д)', 'Основное заболевание матери', 'Код МКБ-10 г(д)','Другие заболевания матери', 'Код МКБ-10 д(д)', 'Другие обстоятельства мертворождения', 'Установил причины смерти', 'Адрес МО','Краткое наименование', 'Основания установления причин смерти', 'Осмотр трупа', 'Записи в мед.док.', 'Предшествующего наблюдения','Вскрытие', 'Статус МС', 'Взамен', 'Дубликат', 'Испорченное', 'Напечатано', 'в случае смерти результате ДТП'
                ]
    sum_colum = len(names)
    names_found = []
    num_colum = 0
    try:
        file =  pd.read_csv(file,header=None, delimiter=';', engine='python')
    except:
        return [False,'Файл не читается','',0]
    for head in range(len(file)):
        if num_colum != sum_colum:
            coll = file.loc[head].tolist()
            num_colum = 0
            collum = []
            check = True
            error_text = '' 
            for name in names:
                k = 0
                f = 0
                for col in coll:
                    if str(col).replace(' ','') == name.replace(' ',''): 
                        collum.append(k)
                        names_found.append(name)
                        num_colum += 1 
                        f = 1
                    k += 1
        else:
            break
    if len(names_found) < sum_colum:
        check = False
        error_text = ' Не найдена колонка ' + str(list(set(names) - set(names_found) ) ) + ';'
        return check,error_text,collum, head - 1
    return check,error_text,collum, head - 1

def slojit_fr(a):
    def read_part_df(excel,number):
        nameSheetShablon = "Sheet1"
        df = pd.read_excel(excel, sheet_name= nameSheetShablon, dtype = str, skiprows = 1, head= 1)
        send_all('прочтен файлик номер '+ list_numbers[number])
        return df
    def file_1(svod):
        with pd.ExcelWriter(new_fedreg) as writer:
            svod.to_excel(writer,index=False)
        return 1
    def file_2(svod):
        df = svod
        del df['СНИЛС']
        del df['ФИО']
        with pd.ExcelWriter(new_iach) as writer:
            df.to_excel(writer,index=False)
        return 1
    pathFolderFedRegParts = os.getenv('path_robot') +r'\_ФР_по_частям'
    date = datetime.datetime.today().strftime("%Y_%m_%d")
    _list = []
    list_numbers = ['раз', 'двас', 'трис']
    pool = ThreadPool(processes=2)
    i = 0
    for excel in glob.glob(pathFolderFedRegParts + r'\Федеральный регистр*.xlsx'):
        thread = pool.apply_async(read_part_df,(excel,i)).get()
        _list.append(thread)
        i+=1

    svod = pd.DataFrame() 
    svod = pd.concat(_list)

    svod = svod[svod["Дата создания РЗ"].notnull()] 
    svod["п/н"] = range(1, len(svod)+1)
    
    new_fedreg = pathFolderFedRegParts + r'\Федеральный регистр лиц, больных COVID-19 - ' + date + '.xlsx'
    new_iach = pathFolderFedRegParts + r'\Федеральный регистр лиц, больных COVID-19 - ' + date + '_ИАЦ.xlsx'
#    shutil.copyfile(pathFolderFedRegParts + r'\ФР.xlsx', new_fedreg)
#    shutil.copyfile(pathFolderFedRegParts + r'\ФР_ИАЦ.xlsx', new_iach)
#    wb= openpyxl.load_workbook(new_fedreg)
#    ws = wb[nameSheetShablon]
#    rows = dataframe_to_rows(svod, index=False, header=False)
#    for r_idx, row in enumerate(rows, 1):
#        for c_idx, value in enumerate(row, 1):
#            ws.cell(row=r_idx, column=c_idx, value=value)
#    wb.save(new_fedreg)
#    svod = svod.drop(['СНИЛС', 'ФИО'], axis=1)
#    wb= openpyxl.load_workbook(new_iach)
#    ws = wb[nameSheetShablon]
#    rows = dataframe_to_rows(svod, index=False, header=False)
    try:
        del df['Unnamed: 24']
    except:
        pass

    pool = ThreadPool(processes=2)
    pool.apply_async(file_1,(svod,)).get()
    pool.apply_async(file_2,(svod,)).get()

#    for r_idx, row in enumerate(rows, 3):
#        for c_idx, value in enumerate(row, 1):
#            ws.cell(row=r_idx, column=c_idx, value=value)
#    wb.save(new_iach)

    NumberForMG = len(svod[svod['Диагноз'].isin(['U07.1']) \
                    & svod['Исход заболевания'].isnull() \
                    & svod['Вид лечения'].isin(['Стационарное лечение'])])

    NumberFor1 = len(svod[svod['Диагноз'].isin(['U07.1']) ])

    NumberFor2 = len(svod[svod['Посмертный диагноз'].isin(['U07.1']) \
                        & svod['Исход заболевания'].isin(['Смерть'])])

    day = pd.to_datetime(svod['Дата изменения РЗ'], format='%d.%m.%Y').max().date()
    count_vizd_old = pd.read_sql ("""SELECT [value_count] FROM [robo].[values]
                where id = (select max(id) from [robo].[values] where [value_name] = 'Всего выздоровело от COVID' 
                and date_rows = (select max(date_rows) from [robo].[values] ) )""",con).iat[0,0]
    count_vizd_new = len(svod[svod['Исход заболевания'].isin(['Выздоровление']) & svod['Диагноз'].isin(['U07.1']) ])

    NumberFor3 = count_vizd_new - count_vizd_old
    # расчет людей на стационарном лечении
    svod['Возраст'] = (pd.to_datetime(svod['Диагноз установлен'], format="%d.%m.%Y") - pd.to_datetime(svod['Дата рождения'], format="%d.%m.%Y")) / numpy.timedelta64(1, 'Y')
    NumberFor4 = len(svod.loc[(svod['Исход заболевания'].isnull()) \
            & (svod['Диагноз'].isin(['U07.1','U07.2']) | svod['Диагноз'].str.contains('J1') ) \
            & (svod['Вид лечения'].isin(['Стационарное лечение']))])
    NumberFor5 = len(svod.loc[(svod['Исход заболевания'].isnull()) & ( svod['Возраст'] >= 60) \
            & (svod['Диагноз'].isin(['U07.1','U07.2']) | svod['Диагноз'].str.contains('J1') ) \
            & (svod['Вид лечения'].isin(['Стационарное лечение']))])
    NumberFor6 = len(svod.loc[(svod['Исход заболевания'].isnull()) & ( svod['Возраст'] >= 70) \
            & (svod['Диагноз'].isin(['U07.1','U07.2']) | svod['Диагноз'].str.contains('J1') ) \
            & (svod['Вид лечения'].isin(['Стационарное лечение']))])

    otvet = 'Вроде все получилось \n' + 'По цифрам\n' \
            + 'На стационарном лечении: ' + str(NumberForMG) + '\n' \
            + 'Всего заболело: ' + str(NumberFor1) +'\n' \
            + 'Всего умерло: '+ str(NumberFor2) + '\n' \
            + 'Всего выздоровело за ' + str(day) + ' : '+ str(NumberFor3) + '\n'\
            + 'Сейчас на стационаром лечении: ' + str(NumberFor4) + '\n' \
            + 'Сейчас на стационаром лечении старше 60: ' + str(NumberFor5) + '\n' \
            + 'Сейчас на стационаром лечении старше 70: ' + str(NumberFor6) 
    return otvet

def excel_to_csv_old(file_excel):
    file_csv = file_excel[:-4] + 'csv'
    sheet = xlrd.open_workbook(file_excel).sheet_by_index(0) 
    col = csv.writer(open(file_csv, 'w', newline=""),delimiter=";") 
    
    for row in range(sheet.nrows): 
        col.writerow(sheet.row_values(row))
    return file_csv

def excel_to_csv(file_excel):
    file_csv = file_excel[:-4] + 'csv'
    excel = openpyxl.load_workbook(file_excel)
    sheet = excel.active 
    col = csv.writer(open(file_csv, 'w', newline=""),delimiter=';')
    for r in sheet.rows: 
        col.writerow([cell.value for cell in r]) 
    return file_csv

def sql_execute(sql):
    Session = sessionmaker(bind=con)
    session = Session()
    session.execute(sql)
    session.commit()
    session.close()

def load_fr(a):
    def fr_to_sql(df):
        df  = df[df['Дата создания РЗ'] != '']
        del df ['Ведомственная принадлежность']
        del df ['Осложнение основного диагноза']
        # ==== Репорт о количестве выздоровевших =====
        report = pd.DataFrame()
        report.loc[0,'date_rows'] = pd.to_datetime(df['Дата создания РЗ'],format='%d.%m.%Y').max().date()
        report.loc[0,'value_name'] = 'Всего выздоровело от COVID'
        report.loc[0,'value_count'] = len(df[df['Исход заболевания'].isin(['Выздоровление']) & df['Диагноз'].isin(['U07.1'])])
        report.to_sql('values',con,schema='robo',index=False,if_exists='append')
        # ============
        send_all('Файл в памяти, количество строк: ' + str(len(df)) )
        sql_execute('TRUNCATE TABLE [dbo].[cv_input_fr]')
        send_all('Очистил input_fr')
        df.to_sql('cv_input_fr',con,schema='dbo',if_exists='append',index=False)
        send_all('Загрузил input_fr, запускаю процедуру')
        sql_execute('EXEC [dbo].[cv_Load_FedReg]')
        if check_table('fedreg'):
            send_all('Федеральный регистр успешно загружен')
            return 1
        else:
            send_all('Произошла какая-то проблема с загрузкой фр')
            return 0
    send_all('Я проснулся и хочу грузить фр!')
    search = search_file('fr')
    if search[0] and search[1]:
        send_all('Вот нашёлся такой файлик:\n' + search[2].split('\\')[-1])
        send_all('Сейчас я его проверю...')
        check = check_file(search[2],'fr')
        if check[0]:
            send_all('Файл прошёл проверку, начинаю грузить в память')
            df = pd.read_csv(search[2],header = check[3], usecols = check[2], na_filter = False, dtype = str, delimiter=';', engine='python')
            print(df.head(3))
            fr_to_sql(df)
            return 1
        else:
            send_all('Файл не прошёл проверку!')
            send_all(check[1])
            return 0
    else:
        if search[0]:
            send_all('Нее... я не хочу работать с xlsx, щас конвертирую!')
            file_csv = excel_to_csv(search[2]) 
            send_all('Результат:\n' + file_csv.split('\\')[-1])
            check = check_file(file_csv,'fr')
            if check[0]:
                send_all('Файл прошёл проверку, начинаю грузить в память')
                df = pd.read_csv(file_csv,header = check[3], usecols = check[2], na_filter = False, dtype = str, delimiter=';', engine='python')
                fr_to_sql(df)
                return 1
            else:
                send_all('Файл не прошёл проверку!')
                send_all(check[1])
                return 0
        else:
            send_all('Но я не нашёл файла федерального регистра (((')
            return 0
    return 1

def load_fr_death(a):
    def fr_death_to_sql(df):
        send_all('Обрезаю слишком длинные строки')
        for column in df.columns:
            for i in range(len(df)):
                df.loc[i,column]= str(df.at[i,column])[:255]
        send_all('Убираю Nan из таблицы')
        for column in df.columns:
            df[column] = df[column].str.replace(r'nan','')
        send_all('Отправляю в базу')
        df.to_sql('cv_input_fr_d_all_2',con,schema='dbo',if_exists='replace',index = False)
        send_all('Запускаю процедуры')
        sql_execute("""
                EXEC   [dbo].[Insert_Table_cv_input_fr_d_all_2]
                EXEC   [dbo].[cv_from_d_all_to_d_covid]
                EXEC   [dbo].[cv_Load_FedReg_d_All]
                EXEC   [dbo].[cv_Load_FedReg_d_covid]
                    """)
        send_all('Успешно загружено')
        return 1
    send_all('Пришло время спокойных')
    search = search_file('fr_death')
    if search[0] and search[1]:
        send_all('Найден файл:\n' + search[2].split('\\')[-1])
        send_all('Сейчас мы его рассмотрим...')
        check = check_file(search[2],'fr_death')
        if check[0]:
            send_all('Похоже файл в порядке, попробую загрузить')
            df = pd.read_csv(search[2],header = check[3], usecols = check[2], na_filter = False, dtype = str, delimiter=';', engine='python')
            fr_death_to_sql(df)
            return 1
        else:
            send_all('Что-то с файлом')
            send_all(str(check))
            return 0
    else:
        if search[0]:
            send_all('Давайте всё-таки работать с csv, конвертирую')
            file_csv = excel_to_csv(search[2]) 
            send_all('Результат:\n' + file_csv.split('\\')[-1])
            check = check_file(file_csv,'fr_death')
            if check[0]:
                send_all('Теперь можно и загрузить в память')
                df = pd.read_csv(file_csv,header = check[3], usecols = check[2], na_filter = False, dtype = str, delimiter=';', engine='python')
                fr_death_to_sql(df)
                return 1
            else:
                send_all('Файл не прошёл проверку!')
                send_all(check[1])
                return 0
        else:
            send_all('Не найден файл умерших')
        return 0

def load_fr_lab(a):
    def fr_lab_to_sql(df):
        send_all('Ну, тут надо переименовать колонки и можно грузить')
        i = 1
        for column in df.columns:
            df.rename(columns = {column: str(i)}, inplace = True)
            i+=1
        df = df.dropna(subset=['1','2','3','4','5','6'])
        send_all('Всего строк в  лабе '+ str(len(df)))
        df.to_sql('cv_input_fr_lab_2',con,schema='dbo',if_exists='replace',index = False)
        send_all('Остались процедуры в базе')
        sql_execute("""
                    EXEC   [dbo].[Insert_Table_cv_input_fr_lab_2]
                    EXEC   [dbo].[cv_load_frlab]
                    """)
        if check_table('fedreg_lab'):
            send_all('Лаборатория успешно загружена')
            return 1
        else:
            send_all('Какая-то проблема с загрузкой лаборатории')
            return 0
    send_all('Посмотрим на файлик лабораторных исследований')
    search = search_file('fr_lab')
    if search[0] and search[1]:
        send_all('Найден файл:\n' + search[2].split('\\')[-1])
        send_all('Сейчас мы его рассмотрим...')
        check = check_file(search[2],'fr_lab')
        if check[0]:
            send_all('Неужели файл все-таки можно загрузить?')
            df = pd.read_csv(search[2],header = check[3], usecols = check[2], na_filter = False, dtype = str, delimiter=';', engine='python')
            fr_lab_to_sql(df)
            return 1
        else:
            send_all('Не, плохой файл!')
            send_all(str(check))
            return 0
    else:
        if search[0]:
            send_all('Ну, это точно нужно перевести в csv!')
            file_csv = excel_to_csv(search[2]) 
            send_all('Результат:\n' + file_csv.split('\\')[-1])
            check = check_file(file_csv,'fr_lab')
            if check[0]:
                send_all('Теперь можно и загрузить в память')
                df = pd.read_csv(file_csv,header = check[3], usecols = check[2], na_filter = False, dtype = str, delimiter=';', engine='python')
                fr_lab_to_sql(df)
                return 1
            else:
                send_all('Файл не прошёл проверку!')
                send_all(check[1])
                return 0
        else:
            send_all('Не найден файл лаборатории')
        return 0

def load_UMSRS(a):
    def UMSRS_to_sql(df):
        df.to_sql('cv_input_umsrs_2',con,schema='dbo',if_exists='append',index = False)
        send_all('Данные загружены в input_umsrs_2, запускаю процедурки')
        sql_execute("""
                    EXEC   [dbo].[Insert_Table_cv_input_umsrs_2]
                    EXEC   [dbo].[cv_Load_UMSRS]
                    """)
        if check_table('umsrs'):
            send_all('Успешно выполнено!')
            return 1
        else:
            send_all('Какая-то проблема с загрузкой УМСРС')
            return 0
    send_all('А теперь будем грузить УМСРС')
    search = search_file('UMSRS')
    if search[0] and search[1]:
        send_all('файл уже сконвертирован:\n' + search[2].split('\\')[-1])
        send_all('Посмотрим что внутри...')
        check = check_file(search[2],'UMSRS')
        if check[0]:
            send_all('Файл прошёл проверку, начинаю грузить в память')
            df = pd.read_csv(search[2],header = check[3], usecols = check[2], names = check[2], na_filter = False, dtype = str, delimiter=';', engine='python')
            UMSRS_to_sql(df)
            return 1
        else:
            send_all('Файл не прошёл проверку!')
            send_all(check[1])
            return 0
    else:
        if search[0]:
            send_all('Нее... я не хочу работать с xlsx, щас конвертирую!')
            file_csv = excel_to_csv(search[2]) 
            send_all('Результат:\n' + file_csv.split('\\')[-1])
            check = check_file(file_csv,'UMSRS')
            if check[0]:
                send_all('Файл прошёл проверку, начинаю грузить в память')
                df = pd.read_csv(file_csv,header = check[3], usecols = check[2], names = check[2], na_filter = False, dtype = str, delimiter=';', engine='python')
                UMSRS_to_sql(df)
                return 1
            else:
                send_all('Файл не прошёл проверку!')
                send_all(check[1])
                return 0
        else:
            send_all('Но я не нашёл файла УМСРС! (((')
            return 0 

def load_report_vp_and_cv(a):
    def open_save(file):
        xcl = win32com.client.Dispatch("Excel.Application")
        wb = xcl.Workbooks.Open(file)
        xcl.DisplayAlerts = False
        wb.SaveAs(file)
        xcl.Quit()
    def load_file_mo(file):
        nameMO = pd.read_excel(file, sheet_name= 'Титул', header =3, usecols='H', nrows = 1).iloc[0,0]
        df = pd.read_excel(file, sheet_name= 'Данные1', header =6, usecols='C:AH', nrows = 1)
        df = df.fillna(0)
        df['nameMO'] = nameMO
        os.replace(file, path + r'\\' + os.path.basename(file))
        return df
    def check_data_table(name):
        sql=f"""
            IF (EXISTS (SELECT * FROM {name})) 
                SELECT 1 ELSE SELECT 0 """
        return pd.read_sql(sql,conn).iat[0,0]
    send_all('Продготовка к отчету Мониторинг ВП и COVID')
    files = glob.glob(get_dir('VP_CV') + r'\из_почты\[!~$]*.xls*')
    if len(files) == 0:
        raise my_except('Папка пустая!')
    path = get_dir('VP_CV') + r'\\' + datetime.datetime.now().strftime("%Y%m%d")
    if os.path.exists(path):
        pass
    else:
        try:
            os.mkdir(path)
        except OSError:
            raise my_except('не смог создать папку') 
    list_=[]
    excel = pd.DataFrame()
    for file in files:
        try:
            excel = load_file_mo(file)
        except:
            open_save(file)
            try:
                excel = load_file_mo(file)
            except:
                send_all('не обработался следующий файл \n'+ file.split('\\')[-1])
            else:
                list_.append(excel)
        else:
            list_.append(excel)
    if len(list_):
        df = pd.concat(list_)
    else:
        raise my_except('Не один файл не обработался')
    df = df.set_axis(['vp_03_Power_Count_Departments','vp_04_Power_Count_Allocated_All','vp_05_Power_Count_Allocated','vp_06_Power_Count_Reanimation_All','vp_07_Power_Count_Reanimation','vp_08_Hospital_All','vp_09_Hospital_Day','vp_10_Hospital_Hard_All','vp_11_Hospital_Hard_Reaniamation','vp_12_Hospital_Hard_Ivl','vp_13_ReceivedAll','vp_14_ReceivedDay','vp_15_DischargedAll','vp_16_DischargedDay','vp_17_DiedAll','vp_18_DiedDay','cv_19_Power_Count_Departments','cv_20_Power_Count_Allocated_All','cv_21_Power_Count_Allocated','cv_22_Power_Count_Reanimation_All','cv_23_Power_Count_Reanimation','cv_24_Hospital_All','cv_25_Hospital_Day','cv_26_Hospital_Hard_All','cv_27_Hospital_Hard_Reaniamation','cv_28_Hospital_Hard_Ivl','cv_29_ReceivedAll','cv_30_ReceivedDay','cv_31_DischargedAll','cv_32_DischargedDay','cv_33_DiedAll','cv_34_DiedDay','nameMO'], axis=1, inplace=False)
    try:
        df.to_sql('HistoryFileMO',con,schema='mon_vp',index = False,if_exists='append')
    except:
        raise my_except('не удалось загрузить в базу!')
    if check_data_table('mon_vp.v_DebtorsReport'):
        return short_report('SELECT * FROM mon_vp.v_DebtorsReport')
    else:
        #==========  тут мы грузим исходные данные в первую вкладку отчета
        df = pd.read_sql('SELECT * FROM mon_vp.v_GrandReport' ,conn)
        df1 = df.loc[df.typeMO==1].sort_values(["numSort"]).drop('typeMO',1).drop('numSort',1)
        df2 = df.loc[df.typeMO==2].sort_values(["numSort"]).drop('typeMO',1).drop('numSort',1)
        shablon = get_dir('help') + r'\СводОбщий_' + (datetime.datetime.now() + datetime.timedelta(days=1)).strftime("%d %m %Y") +'.xlsx'
        shutil.copyfile(get_dir('help') + r'\шаблон Мониторинг ВП.xlsx', shablon)

        wb= openpyxl.load_workbook(shablon)
        ws= wb['Данные']
        ws.cell(row=1, column=2, value=(datetime.datetime.now() + datetime.timedelta(days=1)).strftime("%d.%m.%Y"))

        rows = dataframe_to_rows(df1,index=False, header=False)
        for r_idx, row in enumerate(rows,9):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        rows = dataframe_to_rows(df2,index=False, header=False)
        for r_idx, row in enumerate(rows,73):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(shablon)

        #  ========== сейчас мы загрузим и сохраним данные по проверкам

        df=pd.read_sql("exec mon_vp.p_CheckMonitorVpAndCovid",conn)
        part_one = df.iloc[:,range(26)]
        part_two = df.iloc[:,[0] + list(range(26,58,1)) ]

        wb= openpyxl.load_workbook(shablon)
        ws= wb['Проверка']

        rows = dataframe_to_rows(part_one,index=False, header=False)
        for r_idx, row in enumerate(rows,3):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        wb.save(shablon)

        wb= openpyxl.load_workbook(shablon)
        ws= wb['Разница']

        rows = dataframe_to_rows(part_two,index=False, header=False)
        for r_idx, row in enumerate(rows,7):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(shablon)
        try:
            shutil.copyfile(shablon, get_dir('VP_CV') + '\\' + shablon.split('\\')[-1])
        except PermissionError:
            raise my_except('Закройте файлик! Не могу скопировать')
        return shablon

def load_report_guber(a):
    def load_sheet(file, sheetName, ColumsName, startRows, header_):
        df = pd.read_excel(file, sheet_name= sheetName,header = None , usecols=ColumsName,  skiprows = startRows)
        df = df.set_axis(header_, axis=1, inplace=False)
        df["idRows"] =pd.to_numeric(df["idRows"]) 
        df = df.sort_values(["idRows"])
        df = df.loc[df["idRows"].notnull()]
        df = df.drop_duplicates()
        df = df.fillna(0)
        return df
    def load_file(file, sheetName, ColumsName, startRows, header_, tableName):    
        list_.clear()
        try:
            excel = load_sheet(file, sheetName, ColumsName, startRows, header_)
        except:
            pass
        else:
            list_.append(excel)
            df = pd.concat(list_)
        try:
            df.to_sql(tableName,con,schema='mon_vp',index = False,if_exists='append')
        except:
            pass
    def check_data_tab(name):
        sql=f"""
        IF (EXISTS (SELECT * FROM {name})) 
        SELECT 1 ELSE SELECT 0 """
        return pd.read_sql(sql,con).iat[0,0]
    directory = get_dir('MG')
    path = directory + '\\' + datetime.datetime.now().strftime("%Y%m%d")
    try:
        os.mkdir(path)
    except:
        pass
    df = pd.DataFrame()
    list_ = []
    header_vp = ['idRows','nameMO','indicators','vp_Received_Count_All_SPb','vp_Received_Count_All_LO','vp_Received_Count_toDay_Spb','vp_Received_Count_toDay_LO','vp_Discharged_Count_All_SPb','vp_Discharged_Count_All_LO','vp_Discharged_Count_toDay_Spb','vp_Discharged_Count_toDay_LO','vp_Died_Count_All_SPb','vp_Died_Count_All_LO','vp_Died_Count_toDay_Spb','vp_Died_Count_toDay_LO','vp_Hospital_Count_All','vp_Hospital_Count_Spb','vp_Hospital_Count_LO','vp_Hospital_Count_Ivl'] 
    header_cv = ['idRows','nameMO','indicators','cv_Diagnosis_Count_All_SPb','cv_Diagnosis_Count_All_LO','cv_Diagnosis_Count_toDay_Spb','cv_Diagnosis_Count_toDay_LO','cv_Discharged_Count_All_SPb','cv_Discharged_Count_All_LO','cv_Discharged_Count_toDay_Spb','cv_Discharged_Count_toDay_LO','cv_Died_Count_All_SPb','cv_Died_Count_All_LO','cv_Died_Count_toDay_Spb','cv_Died_Count_toDay_LO','cv_Hospital_Count_All','cv_Hospital_Count_Spb','cv_Hospital_Count_LO','cv_Hospital_Count_Ivl']
    header_ivl = ['idRows','nameMO','ivl_Invaz_Count_All','ivl_Invaz_Count_Busy','ivl_Invaz_Count_Free_All','ivl_Invaz_Count_Faulty','ivl_NeInvaz_Count_All','ivl_NeInvaz_Count_Busy','ivl_NeInvaz_Count_Free_All','ivl_NeInvaz_Count_Faulty','ivl_Pacient_Count_All','ivl_Pacient_Count_Covid']
    header_bunk = ['idRows','nameMO','bn_Count_All','bn_Count_Ill_All','bn_Count_Ill_Faulty','bn_Count_Ill_Free']

    for file in glob.glob(directory + r'\из_почты\[!~$]*.xls*'):
        load_file(file, 'Cвод_ОРВИ_и_Пневм', 'A:S', 4, header_vp, 'ReportGubernator_Pnevm')
        load_file(file, 'Свод_COVID', 'A:S', 4, header_cv, 'ReportGubernator_Covid')
        load_file(file, 'Свод_ИВЛ', 'A:L', 3, header_ivl, 'ReportGubernator_Ivl')
        load_file(file, 'Свод_Койки', 'A:F', 2, header_bunk, 'ReportGubernator_Bunk')
        os.replace(file, path + r'\\' + os.path.basename(file))

    if check_data_tab('mon_vp.v_DebtorsReportGubernator'):
        return short_report('SELECT * FROM mon_vp.v_DebtorsReportGubernator')
    else:
        shutil.copyfile(get_dir('help') + r'\big_finger.png', get_dir('temp') + r'\big_finger.png')
        return get_dir('temp') + r'\big_finger.png'
