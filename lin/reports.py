import os,subprocess,datetime,shutil,glob,sqlalchemy
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

server  = os.getenv('server')
server_parus  = os.getenv('server_parus')
user    = os.getenv('mysqldomain') + '\\' + os.getenv('mysqluser') # Тут правильный двойной слеш!
passwd  = os.getenv('mypassword')
dbase   = os.getenv('db')
db_parus   = os.getenv('db_parus')

eng1 = sqlalchemy.create_engine(f"mssql+pymssql://{user}:{passwd}@{server}/{dbase}",pool_pre_ping=True)
try:
    eng2 = sqlalchemy.create_engine(f"mssql+pymssql://{user}:{passwd}@{server_parus}/{db_parus}",pool_pre_ping=True)
    con_parus = eng2.connect()
except:
    print(f"mssql+pymssql://{user}:{passwd}@{server_parus}/{db_parus}")

con = eng1.connect()

def get_dir(name):
    sql = f"SELECT Directory FROM [robo].[directions_for_bot] where NameDir = '{name}' and [linux] = 'True'"
    df = pd.read_sql(sql,con)
    return df.iloc[0,0] 

def sql_execute(sql):
    Session = sessionmaker(bind=con)
    session = Session()
    session.execute(sql)
    session.commit()
    session.close()

def fr_deti(a):
    sql_week    = 'exec [dbo].[Proc_Report_Children_by_week]'
    sql_month   = 'exec [dbo].[Proc_Report_Children_by_month]'
    sql_diagnoz = 'exec [dbo].[Proc_Report_Children_by_diagnoz]'
    file = get_dir('temp') + '/otchet_deti.xlsx'
    week = pd.read_sql(sql_week,con)
    month = pd.read_sql(sql_month,con)
    diagnoz = pd.read_sql(sql_diagnoz,con)
    with pd.ExcelWriter(file) as writer:
        week.to_excel(writer,sheet_name='week',index=False) 
        month.to_excel(writer,sheet_name='month',index=False) 
        diagnoz.to_excel(writer,sheet_name='diagnoz',index=False) 
    return file

def short_report(textsql):
    df = pd.read_sql(textsql,con)
    table_html = get_dir('temp') + '/table.html'
    table_png  = get_dir('temp') + '/table.png'

    df.to_html(table_html,index=False,justify='center')

    command = "/usr/bin/wkhtmltoimage --encoding utf-8 -f png --width 0 " +  table_html + " " + table_png
    try:
        subprocess.check_call(command,  shell=True)
    except:
        raise my_except('Не удалось сделать файл\n' +  subprocess.check_output(command,  shell=True))
    else:
        return table_png

def dead_not_mss(a):
    frNotMss = pd.read_sql('EXEC dbo.p_FedRegNotMss',con)
    date = datetime.datetime.today().strftime("%Y_%m_%d_%H_%M")
    file = get_dir('fr_not_mss') + '/Список_без_МСС_'+ date +'.xlsx'
    with pd.ExcelWriter(file) as writer:
        frNotMss.to_excel(writer,sheet_name='notMSS',index=False)
    return "Список готов\n" + file.split('/')[- 1]

def dynamics(a):
    frNotMss = pd.read_sql('EXEC [dbo].[275_M3]',con)
    date = datetime.datetime.today().strftime("%Y_%m_%d_%H_%M")
    file = get_dir('dynam') + '/Динамика_'+ date +'.xlsx'
    with pd.ExcelWriter(file) as writer:
        frNotMss.to_excel(writer,sheet_name='notMSS',index=False)
    return "Динамика готова\n" + file.split('/')[-1]

def mg_from_guber(a):
    def CreateExcelSvod(nameSheetShablon, startRowsShablons, nameSheetSvod, startRowsSvod, pathSvod, i):
        _list.clear()
        for excel in glob.glob(_path_folder_files_mo) :
            try:
                df = pd1.read_excel(excel, sheet_name=nameSheetShablon, dtype = str, skiprows = startRowsShablons , head= None)
            except:
                pass 
            else:
                _list.append(df)
        svod = pd.DataFrame()
        svod = pd.concat(_list)
        svod["Unnamed: 0"] =pd.to_numeric(svod["Unnamed: 0"])
        svod = svod.sort_values(["Unnamed: 0"])
        svod = svod.loc[svod["Unnamed: 0"].notnull()]
        svod = svod.drop_duplicates()
        
        wb = openpyxl.load_workbook(pathSvod)
        ws = wb[nameSheetSvod]
        rows = dataframe_to_rows(svod,index=False, header=False)

        for r_idx, row in enumerate(rows, startRowsSvod):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        if (i == 1):
            ws['Q2'] = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%d.%m.%Y')
        wb.save(pathSvod)
        return 1
    def UpdateColumsSvod(nameSheet, nameColums, rowStartSvodOld, countRowsSvodOld, rowStartSvodNew, cellStartSvodNew):
        _list.clear()
        try:
            df = pd.read_excel(_path_old_svod, sheet_name= nameSheet , usecols = nameColums, skiprows = rowStartSvodOld, nrows = countRowsSvodOld)
        except:
            pass
        else:
            _list.append(df)
        svod=pd.DataFrame()
        svod=pd.concat(_list)
        wb= openpyxl.load_workbook(_path_svod_vp_cv)
        ws = wb[nameSheet]
        rows = dataframe_to_rows(svod, index=False, header=False)
        for r_idx, row in enumerate(rows, rowStartSvodNew):
            for c_idx, value in enumerate(row, cellStartSvodNew):
                ws.cell(row=r_idx, column=c_idx, value=value)
        wb.save(_path_svod_vp_cv)
        return 1
    _path_folder_report = get_dir('MG')
    _name_file_svod = '/09 стационары для Справки Губернатора'
    _path_folder_files_mo = _path_folder_report + '/Беглов_09/*.xlsx'

    _path_svod_all   = _path_folder_report + _name_file_svod + '_'+ datetime.datetime.now().strftime("%d.%m.%Y_%H_%M") + '.xlsx'
    _path_svod_vp_cv = _path_folder_report + _name_file_svod + '_'+ datetime.datetime.now().strftime("%d.%m.%Y") + '.xlsx'
    _path_old_svod   = _path_folder_report + _name_file_svod + '_'+ (datetime.datetime.now() - datetime.timedelta(days=1)).strftime("%d.%m.%Y") + '.xlsx'

    _list = []
    shutil.copyfile(_path_folder_report + _name_file_svod + '.xlsx', _path_svod_all)
    CreateExcelSvod("Cвод_ОРВИ_и_Пневм", 3, "Cвод_ОРВИ_и_Пневм", 6, _path_svod_all, 0)
    CreateExcelSvod("Свод_COVID", 3, "Свод_COVID", 6, _path_svod_all, 0)
    CreateExcelSvod("Свод_ИВЛ", 2, "Свод_ИВЛ", 5, _path_svod_all, 0)
    CreateExcelSvod("Свод_Койки", 1, "Свод_Койки", 4, _path_svod_all, 0)

    shutil.copyfile(_path_folder_report + _name_file_svod + '2.xlsx', _path_svod_vp_cv)

    CreateExcelSvod("Cвод_ОРВИ_и_Пневм", 3, "Свод", 8, _path_svod_vp_cv, 1)
    CreateExcelSvod("Свод_COVID", 3, "Свод", 97, _path_svod_vp_cv, 1)

    UpdateColumsSvod("Свод","D", 6, 79, 8, 20)
    UpdateColumsSvod("Свод","P", 6, 79, 8, 23)

    UpdateColumsSvod("Свод","D", 94, 200, 96, 20)
    UpdateColumsSvod("Свод","P", 94, 200, 96, 23)
    return 'Все готово'

def parus_43_cov_nulls(a):
    sql="""
    select 
            org.MED_ORGANIZATION_NAME as 'организация',data.BEGIN_DATE as 'Дата',count(*) as 'нулей в накопительных итогах'
      FROM [PARUS].[REPORT_DATA] as data
      inner join PARUS.MED_ORGANIZATIONS as org
            on (data.MED_ORGANIZATION_ID = org.MED_ORGANIZATION_ID)
      inner join PARUS.FORMS as form
            on (data.FORM_ID = form.FORM_ID)
      inner join PARUS.INDICATORS as ind
            on (data.INDICATOR_ID = ind.INDICATOR_ID)
    where form.FORM_CODE = '43 COVID 19'
            and  ind.INDICATOR_CODE  in ('43_covid_05','43_covid_07','43_covid_09_2')
            and len(org.MED_ORGANIZATION_NAME) > 5
            and  data.BEGIN_DATE in (cast(getdate()-2 as date),cast(getdate()-1 as date),cast(getdate()   as date))  
            and NUM_VALUE = '0' 
            group by [MED_ORGANIZATION_NAME],[BEGIN_DATE]
    """
    df = pd.read_sql(sql,con_parus)
    df = pd.pivot_table(df,index = ['организация'], values=['нулей в накопительных итогах'],columns=['Дата'])
    df = df.fillna(0)
    df['нулей в накопительных итогах'] = df['нулей в накопительных итогах'].round().astype(int)
    table_html = get_dir('temp') + '/table.html'
    table_png  = get_dir('temp') + '/table.png'

    df.to_html(table_html,justify='justify')
    subprocess.call('wkhtmltoimage --quiet --encoding utf-8 -f png --width 0 ' +  table_html + ' ' + table_png, shell=True)
    return table_png
