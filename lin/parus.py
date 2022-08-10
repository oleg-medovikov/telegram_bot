import sqlalchemy,cx_Oracle,os,openpyxl,shutil,datetime,subprocess,time,imgkit,requests
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import pandas as pd
from loader import get_dir
from sending import send

base_parus = os.getenv('base_parus')
userName = os.getenv('oracle_user')
password = os.getenv('oracle_pass')
userbase = os.getenv('oracle_base')

server  = os.getenv('server')
user    = os.getenv('mysqldomain') + '\\' + os.getenv('mysqluser') # Тут правильный двойной слеш!
passwd  = os.getenv('mypassword')
dbase   = os.getenv('db')

class my_except(Exception):
    pass

def o_40_covid_by_date(a):
    sql = open('sql/parus/covid_40_by_date.sql', 'r').read()    
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql,con)

    df = df.loc[df.TVSP.notnull()]
    df = df.loc[~df.TVSP.isin(['Пункт вакцинации'])]
    df.V_1 = pd.to_numeric(df.V_1)
    df.V_2 = pd.to_numeric(df.V_2)
    df.DAY = df.DAY.dt.date
    res = df.pivot_table(index=['DIST','TVSP'], columns=['DAY'], values=['V_1','V_2'],fill_value=0, aggfunc=np.sum)
    summ = df.pivot_table(index=['DIST'], columns=['DAY'], values=['V_1','V_2'],fill_value=0, aggfunc=np.sum)
    total = pd.DataFrame(summ.sum()).T

    res['Район'] = res.index.get_level_values(0)
    res['Пункт вакцинации'] = res.index.get_level_values(1)

    summ['Район'] = summ.index.get_level_values(0)
    summ['Пункт вакцинации'] = 'Всего'
    total['Район'] = 'Весь город'
    total['Пункт вакцинации'] = 'Все'

    itog= pd.concat([total,summ,res], ignore_index=True)
    itog = itog.set_index(['Район','Пункт вакцинации'])
    
    file = get_dir('temp') + '/Вакцинация по датам.xlsx'
    with pd.ExcelWriter(file) as writer:
        itog.V_1.to_excel(writer,sheet_name='Первый компонент вакцины')
        itog.V_2.to_excel(writer,sheet_name='Второй компонент вакцины')
    return file 

def svod_40_cov_19(a):
    sql   = open('sql/parus/covid_40_spytnic.sql','r').read()#.replace('SYSDATE', 'SYSDATE - 1')
    sql2  = open('sql/parus/covid_40_spytnic_old.sql','r').read()#.replace('SYSDATE', 'SYSDATE - 1')
    sql3  = open('sql/parus/covid_40_epivak.sql','r').read()#.replace('SYSDATE', 'SYSDATE - 1')
    sql4  = open('sql/parus/covid_40_epivak_old.sql','r').read()#.replace('SYSDATE', 'SYSDATE - 1')
    sql5  = open('sql/parus/covid_40_covivak.sql','r').read()#.replace('SYSDATE', 'SYSDATE - 1')
    sql6  = open('sql/parus/covid_40_covivak_old.sql','r').read()#.replace('SYSDATE', 'SYSDATE - 1')
    sql7  = open('sql/parus/covid_40_revac.sql','r').read()#.replace('SYSDATE', 'SYSDATE - 1')
    sql8  = open('sql/parus/covid_40_revac_old_new.sql','r').read()#.replace('SYSDATE', 'SYSDATE - 1')
    sql9  = open('sql/parus/covid_40_light.sql','r').read()#.replace('SYSDATE', 'SYSDATE - 1')
    sql10 = open('sql/parus/covid_40_light_old.sql','r').read()#.replace('SYSDATE', 'SYSDATE - 1')
    
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        sput = pd.read_sql(sql,con)
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        sput_old = pd.read_sql(sql2,con)
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        epivak = pd.read_sql(sql3,con)
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        epivak_old = pd.read_sql(sql4,con)
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        covivak = pd.read_sql(sql5,con)
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        covivak_old = pd.read_sql(sql6,con)
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        revac = pd.read_sql(sql7,con)
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        revac_old = pd.read_sql(sql8,con)
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        light = pd.read_sql(sql9,con)
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        light_old = pd.read_sql(sql10,con)
  
    send('', 'Запросы к базе выполнены')
    del sput ['ORGANIZATION']
    del sput_old ['ORGANIZATION']
    del epivak ['ORGANIZATION']
    del epivak_old ['ORGANIZATION']
    del covivak ['ORGANIZATION']
    del covivak_old ['ORGANIZATION']
    #del revac ['INDX']    
    #del revac_old ['INDX']
    del light ['ORGANIZATION']
    del light_old ['ORGANIZATION']
    revac = revac.loc[revac['TIP'] == 'Медицинская организация']
    revac_old = revac_old.loc[revac_old['TIP'] == 'Медицинская организация']

    date_otch = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%d.%m.%Y')
    new_name_pred ='40_COVID_19_БОТКИНА_' + date_otch + '_предварительный.xlsx'
    new_name_osn  ='40_COVID_19_БОТКИНА_' + date_otch + '_основной.xlsx'
    shablon_path = get_dir('help')

    shutil.copyfile(shablon_path + '/40_COVID_19_pred.xlsx' , shablon_path  + '/' + new_name_pred)
    shutil.copyfile(shablon_path + '/40_COVID_19_osn.xlsx'  , shablon_path  + '/' + new_name_osn)

    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name_pred)
    ws = wb['Спутник-V']
    rows = dataframe_to_rows(sput,index=False, header=False)
    for r_idx, row in enumerate(rows,5):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['Вчера_Спутник']
    rows = dataframe_to_rows(sput_old,index=False, header=False)
    for r_idx, row in enumerate(rows,5):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['ЭпиВакКорона']
    rows = dataframe_to_rows(epivak,index=False, header=False)
    for r_idx, row in enumerate(rows,5):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    ws = wb['Вчера_ЭпиВак']
    rows = dataframe_to_rows(epivak_old,index=False, header=False)
    for r_idx, row in enumerate(rows,5):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['КовиВак']
    rows = dataframe_to_rows(covivak,index=False, header=False)
    for r_idx, row in enumerate(rows,5):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['Спутник Лайт']
    rows = dataframe_to_rows(light,index=False, header=False)
    for r_idx, row in enumerate(rows,5):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)


    ws = wb['Вчера_КовиВак']
    rows = dataframe_to_rows(covivak_old,index=False, header=False)
    for r_idx, row in enumerate(rows,5):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['Ревакцинация']
    rows = dataframe_to_rows(revac,index=False, header=False)
    for r_idx, row in enumerate(rows,9):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['Вчера_ревакцин']
    rows = dataframe_to_rows(revac_old,index=False, header=False)
    for r_idx, row in enumerate(rows,9):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['Вчера_Спутник Лайт']
    rows = dataframe_to_rows(light_old,index=False, header=False)
    for r_idx, row in enumerate(rows,5):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save( shablon_path  + '/' + new_name_pred) 
    send('', 'Готов предварительный файл')

    # основной отчёт
    del sput[sput.columns[-1]]
    #del sput[sput.columns[-1]]
    #del sput[sput.columns[-1]]

    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name_osn)
    ws = wb['Спутник-V']
    rows = dataframe_to_rows(sput,index=False, header=False)
    for r_idx, row in enumerate(rows,5):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    del epivak[epivak.columns[-1]]
    #del epivak[epivak.columns[-1]]
    #del epivak[epivak.columns[-1]]
    
    ws = wb['ЭпиВакКорона']
    rows = dataframe_to_rows(epivak,index=False, header=False)
    for r_idx, row in enumerate(rows,5):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    del covivak[covivak.columns[-1]]
    #del covivak[covivak.columns[-1]]
    #del covivak[covivak.columns[-1]]

    ws = wb['КовиВак']
    rows = dataframe_to_rows(covivak,index=False, header=False)
    for r_idx, row in enumerate(rows,5):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    

    del light[light.columns[-1]]
    #del light[light.columns[-1]]
    #del light[light.columns[-1]]

    ws = wb['Спутник Лайт']
    rows = dataframe_to_rows(light,index=False, header=False)
    for r_idx, row in enumerate(rows,5):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    del revac['SCEP']
    ws = wb['Ревакцинация']
    rows = dataframe_to_rows(revac,index=False, header=False)
    for r_idx, row in enumerate(rows,9):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save( shablon_path  + '/' + new_name_osn) 

    return(shablon_path  + '/' + new_name_pred + ';' + shablon_path  + '/' + new_name_osn)

def svod_50_cov_19(a):
    sql1  = open('sql/parus/covid_50_polic.sql','r').read()
    sql2  = open('sql/parus/covid_50_stac.sql','r').read()
    sql3  = open('sql/covid/mz_50.sql','r').read()

    sql4  = open('sql/parus/covid_50_polic_old.sql','r').read()
    sql5  = open('sql/parus/covid_50_stac_old.sql','r').read()

    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        polic = pd.read_sql(sql1,con)
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        stac = pd.read_sql(sql2,con)
    with sqlalchemy.create_engine(f"mssql+pymssql://{user}:{passwd}@{server}/{dbase}",pool_pre_ping=True).connect() as con:
       covid = pd.read_sql(sql3,con)
    
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        polic_old = pd.read_sql(sql4,con)
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        stac_old = pd.read_sql(sql5,con)
    
    send('', 'Запросы к базе выполнены')
    date_otch = polic['DAY'].unique()[0]
    
    del polic ['DAY']
    del stac ['DAY']
    covid_pol = covid.loc[covid['Type_Therapy'] == 'Поликлинника']
    del covid_pol ['Type_Therapy']
    del covid_pol ['Count_70_COVID_week']
    del covid_pol ['Count_70_COVID']
    del covid_pol ['Count_70_Pnev_week']
    del covid_pol ['Count_70_Pnev']
    covid_stac = covid.loc[covid['Type_Therapy'] == 'Стационар']
    del covid_stac ['Type_Therapy']
    
    new_name_pred  ='50_COVID_19_' + date_otch + '_предварительный.xlsx'
    shablon_path = get_dir('help')

    shutil.copyfile(shablon_path + '/50_COVID_19_pred.xlsx' , shablon_path  + '/' + new_name_pred)

    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name_pred)
    ws = wb['Разрез МО_ГП']
    rows = dataframe_to_rows(polic,index=False, header=False)
    index_col = [2,9,10,11,12,13,14,15,16,17,18,19,20,21]
    for r_idx, row in enumerate(rows,7):  
        for c_idx, value in enumerate(row,0):
            ws.cell(row=r_idx, column=index_col[c_idx], value=value)
    
    ws = wb['Разрез МО_стац']
    rows = dataframe_to_rows(stac,index=False, header=False)
    index_col = [2,13,14,15,16,17,18,19,20,21,22,23,24]
    for r_idx, row in enumerate(rows,7):  
        for c_idx, value in enumerate(row,0):
            ws.cell(row=r_idx, column=index_col[c_idx], value=value)
    
    ws = wb['Пред.отч_разрез МО_ГП']
    rows = dataframe_to_rows(polic_old,index=False, header=False)
    index_col = [2,9,10,11,12,13,14,15,16,17,18,19,20,21]
    for r_idx, row in enumerate(rows,7):  
        for c_idx, value in enumerate(row,0):
            ws.cell(row=r_idx, column=index_col[c_idx], value=value)
    
    ws = wb['Пред.отч_разрез МО_стац']
    rows = dataframe_to_rows(stac_old,index=False, header=False)
    index_col = [2,13,14,15,16,17,18,19,20,21,22,23,24]
    for r_idx, row in enumerate(rows,7):  
        for c_idx, value in enumerate(row,0):
            ws.cell(row=r_idx, column=index_col[c_idx], value=value)
    
    ws = wb['ФР_ГП']
    rows = dataframe_to_rows(covid_pol,index=False, header=False)
    for r_idx, row in enumerate(rows,7):  
        for c_idx, value in enumerate(row,2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['ФР_стац']
    rows = dataframe_to_rows(covid_stac,index=False, header=False)
    for r_idx, row in enumerate(rows,7):  
        for c_idx, value in enumerate(row,2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save( shablon_path  + '/' + new_name_pred) 
    return shablon_path  + '/' + new_name_pred

def parus_43_cov_nulls(a):
    sql=open('sql/parus/covid_43_nulls.sql','r').read()
    
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql,con)

    df = df.fillna(0)
    table_html = get_dir('temp') + '/table.html'
    table_png  = get_dir('temp') + '/table.png'

    df.to_html(table_html,justify='center', index=False)

    command = "/usr/bin/wkhtmltoimage --encoding utf-8 -f png --width 0 " +  table_html + " " + table_png
    try:
        subprocess.check_call(command,  shell=True)
    except:
        raise my_except('Не удалось сделать файл\n' +  subprocess.check_output(command,  shell=True))
    else:
        return table_png

def svod_43_covid_19(a):
    sql = open('sql/parus/covid_43_svod.sql','r').read()
    if datetime.datetime.today().weekday() == 0:
        sql1 = open('sql/parus/covid_43_svod_old3.sql','r').read()
    else:
        sql1 = open('sql/parus/covid_43_svod_old1.sql','r').read()
    
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql,con)
        df_old = pd.read_sql(sql1,con)
    
    df.index     = range(1,len(df) + 1)
    df_old.index = range(1,len(df_old) + 1)

    date_otch = datetime.datetime.now().strftime('%d_%m_%Y')
    new_name = date_otch + '_43_COVID_19_cvod.xlsx'
    shablon_path = get_dir('help')
    shutil.copyfile(shablon_path + '/43 COVID 19.xlsx', shablon_path  + '/' + new_name)

    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    ws = wb['Разрез по МО']
    rows = dataframe_to_rows(df,index=True, header=False)
    for r_idx, row in enumerate(rows,4):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save( shablon_path  + '/' + new_name) 
    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    ws = wb['Вчера']
    rows = dataframe_to_rows(df_old,index=True, header=False)
    for r_idx, row in enumerate(rows,4):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save( shablon_path  + '/' + new_name)
    return  shablon_path  + '/' + new_name

def no_save_43(a):
    sql= open('sql/parus/covid_43_no_save.sql','r').read()
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql,con)

    df = df.fillna(0)
    table_html = get_dir('temp') + '/table.html'
    table_png  = get_dir('temp') + '/table.png'

    df.to_html(table_html,justify='center', index=False)
    command = "/usr/bin/wkhtmltoimage --encoding utf-8 -f png --width 0 " +  table_html + " " + table_png
    try:
        subprocess.check_call(command,  shell=True)
    except:
        raise my_except('Не удалось сделать файл\n' +  subprocess.check_output(command,  shell=True))
    else:
        return table_png

def no_save_50(a):
    sql= open('sql/parus/covid_50_no_save.sql','r').read()
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql,con)

    df = df.fillna(0)
    table_html = get_dir('temp') + '/table.html'
    table_png  = get_dir('temp') + '/table.png'

    df.to_html(table_html,justify='center', index=False)
    command = "/usr/bin/wkhtmltoimage --encoding utf-8 -f png --width 0 " +  table_html + " " + table_png
    try:
        subprocess.check_call(command,  shell=True)
    except:
        raise my_except('Не удалось сделать файл\n' +  subprocess.check_output(command,  shell=True))
    else:
        return table_png

def cvod_26_covid(a):
    sql = open('sql/parus/covid_26_svod.sql','r').read()
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql,con)
    df ['type'] = 'parus' 
    old_file = get_dir('punct_zabor') +'/'+  datetime.datetime.now().strftime('%d.%m.%Y') + ' Пункты отбора.xlsx'
    values = {6 : 0, 7 : 0, 8 : 0, 9 : 0, 10 : 0, 11 : 0, 12 : 0, 13 : 0, 14 : 0 }
    try:
        old = pd.read_excel(old_file,skiprows=3,header=None,sheet_name='Соединение')
    except:
        old = pd.DataFrame()
    else:
        old = old.loc[~(old[2].isnull() & old[3].isnull() & old[5].isnull() ) ].fillna(value=values)
        old [7]  = 0 
        old [9]  = 0
        old [11] = 0
        old [13] = 0
        del old [0]
        del old [14]
        old ['type'] = 'file' 
    if len(old.columns) == len(df.columns):    
        old.columns = df.columns
    

    old_file = get_dir('punct_zabor') +'/'+  datetime.datetime.now().strftime('%d.%m.%Y') + ' Пункты отбора.xlsx'
    new_df = pd.concat([df,old], ignore_index=True).drop_duplicates(subset=['LAB_UTR_MO', 'ADDR_PZ', 'LAB_UTR_02'])
    date = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%d_%m_%Y')
    new_name = date + '_26_COVID_19_cvod.xlsx'
    shablon_path = get_dir('help')

    shutil.copyfile(shablon_path + '/26_COVID_19_svod.xlsx', shablon_path  + '/' + new_name)

    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    ws = wb['Из паруса']
    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,5):  
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)
    ws = wb['Из файла']
    rows = dataframe_to_rows(old,index=False, header=False)
    for r_idx, row in enumerate(rows,5):  
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)
    ws = wb['Соединение']
    rows = dataframe_to_rows(new_df,index=False, header=False)
    for r_idx, row in enumerate(rows,5):  
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save( shablon_path  + '/' + new_name)
    return  shablon_path  + '/' + new_name

def cvod_27_smal(a):
    sql = open('sql/parus/covid_27_smal.sql','r').read()

    #send('', 'начинаю грузить для плотниковой')
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql,con)
    if len(df):
        with sqlalchemy.create_engine(f"mssql+pymssql://{user}:{passwd}@miacbase3/MIAC_DS", pool_pre_ping=True).connect() as c:
            df.to_sql('covid_27',c,schema='Pds',index=False,if_exists='replace')

    #send('', 'Для плотниковой загрузил')
    
    return 1

def cvod_27_covid(a):
    sql1 = open('sql/parus/covid_27_svod.sql','r').read()
    sql2 = open('sql/parus/covid_27_svod_old.sql','r').read()
    sql3 = open('sql/parus/covid_27_svod_2.sql','r').read()
    sql4 = open('sql/parus/covid_27_svod_3.sql','r').read()
    sql5 = open('sql/parus/covid_27_svod_4.sql','r').read()
    
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql1,con)

    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df_old = pd.read_sql(sql2,con)
    
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df_2 = pd.read_sql(sql3,con)
     
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df_3 = pd.read_sql(sql4,con)

    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df_4 = pd.read_sql(sql5,con)

    date = datetime.datetime.now().strftime('%d_%m_%Y')
     
    df_3 = df_3.loc[~(df_3['IDMO'].isnull())]
    for col  in df_3.columns[10:]:
        try:
            df_3[col] = df_3[col].str.replace(',','.').astype(float)
        except:
            send('',col)

   
    df_4 = df_4.loc[~(df_4['IDMO'].isnull())]
    for col  in df_4.columns[10:]:
        try:
            df_4[col] = df_4[col].str.replace(',','.').astype(float)
        except:
            send('',col)

    shablon_path = get_dir('help')

    new_name_1 = shablon_path +'/'+ date + '_1_27_COVID-19.xlsx'
    new_name_2 = shablon_path +'/'+ date + '_2_Результаты_исследований_материала_на_COVID-19.xlsx'
    new_name_3 = shablon_path +'/'+ date + '_3_Кратность_лаб_обсл_на_COVID-19.xlsx'
    new_name_4 = shablon_path +'/'+ date + '_4_Кратность_положительных_лаб_обсл_на_COVID-19.xlsx'

    shutil.copyfile(shablon_path + '/27_COVID_19_svod.xlsx', new_name_1)
        
    wb= openpyxl.load_workbook( new_name_1)
    
    ws = wb['Для заполнения']
    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,4):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    ws = wb['Пред.отч']
    rows = dataframe_to_rows(df_old,index=False, header=False)
    for r_idx, row in enumerate(rows,4):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save( new_name_1 )

    shutil.copyfile(shablon_path + '/27_COVID_19_svod_2.xlsx', new_name_2)
        
    wb= openpyxl.load_workbook( new_name_2)
    
    ws = wb['свод']
    rows = dataframe_to_rows(df_2,index=False, header=False)
    for r_idx, row in enumerate(rows,4):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save( new_name_2 )

    shutil.copyfile(shablon_path + '/27_COVID_19_svod_3.xlsx', new_name_3)
        
    wb= openpyxl.load_workbook( new_name_3)
    
    ws = wb['свод']
    rows = dataframe_to_rows(df_3,index=False, header=False)
    for r_idx, row in enumerate(rows,4):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save( new_name_3 )
     
    shutil.copyfile(shablon_path + '/27_COVID_19_svod_4.xlsx', new_name_4)
    
    wb= openpyxl.load_workbook( new_name_4)
    
    ws = wb['свод']
    rows = dataframe_to_rows(df_4,index=False, header=False)
    for r_idx, row in enumerate(rows,4):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save( new_name_4 )
    return ( new_name_1 + ';' + new_name_2 +';'+ new_name_3 +';'+ new_name_4 )

def cvod_27_regiz(a):
    url = os.getenv('url845').replace('845','870')
    data = requests.get(url, verify=False).json()
    regiz = pd.DataFrame.from_dict(data)
    columns = ['orderresponse_assign_organization_level1_key', 'ShortNameMO','Кол-во тестов',
               'Кол-во ПЦР тестов','Кол-во положительных ПЦР тестов','Кол-во тестов на антитела',
               'Кол-во положительных тестов на антитела','Кол-во тестов на антитела после вакцинации',
               'Кол-во положительных тестов на антитела после вакцинации']
    regiz = regiz[columns] 
    sql = open('sql/parus/covid_27_regiz.sql','r').read()
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql,con)
    
    sql = open('sql/covid/nsi_27.sql','r').read()
    with sqlalchemy.create_engine(f"mssql+pymssql://{user}:{passwd}@miacbase3/NsiBase",pool_pre_ping=True).connect() as con:
       nsi = pd.read_sql(sql,con)
 
    date = (datetime.datetime.now() - datetime.timedelta(days=2)).strftime('%d_%m_%Y')
    new_name = date + '_27_COVID_19_regiz.xlsx'
    shablon_path = get_dir('help')
    shutil.copyfile(shablon_path + '/27_COVID_19_regiz.xlsx', shablon_path  + '/' + new_name)
        
    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    
    ws = wb['parus']
    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,4):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    

    ws = wb['regiz']
    rows = dataframe_to_rows(regiz,index=False, header=True)
    for r_idx, row in enumerate(rows,2):  
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['nsi']
    rows = dataframe_to_rows(nsi,index=False, header=True)
    for r_idx, row in enumerate(rows,1):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save( shablon_path  + '/' + new_name)
    return  shablon_path  + '/' + new_name

def cvod_29_covid(a): 
    if int(time.strftime("%H")) < 16:
        sql = open('sql/parus/covid_29_svod1.sql','r').read()
        date = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%d_%m_%Y')
    else:
        sql = open('sql/parus/covid_29_svod0.sql','r').read()
        date = datetime.datetime.now().strftime('%d_%m_%Y')
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql,con)

    new_name = date + '_29_COVID_19_cvod.xlsx'
    shablon_path = get_dir('help')

    shutil.copyfile(shablon_path + '/29_COVID_19_svod.xlsx', shablon_path  + '/' + new_name)

    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    ws = wb['Для заполнения']
    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,3):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save( shablon_path  + '/' + new_name)
    return  shablon_path  + '/' + new_name

def cvod_33_covid(a):
    sql = open('sql/parus/covid_33_svod.sql','r').read()
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql,con)

    date = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%d_%m_%Y')
    new_name = date + '_33_COVID_19_cvod.xlsx'
    shablon_path = get_dir('help')

    shutil.copyfile(shablon_path + '/33_COVID_19_svod.xlsx', shablon_path  + '/' + new_name)

    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    ws = wb['Свод']
    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,5):  
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save( shablon_path  + '/' + new_name)
    return  shablon_path  + '/' + new_name

def cvod_36_covid(a):
    sql = open('sql/parus/covid_36_svod.sql','r').read()
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql,con)

    date = df['DAY'].unique()[0]
    del df['DAY']
    new_name = date + '_36_COVID_19_cvod.xlsx'
    shablon_path = get_dir('help')

    shutil.copyfile(shablon_path + '/36_COVID_19_svod.xlsx', shablon_path  + '/' + new_name)

    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    ws = wb['Свод']
    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,7):  
        for c_idx, value in enumerate(row, 2):
             ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save( shablon_path  + '/' + new_name)
    return  shablon_path  + '/' + new_name

def cvod_37_covid(a):
    sql = open('sql/parus/covid_37_svod.sql','r').read()
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql,con)

    date = df['DAY'].unique()[0]
    del df['DAY']
    new_name = date + '_37_COVID_19_cvod.xlsx'
    shablon_path = get_dir('help')

    shutil.copyfile(shablon_path + '/37_COVID_19_svod.xlsx', shablon_path  + '/' + new_name)

    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    ws = wb['Свод']
    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,6):  
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save( shablon_path  + '/' + new_name)
    return  shablon_path  + '/' + new_name

def cvod_38_covid(a):
    sql = open('sql/parus/covid_37_svod.sql','r').read()
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql,con)

    date = df['DAY'].unique()[0]
    del df['DAY']
    new_name = date + '_38_COVID_19_cvod.xlsx'
    shablon_path = get_dir('help')

    shutil.copyfile(shablon_path + '/38_COVID_19_svod.xlsx', shablon_path  + '/' + new_name)

    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    ws = wb['Свод']
    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,8):  
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save( shablon_path  + '/' + new_name)
    return  shablon_path  + '/' + new_name

def cvod_51_covid(a):
    sql1 = open('sql/parus/covid_51_svod.sql','r').read()
    sql2 = open('sql/parus/covid_51_svod_all.sql','r').read()

    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql1,con)

    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df_all = pd.read_sql(sql2,con)


    date = df['DAY'].unique()[0]
    del df ['DAY']
    df = df.append(df.sum(numeric_only=True), ignore_index=True)
    df.loc[len(df)-1,'COV_02'] = 'ИТОГО:'
    new_name = date + '_51_COVID_19_cvod.xlsx'
    shablon_path = get_dir('help')

    shutil.copyfile(shablon_path + '/51_COVID_19_svod.xlsx', shablon_path  + '/' + new_name)

    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    ws = wb['Свод по МО']
    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,4):  
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)
    ws = wb['Свод по всем МО']
    rows = dataframe_to_rows(df_all,index=False, header=False)
    for r_idx, row in enumerate(rows,4):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save( shablon_path  + '/' + new_name)

    return shablon_path  + '/' + new_name

def cvod_52_covid(a):
    sql1= open('sql/parus/covid_52_svod.sql','r').read()
    sql2= open('sql/parus/covid_52_svod_old.sql','r').read()


    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql1,con)

    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df_old = pd.read_sql(sql2,con)
        
    date = str(df['DAY'].unique()[0])
    del df ['ORGANIZATION'] 
    del df_old ['ORGANIZATION'] 
    del df ['DAY']
    del df_old ['DAY']
    new_name = date + '_52_COVID_19_pred.xlsx'
    new_name_2 = date + '_52_COVID_19_osn.xlsx'
    shablon_path = get_dir('help')

    shutil.copyfile(shablon_path + '/52_COVID_19_pred.xlsx', shablon_path  + '/' + new_name)
    shutil.copyfile(shablon_path + '/52_COVID_19_osn.xlsx', shablon_path  + '/' + new_name_2)

    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    ws = wb['52 COVID']
    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,11):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    ws = wb['Вчера']
    rows = dataframe_to_rows(df_old,index=False, header=False)
    for r_idx, row in enumerate(rows,11):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save( shablon_path  + '/' + new_name)


    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name_2)
    ws = wb['52 COVID']
    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,11):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save( shablon_path  + '/' + new_name_2)

    return shablon_path  + '/' + new_name +';'+ shablon_path  + '/' + new_name_2

def cvod_28_covid(a):
    sql= open('sql/parus/covid_28_svod.sql','r').read()

    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql,con)
 
    date = str(df['DAY'].unique()[0])
    del df ['DAY']

    new_name = date + '_28_COVID_19.xlsx'
    shablon_path = get_dir('help')

    shutil.copyfile(shablon_path + '/28_COVID_19_svod.xlsx', shablon_path  + '/' + new_name)
 
    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    ws = wb['Свод']
    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,4):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save( shablon_path  + '/' + new_name)
    
    return shablon_path  + '/' + new_name

def cvod_41_covid(a):
    sql= open('sql/parus/covid_41_svod.sql','r').read()

    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql,con)
 
    date = str(df['DAY'].unique()[0])
    del df ['DAY']
    del df ['ORGANIZATION']

    new_name = date + '_41_COVID_19.xlsx'
    shablon_path = get_dir('help')

    shutil.copyfile(shablon_path + '/41_COVID_19_svod.xlsx', shablon_path  + '/' + new_name)
 
    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    ws = wb['Свод']
    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,4):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save( shablon_path  + '/' + new_name)
    
    return shablon_path  + '/' + new_name

def cvod_42_covid(a):
    sql= open('sql/parus/covid_42_svod.sql','r').read()

    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql,con)
 
    date = str(df['DAY'].unique()[0])
    del df ['DAY']
    del df ['ORGANIZATION']

    new_name = date + '_42_COVID_19.xlsx'
    shablon_path = get_dir('help')

    shutil.copyfile(shablon_path + '/42_COVID_19_svod.xlsx', shablon_path  + '/' + new_name)
 
    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    ws = wb['Свод']
    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,4):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save( shablon_path  + '/' + new_name)
    
    return shablon_path  + '/' + new_name

def cvod_49_covid(a):
    sql1= open('sql/parus/covid_49_svod.sql','r').read()
    #sql2= open('sql/parus/covid_49_dolg.sql','r').read()

    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql1,con)
    
    #with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
    #    dolg = pd.read_sql(sql2,con)
    
    spisok_mo = pd.read_excel( get_dir('help')+ '/svod_49_spisok_mo.xlsx')

    for i in range(len(spisok_mo)):
        if spisok_mo.at[i, 'медицинская организация'] in df['ORGANIZATION'].unique():
            spisok_mo = spisok_mo.drop([i])
 
    date = str(df['DAY'].unique()[0])
    del df ['DAY']
    del df ['ORGANIZATION']

    new_name = date + '_49_COVID_19.xlsx'
    shablon_path = get_dir('help')
    

    shutil.copyfile(shablon_path + '/49_COVID_19_svod.xlsx', shablon_path  + '/' + new_name)
 
    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    ws = wb['свод']
    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,5):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    rows = dataframe_to_rows(spisok_mo,index=False, header=False)
    for r_idx, row in enumerate(rows,5):
        for c_idx, value in enumerate(row, 9):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save( shablon_path  + '/' + new_name)
    
    return shablon_path  + '/' + new_name



def cvod_4_3_covid(a):
    medications = open('sql/lists/medications','r').read().split('\n')
    poks = ['_01', '_02', '_03', '_04', '_05', '_06', '_07', '_08', '_09',
            '_10', '_11', '_12', '_13', '_14', '_15', '_16', '_17', '_18',
            '_19', '_20', '_21', '_22', '_23', '_24', '_25', '_26', '_27',
            '_28', '_29', '_30', '_31', '_32', '_33', '_34', '_35', '_36',
            '_37', '_38', '_39', '_40', '_41', '_42', '_43', '_44', '_45',
            '_46', '_47', '_48', '_49', '_50', '_51', '_52', '_53', '_54',
            '_55', '_56', '_57', '_58', '_59']

    shablon =  get_dir('help') + '/4.3_COVID_19_svod.xlsx'
    file = get_dir('help') + '/4.3 COVID 19.xlsx'

    shutil.copyfile(shablon,file)
    wb= openpyxl.load_workbook(file,data_only=False)
    
    for med,pok in zip(medications,poks):
        pokazateli = f"""'4.3_1{pok}' pok1,'4.3_2{pok}' pok2,'4.3_3{pok}' pok3,'4.3_4{pok}' pok4,'4.3_5{pok}' pok5,'4.3_6{pok}' pok6,'4.3_7{pok}' pok7,'4.3_8{pok}' pok8"""
        sql = open('/home/sshuser/telegram_bot/lin/sql/parus/covid_4.3_svod_stac.sql','r').read().replace('pokazateli',pokazateli)

        with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
            df = pd.read_sql(sql,con)
        del df['DAY']
    
        df['POK9'] = 30*(df['POK1'] + df['POK3'])/ df['POK4'] 
        df['POK10'] = (df['POK1'] + df['POK3'])/ df['POK4'] 
        ws = wb[pok.replace('_','')]
        
        # пишем наименование лекарства
        ws.cell(row=3, column=6, value=med)
        # пишем значения в таблицу
        rows = dataframe_to_rows(df,index=False, header=False)
        for r_idx, row in enumerate(rows,5):  
            for c_idx, value in enumerate(row, 5):
                ws.cell(row=r_idx, column=c_idx, value=value)

        ws = wb['main']
        ws.cell(row= 6  + medications.index(med), column=5, value=med)
        
        ws.cell(row= 6  + medications.index(med), column=6, value=df.loc[~(df['POK1'].isnull()), 'POK1'].sum())
        ws.cell(row= 6  + medications.index(med), column=7, value=df.loc[~(df['POK3'].isnull()), 'POK3'].sum())
        ws.cell(row= 6  + medications.index(med), column=8, value=df.loc[~(df['POK4'].isnull()), 'POK4'].sum())
        ws.cell(row= 6  + medications.index(med), column=9, value=df.loc[~(df['POK5'].isnull()), 'POK5'].sum())


        ws.cell(row= 6  + medications.index(med), column=10, value=df.loc[~(df['POK9'].isnull()) & (df['POK9'] != np.inf) , 'POK9'].mean())
        ws.cell(row= 6  + medications.index(med), column=11, value=df.loc[~(df['POK10'].isnull()) & (df['POK10'] != np.inf) , 'POK10'].mean())
        if len(df.loc[(df['POK9'].isnull()) | (df['POK9'] == np.inf) , 'POK9']):
            ws.cell(row= 6  + medications.index(med), column=12, value='Да')
        else:
            ws.cell(row= 6  + medications.index(med), column=12, value='Нет')

    wb.save(file)

    return file

def medical_waste(a):
    sql= open('sql/parus/medical_waste.sql','r').read()

    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql,con)

    for col in df.columns:
        try:
            df[col] = pd.to_numeric(df[col], errors='ignore')
        except:
            pass
    new_name = datetime.datetime.now().strftime('%d_%m_%Y') + '_медицинские_отходы.xlsx'
    shablon_path = get_dir('help')

    shutil.copyfile(shablon_path + '/medical_waste.xlsx', shablon_path  + '/' + new_name)
 
    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    ws = wb['SVOD']
    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,5):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    dolg = set(df.loc[df['YEAR'] == datetime.datetime.now().year -1 , 'ORGANIZATION' ].unique()) \
            -  set(df.loc[df['YEAR'] == datetime.datetime.now().year, 'ORGANIZATION' ].unique())
    
    dolg = pd.DataFrame(dolg)    

    ws = wb['DOLG']
    rows = dataframe_to_rows(dolg,index=False, header=False)
    for r_idx, row in enumerate(rows,2):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save( shablon_path  + '/' + new_name)
    
    return shablon_path  + '/' + new_name

def covid_53_svod(a):
    sql1 = open('sql/parus/covid_53_svod.sql', 'r').read()
    sql2 = open('sql/parus/covid_53_svod_old.sql', 'r').read()

    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql1,con)
    
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        old = pd.read_sql(sql2,con)

    date = df['DAY'].unique()[0]
    del df['DAY']

    sm = df.groupby(by="ORGANIZATION",as_index=False).sum()
    sm['POK03'] = sm ['ORGANIZATION']
    sm ['TYPE'] = 'Медицинская организация'
    
    for i in range(len(sm)):
        k = len(df)
        for col in df.columns:
            try:
                df.loc[k,col] = sm.at[i,col]
            except:
                pass

    df = df.sort_values(by=["ORGANIZATION", "POK02"],na_position='first',ignore_index=True).fillna('')
    
    del df ['ORGANIZATION']
    # вчерашнее 
    del old['DAY']

    sm = old.groupby(by="ORGANIZATION",as_index=False).sum()
    sm['POK03'] = sm ['ORGANIZATION']
    sm ['TYPE'] = 'Медицинская организация'
    
    for i in range(len(sm)):
        k = len(old)
        for col in old.columns:
            try:
                old.loc[k,col] = sm.at[i,col]
            except:
                pass

    old = old.sort_values(by=["ORGANIZATION", "POK02"],na_position='first',ignore_index=True).fillna('')
    
    del old ['ORGANIZATION']
    # ====


    new_name = '53_COVID_БОТКИНА_' + date + '.xlsx'
    new_name2 = '53_COVID_БОТКИНА_' + date + '_основной.xlsx'

    shablon_path = get_dir('help')
    shutil.copyfile(shablon_path + '/53_COVID_19_svod.xlsx', shablon_path  + '/' + new_name)
    shutil.copyfile(shablon_path + '/53_COVID_19.xlsx', shablon_path  + '/' + new_name2)

    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    ws = wb['Спутник-М']

    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,5):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['Вчера_Спутник-М']

    rows = dataframe_to_rows(old,index=False, header=False)
    for r_idx, row in enumerate(rows,5):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save( shablon_path  + '/' + new_name)

    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name2)
    ws = wb['Спутник-М']

    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,5):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save( shablon_path  + '/' + new_name2)

    return shablon_path  + '/' + new_name +';' + shablon_path  + '/' + new_name2

 
def covid_54_svod(a):
    sql1 = open('sql/parus/covid_54_svod.sql', 'r').read()
    sql2 = open('sql/parus/covid_54_error.sql', 'r').read()

    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql1,con)

    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        error = pd.read_sql(sql2,con)

    del error['SUM(OSTATOK)']

    date = df['DAY'].unique()[0]
    del df['DAY']

    new_name = '54_COVID_19_' + date + '.xlsx'

    shablon_path = get_dir('help')
    shutil.copyfile(shablon_path + '/54_COVID_19_svod.xlsx', shablon_path  + '/' + new_name)

    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    
    ws = wb['svod']

    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,4):
        for c_idx, value in enumerate(row, 3):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['ошибки с остатком']

    rows = dataframe_to_rows(error,index=False, header=False)
    for r_idx, row in enumerate(rows,2):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)


    wb.save( shablon_path  + '/' + new_name)

    return shablon_path + '/' + new_name
  
def extra_izv(a):
    sql1 = open('sql/parus/extra_izv.sql', 'r').read()
    sql2 = open('sql/parus/extra_izv_dolg.sql', 'r').read()

    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql1,con)
    
    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        old = pd.read_sql(sql2,con)

    date = df['DAY'].unique()[0]
    del df['DAY']

    dolg = pd.DataFrame(columns=["ORGANIZATION"])

    for org in old["ORGANIZATION"].unique():
        if not org in df["POK01"].unique():
            dolg.loc[len(dolg), "ORGANIZATION"] = org


    new_name = 'Экстренные_извещения_' + date + '.xlsx'

    shablon_path = get_dir('help')
    shutil.copyfile(shablon_path + '/extra_izv.xlsx', shablon_path  + '/' + new_name)

    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    
    ws = wb['svod']

    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,3):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    rows = dataframe_to_rows(dolg,index=False, header=False)
    for r_idx, row in enumerate(rows,3):
        for c_idx, value in enumerate(row, 6):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save( shablon_path  + '/' + new_name)

    return shablon_path + '/' + new_name

def distant_consult(a):
    sql1 = open('sql/parus/distant_svod.sql', 'r').read()
    #sql2 = open('sql/parus/distant_organization.sql', 'r').read()

    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql1,con)
    
    #with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
    #    old = pd.read_sql(sql2,con)

    date = df['DAY'].unique()[0]
    del df['DAY']
    
    new_name = 'Дистанц_Консультации_' + date + '.xlsx'

    shablon_path = get_dir('help')
    shutil.copyfile(shablon_path + '/distant_consult.xlsx', shablon_path  + '/' + new_name)
    
    old = pd.read_excel(shablon_path + '/distant_consult.xlsx', sheet_name = 'эталон')
    
    dolg = pd.DataFrame(columns=["ORGANIZATION"])

    for i in range(len(old)):
        org = old.at[i, 'Полное наименование МО (из ФРМО)']
        if not org in df["POK02"].unique():
            dolg.loc[len(dolg), "ORGANIZATION"] = old.at[i,'Краткое наименование МО (для вывода должников)']
    
    del df ['ORGANIZATION']

    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    
    ws = wb['svod']

    rows = dataframe_to_rows(df,index=False, header=False)
    index_col = [1,2,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38]
    for r_idx, row in enumerate(rows,2):
        for c_idx, value in enumerate(row, 0):
            ws.cell(row=r_idx, column=index_col[c_idx], value=value)

    ws = wb['dolg']
    
    rows = dataframe_to_rows(dolg,index=False, header=False)
    for r_idx, row in enumerate(rows,3):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save( shablon_path  + '/' + new_name)
    
    csv = pd.read_csv(shablon_path + '/dist_cons.csv', sep=';' )
    csv = csv.drop(0) 
    
    k = 0
    for i in index_col:
        try:
            for index in range(len(df)):
                csv.loc[index, csv.columns[i - 1]] = str(df.at[index, df.columns[k]]).replace('.0','')
        except:
            break
        else:
            k+=1

    csv_file = 'temp/шаблон_для_загрузки.csv'
    
    csv.to_csv(csv_file, sep=';', index=False, encoding='cp1251')

    return shablon_path + '/' + new_name +';'+ csv_file
    
def covid_4_2_svod(a):
    sql = open('sql/parus/covid_4.2_svod.sql', 'r').read()

    with cx_Oracle.connect(userName, password, userbase,encoding="UTF-8") as con:
        df = pd.read_sql(sql,con)

    date = df['DAY'].unique()[0]
    del df['DAY']

    new_name = '4.2_COVID_19_' + date + '.xlsx'

    shablon_path = get_dir('help')
    shutil.copyfile(shablon_path + '/4.2_COVID_19_svod.xlsx', shablon_path  + '/' + new_name)

    names = [
            'region','ORGANIZATION','FULLNAME','MIAC_COVID4.2_vrach','MIAC_COVID4.2_telefo',
            'MIAC_COVID4.2_gl','MIAC_COVID4.2_kl','MIAC_COVID4.2_rez','MIAC_COVID4.2_tipmo','MIAC_COVID4.2spm_vs',
            'MIAC_COVID4.2spm_ivs','MIAC_COVID4.2spm_11','MIAC_COVID4.2spm_12','MIAC_COVID4.2spm_13','MIAC_COVID4.2spm_14',
            'MIAC_COVID4.2spm_15','MIAC_COVID4.2spm_16','voditel','MIAC_COVID4.2spm_18','MIAC_COVID4.2pol_101',
            'MIAC_COVID4.2pol_103','MIAC_COVID4.2pol_102','MIAC_COVID4.2pol_104','MIAC_COVID4.2pol_105',
            'MIAC_COVID4.2pol_106','MIAC_COVID4.2pol_107','MIAC_COVID4.2pol_108','MIAC_COVID4.2pol_109',
            'MIAC_COVID4.2pol_110','MIAC_COVID4.2pol_111','MIAC_COVID4.2pol_112','MIAC_COVID4.2pol_113',
            'MIAC_COVID4.2pol_114','MIAC_COVID4.2pol_115','MIAC_COVID4.2pol_116','MIAC_COVID4.2_129','MIAC_COVID4.2_101',
            'MIAC_COVID4.2_102','MIAC_COVID4.2_103','MIAC_COVID4.2_104','MIAC_COVID4.2_105','MIAC_COVID4.2_106',
            'MIAC_COVID4.2_107','MIAC_COVID4.2_108','MIAC_COVID4.2_109','MIAC_COVID4.2_110','MIAC_COVID4.2_111',
            'MIAC_COVID4.2_112','MIAC_COVID4.2_113','MIAC_COVID4.2_114','MIAC_COVID4.2_115','MIAC_COVID4.2_116',
            'MIAC_COVID4.2_117','MIAC_COVID4.2_118','MIAC_COVID4.2_119','MIAC_COVID4.2_120','MIAC_COVID4.2_121',
            'MIAC_COVID4.2_122','MIAC_COVID4.2_123','MIAC_COVID4.2_124','MIAC_COVID4.2_125','MIAC_COVID4.2_126',
            'MIAC_COVID4.2_127','MIAC_COVID4.2_128','MIAC_COVID4.2_229','MIAC_COVID4.2_201','MIAC_COVID4.2_202',
            'MIAC_COVID4.2_203','MIAC_COVID4.2_204','MIAC_COVID4.2_205','MIAC_COVID4.2_206','MIAC_COVID4.2_207',
            'MIAC_COVID4.2_208','MIAC_COVID4.2_209','MIAC_COVID4.2_210','MIAC_COVID4.2_211','MIAC_COVID4.2_212',
            'MIAC_COVID4.2_213','MIAC_COVID4.2_214','MIAC_COVID4.2_215','MIAC_COVID4.2_216','MIAC_COVID4.2_217',
            'MIAC_COVID4.2_218','MIAC_COVID4.2_219','MIAC_COVID4.2_220','MIAC_COVID4.2_221','MIAC_COVID4.2_222',
            'MIAC_COVID4.2_223','MIAC_COVID4.2_224','MIAC_COVID4.2_225','MIAC_COVID4.2_226','MIAC_COVID4.2_227',
            'MIAC_COVID4.2_228','COVID4.2pol_109.1','COVID4.2pol_109.2','COVID4.2pol_109.3','COVID4.2pol_112.1',
            'COVID4.2pol_112.2','COVID4.2pol_116.1','COVID4.2spm_16.1','COVID4.2spm_16.2','COVID4.2spm_16.3',
            'COVID4.2spm_16.4','COVID4.2spm_16.5','COVID4.2spm_16.6','MIAC_COVID4.2_120.1','MIAC_COVID4.2_220.1','felsheri'
            ]
    

    ot = pd.DataFrame(columns=names)

    ot['ORGANIZATION'] = pd.Series(df['ORGANIZATION'].unique())
    ot['FULLNAME'] = ot['ORGANIZATION']

    for i in range(len(df)):
        if df.at[i,'POKAZATEL'] in names:
            ot.loc[ot['ORGANIZATION'] == df.at[i,'ORGANIZATION'], df.at[i,'POKAZATEL'] ] = df.at[i,'VALUE']

    ot['region'] = 'г. Санкт-Петербург'
    ot['MIAC_COVID4.2_vrach'] = ot['MIAC_COVID4.2_vrach'].str.replace('\n', '').str.replace('\r', '').str.replace('\t', ' ')
    ot['MIAC_COVID4.2_telefo'] = ot['MIAC_COVID4.2_telefo'].str.replace('\n', '').str.replace('\r', '').str.replace('\t', ' ')

    ot = ot.fillna(0)

    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)

    ws = wb['svod']

    rows = dataframe_to_rows(ot,index=False, header=False)

    for r_idx, row in enumerate(rows,2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save( shablon_path  + '/' + new_name)

    return shablon_path + '/' + new_name
 

