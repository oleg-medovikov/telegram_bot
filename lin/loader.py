#Работа с базами данных
import pandas as pd
import subprocess,glob,warnings,datetime,os,xlrd,csv,openpyxl,shutil,threading,time,numpy,sqlalchemy
from sending import send
from sqlalchemy.orm import sessionmaker
from openpyxl.utils.dataframe import dataframe_to_rows
from reports import short_report
from multiprocessing import Process
from multiprocessing.pool import ThreadPool
from reports import short_report
#from pyexcelerate import Workbook

warnings.filterwarnings('ignore')

class my_except(Exception):
     pass

server  = os.getenv('server')
user    = os.getenv('mysqldomain') + '\\' + os.getenv('mysqluser') # Тут правильный двойной слеш!
passwd  = os.getenv('mypassword')
dbase   = os.getenv('db')

eng = sqlalchemy.create_engine(f"mssql+pymssql://{user}:{passwd}@{server}/{dbase}",pool_pre_ping=True)
con = eng.connect()

def get_dir(name):
    sql = f"SELECT Directory FROM [robo].[directions_for_bot] where NameDir = '{name}' and [linux] = 'True'"
    with sqlalchemy.create_engine(f"mssql+pymssql://{user}:{passwd}@{server}/{dbase}",pool_pre_ping=True).connect() as con:
        df = pd.read_sql(sql,con)
    return df.iloc[0,0] 

def check_table(name):
    sql=f"""SELECT distinct case when (cast([datecreated] as date) = cast(getdate() as date))
        then 1
        else 0
        end AS 'Check'
            FROM [dbo].[cv_{name}] 
            where cast([datecreated] as date) > '20210812'
            """

    with sqlalchemy.create_engine(f"mssql+pymssql://{user}:{passwd}@{server}/{dbase}",pool_pre_ping=True).connect() as con:
        df = pd.read_sql(sql,con).iat[0,0]
    return df 

def search_file(category):
    path = get_dir('path_robot') + '/' + datetime.datetime.now().strftime("%Y_%m_%d")
    if category == 'fr':
        pattern = path + '/Федеральный регистр лиц, больных *[!ИАЦ].xlsx'
    if category == 'fr_death':
        pattern = path + '/*Умершие пациенты*.xlsx'
    if category == 'fr_lab':
        pattern = path + '/*Отчёт по лабораторным*.xlsx'
    if category == 'UMSRS':
        pattern = path + '/*УМСРС*.xlsx'
    if category == 'Vakcina':
        pattern = path + '/Федеральный регистр вакцинированных*.xlsx'

    file_excel = glob.glob(pattern)
    file_csv = glob.glob(pattern[:-4] + 'csv')
    if len(file_csv):
        return True, True, file_csv[0]
    else:
        if len(file_excel):
            return True, False, file_excel[0]
        else:
            return False, False, None

def check_file(file,category):
    if category == 'fr':
        names = [
    'п/н','Дата создания РЗ','УНРЗ','Дата изменения РЗ','СНИЛС','ФИО','Пол','Дата рождения','Диагноз','Диагноз установлен','Осложнение основного диагноза','Субъект РФ','Медицинская организация','Ведомственная принадлежность','Вид лечения','Дата исхода заболевания','Исход заболевания','Степень тяжести','Посмертный диагноз','ИВЛ','ОРИТ','МО прикрепления','Медицинский работник'
                ]
    if category == 'fr_death':
        names = [
    '1', '2', '3', '4', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22','23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '34', '35', '36', '37', '38', '39', '40', '41', '42','43', '44', '45', '46', '47', '48', '49', '50', '51', '52', '53'
                ]
    if category == 'fr_lab':
        names = [
    'Субъект', 'УНРЗ', 'Мед. организация', 'Основной диагноз', 'Наименование лаборатории', 'Дата лабораторного теста','Тип лабораторного теста', """Результат теста
(положительный/
отрицательный)""", 'Этиология пневмония','Дата первого лабораторного подтверждения COVID-19','Дата последнего лабораторного подтверждения COVID-19'
                ]
    if category == 'UMSRS':
        names = [
    '№ п/п', 'Номер свидетельства о смерти', 'Дата выдачи', 'Категория МС', 'Фамилия', 'Имя', 'Отчество', 'Пол', 'Дата рождения','Дата смерти', 'Возраст', 'Страна', 'Республика', 'Субъект', 'Область', 'Район', 'Город', 'Населенный пункт', 'Элемент планировочной структуры','Район СПБ', 'Улица', 'Дом', 'Корпус', 'Строение', 'Квартира', 'Страна смерти', 'Республика смерти','Субъект смерти', 'Область смерти','Район смерти', 'Город смерти', 'Населенный пункт смерти', 'Элемент планировочной структуры смерти', 'Район СПБ смерти', 'Улица смерти','Дом смерти', 'Корпус смерти', 'Строение смерти', 'Квартира смерти', 'Место смерти', 'Код МКБ-10 а', 'Болезнь или состояние, непосред приведшее к смерти','Код МКБ-10 б', 'Патол. состояние, кот. привело к указанной причине', 'Код МКБ-10 в', 'Первоначальная причина смерти', 'Код МКБ-10 г','Внешняя причина при травмах и отравлениях','Код II', 'Прочие важние состояния', 'Код МКБ-10 а(д)', 'Основное заболевание плода или ребенка','Код МКБ-10 б(д)', 'Другие заболевания плода или ребенка', 'Код МКБ-10 в(д)', 'Основное заболевание матери', 'Код МКБ-10 г(д)','Другие заболевания матери', 'Код МКБ-10 д(д)', 'Другие обстоятельства мертворождения', 'Установил причины смерти', 'Адрес МО','Краткое наименование', 'Основания установления причин смерти', 'Осмотр трупа', 'Записи в мед.док.', 'Предшествующего наблюдения','Вскрытие', 'Статус МС', 'Взамен', 'Дубликат', 'Испорченное', 'Напечатано', 'в случае смерти результате ДТП'
                ]
    if category == 'Vakcina':
        names = [
    'п/н', 'Дата создания РЗ', 'Дата обновления РЗ', 'УНРЗ', 'СНИЛС', 'ФИО', 'Пол', 'Дата рождения', 'Вакцина','Дата вакцинации', 'Субъект РФ', 'Медицинская организация', 'Ведомственная принадлежность', 'Структурное подразделение', 'Статус иммунизации', 'Дневник самонаблюдения', 'Сведения об осложнениях', 'Наличие пациента в Регистре COVID'      
                ]

    sum_colum = len(names)
    names_found = []
    num_colum = 0
    try:
        if category == 'fr':
            file =  pd.read_csv(file,header=None, delimiter=';', engine='python', encoding='utf-8') #,encoding = 'cp1250')
        else:
            file =  pd.read_csv(file,header=None, delimiter=';', engine='python')
    except:
        return [False,'Файл не читается','',0]
    for head in range(len(file)):
        if num_colum != sum_colum:
            coll = file.loc[head].tolist()
            num_colum = 0
            collum = []
            check = True
            error_text = '' 
            for name in names:
                k = 0
                f = 0
                for col in coll:
                    if str(col).replace(' ','') == name.replace(' ',''): 
                        collum.append(k)
                        names_found.append(name)
                        num_colum += 1 
                        f = 1
                    k += 1
        else:
            break
    if len(names_found) < sum_colum:
        check = False
        error_text = ' Не найдена колонка ' + str(list(set(names) - set(names_found) ) ) + ';'
        return check,error_text,collum, head - 1
    return check,error_text,collum, head - 1

def load_vaccina(a):
    files = glob.glob(get_dir('path_robot') + '/_ФР_по_частям/Федеральный регистр вакцинированных*')
    if not len(files):
        raise my_except('В папке нет файлов!')
    names = [
    'п/н', 'Дата создания РЗ', 'Дата обновления РЗ', 'УНРЗ', 'СНИЛС', 'ФИО', 'Пол', 'Дата рождения', 'Вакцина','Дата вакцинации', 'Субъект РФ', 'Медицинская организация', 'Ведомственная принадлежность', 'Структурное подразделение', 'СНИЛС медицинского работника','Статус иммунизации', 'Дневник самонаблюдения', 'Сведения об осложнениях', 'Наличие пациента в Регистре COVID'      
        ]
    cols = ['number','Data_sozdaniya_RZ','Data_obnovleniya_RZ','UNRZ','SNILS','FIO','Pol','Data_rozhdeniya','Vaktsina','Data_vaktsinatsii','Subyekt_RF','Meditsinskaya_organizatsiya','Vedomstvennaya_prinadlezhnost','Strukturnoye_podrazdeleniye','SNILS_MO_employee','Status_immunizatsii','Dnevnik_samonablyudeniya','Ob_informatsii_oslozhneniyakh','Nalichiye_patsiyenta_v_Registre_COVID']
    list_ = []
    for file in files:
        part = pd.DataFrame()
        try:
            part = pd.read_excel(file, usecols=names,  header=0,skipfooter=1 )
        except:
            try:
                part = pd.read_excel(file, usecols=names,  header=1,skiprows = 1, skipfooter=1 )
            except Exception as e:
                send('admin', file.rsplit('/',1)[-1] +'\n'+ str(e) )
        
        if len(part):
            send('admin','Я прочёл ' + file.rsplit('/',1)[-1])
            list_.append(part)
        else:
            send('admin','Я не смог прочитать ' + file.rsplit('/',1)[-1])

    df = pd.concat(list_)
    df.columns = cols
    send('admin',str(df.columns))
    send('admin','Файл в памяти, количество строк: ' + str(len(df)) )
    sql_execute('TRUNCATE TABLE [dbo].[cv_fedRegVakcin]')
    send('admin','Очистил cv_fedRegVakcin')

    with sqlalchemy.create_engine(f"mssql+pymssql://{user}:{passwd}@{server}/{dbase}",pool_pre_ping=True).connect() as con:
        df.to_sql('cv_fedRegVakcin',con,schema='dbo',if_exists='replace',index=False)
    if check_table('fedRegVakcin'):
        send('admin','Федеральный регистр успешно загружен')
        return 1
    else:
        send('admin','Произошла какая-то проблема с загрузкой вакцинируемых')
        return 0

def slojit_fr(a):
    import xlrd
    xlrd.xlsx.ensure_elementtree_imported(False, None)
    xlrd.xlsx.Element_has_iter = True

    file = '/mnt/COVID-списки/Входящие списки/для загрузки/Robot/_ФР_по_частям/Федеральный регистр лиц, больных COVID-19 - 2021-04-18T203626.396.xlsx'
    names = ['п/н','Дата создания РЗ','УНРЗ','Дата изменения РЗ','СНИЛС','ФИО','Пол','Дата рождения','Диагноз','Диагноз установлен','Осложнение основного диагноза',
             'Субъект РФ','Медицинская организация','Ведомственная принадлежность','Вид лечения','Дата исхода заболевания','Исход заболевания','Степень тяжести',
             'Посмертный диагноз','ИВЛ','ОРИТ','МО прикрепления','Медицинский работник']  

    path = get_dir('path_robot') + '/_ФР_по_частям'
    date = datetime.datetime.today().strftime("%Y_%m_%d")
    nameSheetShablon = "Sheet1"
    _list = []
    
    files = glob.glob(path + '/Федеральный регистр лиц*.xlsx')
    if not len(files):
        raise my_except('В папке нет файлов!')

    for excel in files:
        try:
            df = pd.read_excel(excel,  header= 1, usecols=names,  engine='xlrd',skipfooter=1 )
        except:
            raise my_except('Какой-то непонятный файл ' + excel.rsplit('/',1)[-1])
        send('epid','прочтён файл ' + excel.rsplit('/',1)[-1])
        _list.append(df)
    
    svod = pd.concat(_list)

    #svod = svod[svod["Дата создания РЗ"].notnull()] 
    svod["п/н"] = range(1, len(svod)+1)
    tumorow = (datetime.datetime.today() + datetime.timedelta(days=1)).strftime("%Y_%m_%d")
    
    new_fedreg      = get_dir('robot') +'/'+ tumorow + '/Федеральный регистр лиц, больных - ' + date + '.csv'
    new_fedreg_temp = get_dir('temp') + '/Федеральный регистр лиц, больных - ' + date + '.csv'
    new_iach        = get_dir('covid_iac2') + '/Федеральный регистр лиц, больных - ' + date + '_ИАЦ.csv'
    new_iach_temp   = get_dir('temp') + '/Федеральный регистр лиц, больных - ' + date + '_ИАЦ.csv'
    otchet_9        = glob.glob(get_dir('robot') +'/'+ tumorow +'/9. Отчет по пациентам COVID-центр*.xlsx' )
    if len(otchet_9):
        otchet_9_new = get_dir('covid_iac2') +'/'+ otchet_9[0].rsplit('/',1)[1]
        shutil.copyfile(otchet_9[0],otchet_9_new)
        send('epid','Отчет №9 скопирован в папку иац')
    else:
        send('epid','Не удалось найти отчет №9')
    
    NumberForMG = len(svod[svod['Диагноз'].isin(['U07.1']) \
                    & svod['Исход заболевания'].isnull() \
                    & svod['Вид лечения'].isin(['Стационарное лечение'])])

    NumberFor1 = len(svod[svod['Диагноз'].isin(['U07.1']) ])

    NumberFor2 = len(svod[svod['Посмертный диагноз'].isin(['U07.1']) \
                        & svod['Исход заболевания'].isin(['Смерть'])])

    day = pd.to_datetime(svod['Дата изменения РЗ'], format='%d.%m.%Y').max().date()
    count_vizd_old = pd.read_sql ("""SELECT [value_count] FROM [robo].[values]
                where id = (select max(id) from [robo].[values] where [value_name] = 'Всего выздоровело от COVID' 
                and date_rows = (select max(date_rows) from [robo].[values] where [value_name] = 'Всего выздоровело от COVID'
                and  date_rows != cast(getdate() as date) ) )""",con).iat[0,0]
    count_vizd_new = len(svod[svod['Исход заболевания'].isin(['Выздоровление']) & svod['Диагноз'].isin(['U07.1']) ])

    NumberFor3 = count_vizd_new - count_vizd_old
    # расчет людей на стационарном лечении
    svod['Возраст'] = (pd.to_datetime(svod['Диагноз установлен'], format="%d.%m.%Y") - pd.to_datetime(svod['Дата рождения'], format="%d.%m.%Y")) / numpy.timedelta64(1, 'Y')
    NumberFor4_1 = len(svod.loc[(svod['Исход заболевания'].isnull()) \
            & (svod['Диагноз'].isin(['U07.1','U07.2']) | svod['Диагноз'].str.contains('J1[2-8]') ) \
            & (svod['Вид лечения'].isin(['Стационарное лечение']))])
    NumberFor4_2 = len(svod.loc[(svod['Исход заболевания'].isnull()) & ( svod['Возраст'] < 18) \
            & (svod['Диагноз'].isin(['U07.1','U07.2']) | svod['Диагноз'].str.contains('J1[2-8]') ) \
            & (svod['Вид лечения'].isin(['Стационарное лечение']))])
    NumberFor4_3 = len(svod.loc[(svod['Исход заболевания'].isnull()) & ( svod['Возраст'] >= 60) \
            & (svod['Диагноз'].isin(['U07.1','U07.2']) | svod['Диагноз'].str.contains('J1[2-8]') ) \
            & (svod['Вид лечения'].isin(['Стационарное лечение']))])
    NumberFor4_4 = len(svod.loc[(svod['Исход заболевания'].isnull()) & ( svod['Возраст'] >= 70) \
            & (svod['Диагноз'].isin(['U07.1','U07.2']) | svod['Диагноз'].str.contains('J1[2-8]') ) \
            & (svod['Вид лечения'].isin(['Стационарное лечение']))])

    NumberFor5_1 = len(svod.loc[(svod['Исход заболевания'].isnull()) \
            & (svod['Диагноз'].isin(['U07.1','U07.2']) | svod['Диагноз'].str.contains('J1[2-8]') ) \
            & (svod['Вид лечения'].isin(['Амбулаторное лечение']))])
    NumberFor5_2 = len(svod.loc[(svod['Исход заболевания'].isnull()) & ( svod['Возраст'] < 18) \
            & (svod['Диагноз'].isin(['U07.1','U07.2']) | svod['Диагноз'].str.contains('J1[2-8]') ) \
            & (svod['Вид лечения'].isin(['Амбулаторное лечение']))])
    NumberFor5_3 = len(svod.loc[(svod['Исход заболевания'].isnull()) & ( svod['Возраст'] >= 60) \
            & (svod['Диагноз'].isin(['U07.1','U07.2']) | svod['Диагноз'].str.contains('J1[2-8]') ) \
            & (svod['Вид лечения'].isin(['Амбулаторное лечение']))])
    NumberFor5_4 = len(svod.loc[(svod['Исход заболевания'].isnull()) & ( svod['Возраст'] >= 70) \
            & (svod['Диагноз'].isin(['U07.1','U07.2']) | svod['Диагноз'].str.contains('J1[2-8]') ) \
            & (svod['Вид лечения'].isin(['Амбулаторное лечение']))])

    NumberFor7 = len(svod.loc[(svod['Исход заболевания'].str.contains('Выздоровление')) \
            & (svod['Диагноз'].isin(['U07.1','U07.2']) ) \
            & ( pd.to_datetime(svod['Дата исхода заболевания'], format='%d.%m.%Y', errors='ignore' )  > (datetime.datetime.now() - datetime.timedelta(days=181) ) ) ] )
    
    otvet = 'По цифрам\n' \
            + 'На стационарном лечении (U07.1): ' + format(NumberForMG,'n') + '\n' \
            + 'Всего заболело: ' + format(NumberFor1,'n') +'\n' \
            + 'Всего умерло: '+ format(NumberFor2,'n') + '\n' \
            + 'Всего выздоровело за ' + str(day) + ' : '+ format(NumberFor3, 'n') + '\n'\
            + 'Сейчас на стационаром лечении: ' + format(NumberFor4_1, 'n') + '\n' \
            + 'Сейчас на стационаром лечении младше 18: ' + format(NumberFor4_2, 'n') + '\n' \
            + 'Сейчас на стационаром лечении старше 60: ' + format(NumberFor4_3, 'n') + '\n' \
            + 'Сейчас на стационаром лечении старше 70: ' + format(NumberFor4_4, 'n')  + '\n' \
            + 'Сейчас на амбулаторном лечении: ' + format(NumberFor5_1, 'n') + '\n' \
            + 'Сейчас на амбулаторном лечении младше 18: ' + format(NumberFor5_2, 'n') + '\n' \
            + 'Сейчас на амбулаторном лечении старше 60: ' + format(NumberFor5_3, 'n') + '\n' \
            + 'Сейчас на амбулаторном лечении старше 70: ' + format(NumberFor5_4, 'n')  + '\n' \
            + 'Всего выздоровело от ковида за последние 180 дней: ' + format(NumberFor7, 'n')
    
    send('epid',otvet)
    count_deti_ill   = len(svod.loc[ (svod['Диагноз'].isin(['U07.1','U07.2'])) & ( svod['Возраст'] < 18) ] )
    count_deti_rec   = len(svod.loc[ (svod['Диагноз'].isin(['U07.1','U07.2'])) & ( svod['Возраст'] < 18) \
            & (svod['Исход заболевания'].str.contains('Выздоровление')) ])
    count_deti_death = len(svod.loc[ (svod['Посмертный диагноз'].isin(['U07.1','U07.2'])) & ( svod['Возраст'] < 18)  \
            & (svod['Исход заболевания'].isin(['Смерть']) ) ]) 

    count_deti_amb   = len(svod.loc[ (svod['Диагноз'].isin(['U07.1','U07.2'])) & ( svod['Возраст'] < 18) \
            & (svod['Вид лечения'].isin(['Амбулаторное лечение']) ) & (svod['Исход заболевания'].isnull()) ] )

    count_deti_stach = len(svod.loc[ (svod['Диагноз'].isin(['U07.1','U07.2'])) & ( svod['Возраст'] < 18) \
            & (svod['Вид лечения'].isin(['Стационарное лечение']) ) & (svod['Исход заболевания'].isnull()) ] )

    count_deti_ill_old = pd.read_sql ("""SELECT [value_count] FROM [robo].[values]
                where id = (select max(id) from [robo].[values] where [value_name] = 'Всего детей заболело от COVID' 
                and date_rows = (select max(date_rows) from [robo].[values] where [value_name] = 'Всего детей заболело от COVID'
                and  date_rows != cast(getdate() as date) ) )""",con).iat[0,0]
    
    count_deti_rec_old = pd.read_sql ("""SELECT [value_count] FROM [robo].[values]
                where id = (select max(id) from [robo].[values] where [value_name] = 'Всего детей выздоровело от COVID' 
                and date_rows = (select max(date_rows) from [robo].[values] where [value_name] = 'Всего детей выздоровело от COVID'
                and  date_rows != cast(getdate() as date) ) )""",con).iat[0,0]

    count_deti_death_old = pd.read_sql ("""SELECT [value_count] FROM [robo].[values]
                where id = (select max(id) from [robo].[values] where [value_name] = 'Всего детей умерло от COVID' 
                and date_rows = (select max(date_rows) from [robo].[values] where [value_name] = 'Всего детей умерло от COVID'
                and  date_rows != cast(getdate() as date) ) )""",con).iat[0,0]

    otvet2 = "Отдельно по детям, больным COVID-19:\n" \
            + 'Всего детей заболело: ' +  format(count_deti_ill,'n') +'\n' \
            + 'Всего детей заболело за день: ' +  format(count_deti_ill - count_deti_ill_old,'n') +'\n' \
            + 'Всего детей выздоровело: ' +  format(count_deti_rec,'n') +'\n' \
            + 'Всего детей выздоровело за день: ' +  format(count_deti_rec - count_deti_rec_old,'n') +'\n' \
            + 'Всего детей умерло: ' +  format(count_deti_death,'n') +'\n' \
            + 'Всего детей умерло за день: ' +  format(count_deti_death - count_deti_death_old,'n') +'\n' \
            + 'Всего детей с COVID-19 на амбулаторном: ' +  format(count_deti_amb,'n') +'\n' \
            + 'Всего детей с COVID-19 на стационарном: ' +  format(count_deti_stach,'n')

    send('epid', otvet2)
    # считаю школьников
    count_deti_ill   = len(svod.loc[ (svod['Диагноз'].isin(['U07.1','U07.2'])) & ( svod['Возраст'] < 18) & ( svod['Возраст'] > 6) ] )
    count_deti_rec   = len(svod.loc[ (svod['Диагноз'].isin(['U07.1','U07.2'])) & ( svod['Возраст'] < 18) & ( svod['Возраст'] > 6)\
            & (svod['Исход заболевания'].str.contains('Выздоровление')) ])
    count_deti_death = len(svod.loc[ (svod['Посмертный диагноз'].isin(['U07.1','U07.2'])) & ( svod['Возраст'] < 18) & ( svod['Возраст'] > 6) \
            & (svod['Исход заболевания'].isin(['Смерть']) ) ]) 

    count_deti_amb   = len(svod.loc[ (svod['Диагноз'].isin(['U07.1','U07.2'])) & ( svod['Возраст'] < 18) & ( svod['Возраст'] > 6)\
            & (svod['Вид лечения'].isin(['Амбулаторное лечение']) ) & (svod['Исход заболевания'].isnull()) ] )

    count_deti_stach = len(svod.loc[ (svod['Диагноз'].isin(['U07.1','U07.2'])) & ( svod['Возраст'] < 18) & ( svod['Возраст'] > 6)\
            & (svod['Вид лечения'].isin(['Стационарное лечение']) ) & (svod['Исход заболевания'].isnull()) ] )

    count_deti_ill_old = pd.read_sql ("""SELECT [value_count] FROM [robo].[values]
                where id = (select max(id) from [robo].[values] where [value_name] = 'Всего школьников заболело от COVID' 
                and date_rows = (select max(date_rows) from [robo].[values] where [value_name] = 'Всего школьников заболело от COVID'
                and  date_rows != cast(getdate() as date) ) )""",con).iat[0,0]
    
    count_deti_rec_old = pd.read_sql ("""SELECT [value_count] FROM [robo].[values]
                where id = (select max(id) from [robo].[values] where [value_name] = 'Всего школьников выздоровело от COVID' 
                and date_rows = (select max(date_rows) from [robo].[values] where [value_name] = 'Всего школьников выздоровело от COVID'
                and  date_rows != cast(getdate() as date) ) )""",con).iat[0,0]

    count_deti_death_old = pd.read_sql ("""SELECT [value_count] FROM [robo].[values]
                where id = (select max(id) from [robo].[values] where [value_name] = 'Всего школьников умерло от COVID' 
                and date_rows = (select max(date_rows) from [robo].[values] where [value_name] = 'Всего школьников умерло от COVID'
                and  date_rows != cast(getdate() as date) ) )""",con).iat[0,0]

    otvet3 = "Отдельно по школьникам, больным COVID-19:\n" \
            + 'Всего заболело: ' +  format(count_deti_ill,'n') +'\n' \
            + 'Всего заболело за день: ' +  format(count_deti_ill - count_deti_ill_old,'n') +'\n' \
            + 'Всего выздоровело: ' +  format(count_deti_rec,'n') +'\n' \
            + 'Всего выздоровело за день: ' +  format(count_deti_rec - count_deti_rec_old,'n') +'\n' \
            + 'Всего умерло: ' +  format(count_deti_death,'n') +'\n' \
            + 'Всего умерло за день: ' +  format(count_deti_death - count_deti_death_old,'n') +'\n' \
            + 'Всего с COVID-19 на амбулаторном: ' +  format(count_deti_amb,'n') +'\n' \
            + 'Всего с COVID-19 на стационарном: ' +  format(count_deti_stach,'n')

    send('epid', otvet3)


    send('epid','Начинаю записывать файлы')

    #with pd.ExcelWriter(new_fedreg_temp) as writer:
    #    svod.to_excel(writer,index=False)

    svod.to_csv(new_fedreg_temp,index=False,sep=";", encoding='utf-8')#encoding='cp1251')
    try:
        shutil.move(new_fedreg_temp,new_fedreg)
    except:
        shutil.move(new_fedreg_temp, path +'/' + new_fedreg_temp.rsplit('/',1)[-1])

    send('epid','Записан файл фед регистра')

    del svod['СНИЛС']
    del svod['ФИО']

    #with pd.ExcelWriter(new_iach_temp) as writer:
    #    svod.to_excel(writer,index=False)

    svod.to_csv(new_iach_temp,index=False,sep=";",encoding='cp1251')#,encoding='cp1251')

    shutil.move(new_iach_temp,new_iach)
    send('epid','Записан файл иац')
    
    return 'Фух, закончил'

def excel_to_csv_old(file_excel):
    file_csv = file_excel[:-4] + 'csv'
    sheet = xlrd.open_workbook(file_excel).sheet_by_index(0) 
    col = csv.writer(open(file_csv, 'w', encoding='cp1251',  newline=""),delimiter=";") 
    
    for row in range(sheet.nrows): 
        col.writerow(sheet.row_values(row))
    return file_csv

def excel_to_csv(file_excel):
    file_csv = file_excel[:-4] + 'csv'
    excel = openpyxl.load_workbook(file_excel)
    sheet = excel.active 
    col = csv.writer(open(file_csv, 'w',newline=""),delimiter=';')
    for r in sheet.rows: 
        col.writerow([cell.value for cell in r]) 
    return file_csv

def sql_execute(sql):
    Session = sessionmaker(bind=con)
    session = Session()
    session.execute(sql)
    session.commit()
    session.close()

def load_fr(a):
    def fr_to_sql(df):
        df  = df[df['Дата создания РЗ'] != '']
        del df ['Ведомственная принадлежность']
        del df ['Осложнение основного диагноза']

        df ['Возраст'] = (pd.to_datetime(df['Диагноз установлен'], format='%d.%m.%Y') \
                - pd.to_datetime(df['Дата рождения'], format='%d.%m.%Y' ))/ numpy.timedelta64(1, 'Y') 

        # ==== Репорт о количестве выздоровевших =====
        report = pd.DataFrame()
        report.loc[0,'date_rows'] = pd.to_datetime(df['Дата создания РЗ'],format='%d.%m.%Y').max().date()
        report.loc[0,'value_name'] = 'Всего выздоровело от COVID'
        report.loc[0,'value_count'] = len(df[df['Исход заболевания'].isin(['Выздоровление']) & df['Диагноз'].isin(['U07.1'])])

        report.loc[1,'date_rows'] = pd.to_datetime(df['Дата создания РЗ'],format='%d.%m.%Y').max().date()
        report.loc[1,'value_name'] = 'Всего на амбулаторном лечении от COVID'
        report.loc[1,'value_count'] = len(df.loc[(df['Исход заболевания'] == '' ) \
                & (df['Диагноз'].isin(['U07.1','U07.2']) | df['Диагноз'].str.contains('J1[2-8]') ) \
                & (df['Вид лечения'].isin(['Амбулаторное лечение']))] )

        report.loc[2,'date_rows'] = pd.to_datetime(df['Дата создания РЗ'],format='%d.%m.%Y').max().date()
        report.loc[2,'value_name'] = 'Всего детей заболело от COVID'
        report.loc[2,'value_count'] = len(df.loc[ (df['Диагноз'].isin(['U07.1','U07.2'])) & ( df['Возраст'] < 18) ] )

        report.loc[3,'date_rows'] = pd.to_datetime(df['Дата создания РЗ'],format='%d.%m.%Y').max().date()
        report.loc[3,'value_name'] = 'Всего детей выздоровело от COVID'
        report.loc[3,'value_count'] =  len(df.loc[ (df['Диагноз'].isin(['U07.1','U07.2'])) & ( df['Возраст'] < 18) \
                & (df['Исход заболевания'].str.contains('Выздоровление')) ])

        report.loc[4,'date_rows'] = pd.to_datetime(df['Дата создания РЗ'],format='%d.%m.%Y').max().date()
        report.loc[4,'value_name'] = 'Всего детей умерло от COVID'
        report.loc[4,'value_count'] = len(df.loc[ (df['Посмертный диагноз'].isin(['U07.1','U07.2'])) & ( df['Возраст'] < 18)  \
            & (df['Исход заболевания'].isin(['Смерть']) ) ]) 

        report.loc[5,'date_rows'] = pd.to_datetime(df['Дата создания РЗ'],format='%d.%m.%Y').max().date()
        report.loc[5,'value_name'] = 'Всего школьников заболело от COVID'
        report.loc[5,'value_count'] = len(df.loc[ (df['Диагноз'].isin(['U07.1','U07.2'])) & ( df['Возраст'] < 18) & (df['Возраст'] > 6 ) ] )

        report.loc[6,'date_rows'] = pd.to_datetime(df['Дата создания РЗ'],format='%d.%m.%Y').max().date()
        report.loc[6,'value_name'] = 'Всего школьников выздоровело от COVID'
        report.loc[6,'value_count'] =  len(df.loc[ (df['Диагноз'].isin(['U07.1','U07.2'])) & ( df['Возраст'] < 18) & (df['Возраст'] > 6 ) \
                & (df['Исход заболевания'].str.contains('Выздоровление')) ])

        report.loc[7,'date_rows'] = pd.to_datetime(df['Дата создания РЗ'],format='%d.%m.%Y').max().date()
        report.loc[7,'value_name'] = 'Всего школьников умерло от COVID'
        report.loc[7,'value_count'] = len(df.loc[ (df['Посмертный диагноз'].isin(['U07.1','U07.2'])) & ( df['Возраст'] < 18) & (df['Возраст'] > 6) \
            & (df['Исход заболевания'].isin(['Смерть']) ) ]) 
        
        del df['Возраст']
        report.to_sql('values',con,schema='robo',index=False,if_exists='append')
        # ============================================
        send('admin','Файл в памяти, количество строк: ' + str(len(df)) )
        sql_execute('TRUNCATE TABLE [dbo].[cv_input_fr]')
        send('admin','Очистил input_fr')
        df.to_sql('cv_input_fr',con,schema='dbo',if_exists='append',index=False)
        send('admin','Загрузил input_fr, запускаю процедуру')
        sql_execute('EXEC [dbo].[cv_Load_FedReg]')
        sql_execute('EXEC [mz].[p_Recalculate_for_50_Report]')
        send('admin','Расчёт замечаний МЗ')
        if check_table('fedreg'):
            send('admin','Федеральный регистр успешно загружен')
            return 1
        else:
            send('admin','Произошла какая-то проблема с загрузкой фр')
            return 0
    send('admin','Я проснулся и хочу грузить фр!')
    search = search_file('fr')
    if search[0] and search[1]:
        send('admin','Вот нашёлся такой файлик:\n' + search[2].split('/')[-1])
        send('admin','Сейчас я его проверю...')
        check = check_file(search[2],'fr')
        if check[0]:
            send('admin','Файл прошёл проверку, начинаю грузить в память')
            df = pd.read_csv(search[2],header = check[3], usecols = check[2], na_filter = False, dtype = str, delimiter=';', engine='python',encoding = 'utf-8')
            print(df.head(3))
            fr_to_sql(df)
            return 1
        else:
            send('admin','Файл не прошёл проверку!\n' + check[1])
            return 0
    else:
        if search[0]:
            send('admin','Нее... я не хочу работать с xlsx, щас конвертирую!')
            file_csv = excel_to_csv(search[2]) 
            send('admin','Результат:\n' + file_csv.split('/')[-1])
            check = check_file(file_csv,'fr')
            if check[0]:
                send('admin','Файл прошёл проверку, начинаю грузить в память')
                df = pd.read_csv(file_csv,header = check[3], usecols = check[2], na_filter = False, dtype = str, delimiter=';', engine='python',encoding = 'utf-8')
                fr_to_sql(df)
                return 0
            else:
                send('admin', 'Файл не прошёл проверку!\n' + check[1])
                return 0
        else:
            send('admin','Но я не нашёл файла федерального регистра (((')
            return 0
    return 1

def load_fr_death(a): 
    def fr_death_to_sql(df):
        send('admin','Обрезаю слишком длинные строки')
        df = df.apply(lambda x: x.loc[::].str[:255] )
        send('admin','Убираю Nan из таблицы')
        df = df.apply(lambda x: x.loc[::].str.replace('nan','') )
        send('admin','Отправляю в базу')
        df.to_sql('cv_input_fr_d_all_2',con,schema='dbo',if_exists='replace',index = False)
        send('admin','Запускаю процедуры')
        sql_execute("""
                EXEC   [dbo].[Insert_Table_cv_input_fr_d_all_2]
                EXEC   [dbo].[cv_from_d_all_to_d_covid]
                EXEC   [dbo].[cv_Load_FedReg_d_All]
                EXEC   [dbo].[cv_Load_FedReg_d_covid]
                    """)
        send('admin','Успешно загружено')
        return 1
    send('admin','Пришло время спокойных')
    search = search_file('fr_death')
    if search[0] and search[1]:
        send('admin','Найден файл:\n' + search[2].split('/')[-1] + '\n Сейчас мы его рассмотрим')
        check = check_file(search[2],'fr_death')
        if check[0]:
            send('admin','Похоже файл в порядке, попробую загрузить')
            df = pd.read_csv(search[2],header = check[3], usecols = check[2], na_filter = False, dtype = str, delimiter=';', engine='python')
            fr_death_to_sql(df)
            return 1
        else:
            send('admin','Что-то с файлом ' + str(check))
            return 0
    else:
        if search[0]:
            send('admin','Давайте всё-таки работать с csv, конвертирую')
            file_csv = excel_to_csv(search[2]) 
            send('admin','Результат:\n' + file_csv.split('/')[-1])
            check = check_file(file_csv,'fr_death')
            if check[0]:
                send('admin','Теперь можно и загрузить в память')
                df = pd.read_csv(file_csv,header = check[3], usecols = check[2], na_filter = False, dtype = str, delimiter=';', engine='python')
                fr_death_to_sql(df)
                return 1
            else:
                send('admin','Файл не прошёл проверку!\n' + check[1])
                return 0
        else:
            send('admin','Не найден файл умерших')
        return 0

def load_fr_lab(a): 
    def fr_lab_to_sql(df):
        send('admin','Ну, тут надо переименовать колонки и можно грузить')
        i = 1
        for column in df.columns:
            df.rename(columns = {column: str(i)}, inplace = True)
            i+=1
        df = df.dropna(subset=['1','2','3','4','5','6'])
        send('admin','Всего строк в  лабе '+ str(len(df)))
        df.to_sql('cv_input_fr_lab_2',con,schema='dbo',if_exists='replace',index = False)
        send('admin','Остались процедуры в базе')
        sql_execute("""
                    EXEC   [dbo].[Insert_Table_cv_input_fr_lab_2]
                    EXEC   [dbo].[cv_load_frlab]
                    """)
        if check_table('fedreg_lab'):
            send('admin','Лаборатория успешно загружена')
            return 1
        else:
            send('admin','Какая-то проблема с загрузкой лаборатории')
            return 0
    send('admin','Посмотрим на файлик лабораторных исследований')
    search = search_file('fr_lab')
    if search[0] and search[1]:
        send('admin','Найден файл:\n' + search[2].split('/')[-1])
        send('admin','Сейчас мы его рассмотрим...')
        check = check_file(search[2],'fr_lab')
        if check[0]:
            send('admin','Неужели файл все-таки можно загрузить?')
            df = pd.read_csv(search[2],header = check[3], usecols = check[2], na_filter = False, dtype = str, delimiter=';', engine='python')
            fr_lab_to_sql(df)
            return 1
        else:
            send('admin','Не, это плохой файл!\n' + str(check))
            return 0
    else:
        if search[0]:
            send('admin','Ну, это точно нужно перевести в csv!')
            file_csv = excel_to_csv(search[2]) 
            send('admin','Результат:\n' + file_csv.split('/')[-1])
            check = check_file(file_csv,'fr_lab')
            if check[0]:
                send('admin','Теперь можно и загрузить в память')
                df = pd.read_csv(file_csv,header = check[3], usecols = check[2], na_filter = False, dtype = str, delimiter=';', engine='python')
                fr_lab_to_sql(df)
                return 1
            else:
                send('admin','Файл не прошёл проверку!\n' + check[1])
                return 0
        else:
            send('admin','Не найден файл лаборатории')
        return 0
def load_UMSRS(a): 
    def UMSRS_to_sql(df):
        df.to_sql('cv_input_umsrs_2',con,schema='dbo',if_exists='replace',index = False)
        send('admin','Данные загружены в input_umsrs_2, запускаю процедурки')
        sql_execute("""
                    EXEC   [dbo].[Insert_Table_cv_input_umsrs_2]
                    EXEC   [dbo].[cv_Load_UMSRS]
                    """)
        if check_table('umsrs'):
            send('admin','Успешно выполнено!')
            return 1
        else:
            send('admin','Какая-то проблема с загрузкой УМСРС')
            return 0

    send('admin','А теперь будем грузить УМСРС')
    search = search_file('UMSRS')
    if search[0]:
        names = ['№ п/п', 'Номер свидетельства о смерти', 'Дата выдачи', 'Категория МС', 'Фамилия', 'Имя', 'Отчество', 'Пол', 'Дата рождения','Дата смерти', 'Возраст', 'Страна',  'Субъект',  'Район', 'Город', 'Населенный пункт', 'Элемент планировочной структуры','Район СПБ', 'Улица', 'Дом', 'Корпус', 'Строение', 'Квартира', 'Страна смерти', 'Субъект смерти', 'Район смерти', 'Город смерти', 'Населенный пункт смерти', 'Элемент планировочной структуры смерти', 'Район СПБ смерти', 'Улица смерти','Дом смерти', 'Корпус смерти', 'Строение смерти', 'Квартира смерти', 'Место смерти', 'Код МКБ-10 а', 'Болезнь или состояние, непосред приведшее к смерти','Код МКБ-10 б', 'Патол. состояние, кот. привело к указанной причине', 'Код МКБ-10 в', 'Первоначальная причина смерти', 'Код МКБ-10 г','Внешняя причина при травмах и отравлениях','Код II-1', 'Прочие важные состояния-1', 'Код МКБ-10 а(д)', 'Основное заболевание плода или ребенка','Код МКБ-10 б(д)', 'Другие заболевания плода или ребенка', 'Код МКБ-10 в(д)', 'Основное заболевание матери', 'Код МКБ-10 г(д)','Другие заболевания матери', 'Код МКБ-10 д(д)', 'Другие обстоятельства мертворождения', 'Установил причины смерти', 'Адрес МО','Краткое наименование', 'Осмотр трупа', 'Записи в мед.док.', 'Предшествующего наблюдения','Вскрытие', 'Статус МС', 'Взамен', 'Дубликат', 'Испорченное', 'Напечатано', 'в случае смерти результате ДТП'] 
        send('admin','Найден файл, сейчас прочту')
        try:
            df = pd.read_excel(search[2],usecols=names, header=7)
        except Exception as e:
            raise my_except('Не смог обработать файл\n' + search[2] + '\n' + str(e) )
        else:
            df.columns = range(len(df.columns))
            df.head()
            UMSRS_to_sql(df)
        return 0 
    else:
        send('admin', 'Не найден файл УМСРС')

def load_report_vp_and_cv(a): 
    def open_save(file): 
        pass
    def load_file_mo(file):
        nameMO = pd.read_excel(file, sheet_name= 'Титул', header =3, usecols='H', nrows = 1).iloc[0,0]
        df = pd.read_excel(file, sheet_name= 'Данные1', header =6, usecols='C:AH', nrows = 1)
        df = df.fillna(0)
        df['nameMO'] = nameMO
        os.replace(file, path + '/' + os.path.basename(file))
        return df 
    def check_data_table(name):
        sql=f"""
            IF (EXISTS (SELECT * FROM {name})) 
                SELECT 1 ELSE SELECT 0 """
        return pd.read_sql(sql,con).iat[0,0]
    send('epid','Продготовка к отчету Мониторинг ВП и COVID')
    files = glob.glob(get_dir('VP_CV') + '/из_почты/[!~$]*.xls*')
    if len(files) == 0:
        raise my_except('Папка пустая!')
    path = get_dir('VP_CV') + '/' + datetime.datetime.now().strftime("%Y%m%d")
    if os.path.exists(path):
        pass
    else:
        try:
            os.mkdir(path)
        except OSError:
            raise my_except('не смог создать папку') 
    list_=[]
    excel = pd.DataFrame()
    for file in files:
        try:
            excel = load_file_mo(file)
        except:
            open_save(file)
            try:
                excel = load_file_mo(file)
            except:
                send('epid','не обработался следующий файл \n'+ file.split('/')[-1])
            else:
                list_.append(excel)
        else:
            list_.append(excel)
    if len(list_):
        df = pd.concat(list_)
    else:
        raise my_except('Не один файл не обработался')
    df = df.set_axis(['vp_03_Power_Count_Departments','vp_04_Power_Count_Allocated_All','vp_05_Power_Count_Allocated','vp_06_Power_Count_Reanimation_All','vp_07_Power_Count_Reanimation','vp_08_Hospital_All','vp_09_Hospital_Day','vp_10_Hospital_Hard_All','vp_11_Hospital_Hard_Reaniamation','vp_12_Hospital_Hard_Ivl','vp_13_ReceivedAll','vp_14_ReceivedDay','vp_15_DischargedAll','vp_16_DischargedDay','vp_17_DiedAll','vp_18_DiedDay','cv_19_Power_Count_Departments','cv_20_Power_Count_Allocated_All','cv_21_Power_Count_Allocated','cv_22_Power_Count_Reanimation_All','cv_23_Power_Count_Reanimation','cv_24_Hospital_All','cv_25_Hospital_Day','cv_26_Hospital_Hard_All','cv_27_Hospital_Hard_Reaniamation','cv_28_Hospital_Hard_Ivl','cv_29_ReceivedAll','cv_30_ReceivedDay','cv_31_DischargedAll','cv_32_DischargedDay','cv_33_DiedAll','cv_34_DiedDay','nameMO'], axis=1, inplace=False)
    try:
        df.to_sql('HistoryFileMO',con,schema='mon_vp',index = False,if_exists='append')
    except:
        raise my_except('не удалось загрузить в базу!')
    if check_data_table('mon_vp.v_DebtorsReport'):
        return short_report('SELECT * FROM mon_vp.v_DebtorsReport')
    else:
        #==========  тут мы грузим исходные данные в первую вкладку отчета
        df = pd.read_sql('SELECT * FROM mon_vp.v_GrandReport' ,con)
        df1 = df.loc[df.typeMO==1].sort_values(["numSort"]).drop('typeMO',1).drop('numSort',1)
        df2 = df.loc[df.typeMO==2].sort_values(["numSort"]).drop('typeMO',1).drop('numSort',1)
        shablon = get_dir('help') + '/СводОбщий_' + (datetime.datetime.now() + datetime.timedelta(days=1)).strftime("%d %m %Y") +'.xlsx'
        shutil.copyfile(get_dir('help') + '/шаблон Мониторинг ВП.xlsx', shablon)

        wb= openpyxl.load_workbook(shablon)
        ws= wb['Данные']
        ws.cell(row=1, column=2, value=(datetime.datetime.now() + datetime.timedelta(days=1)).strftime("%d.%m.%Y"))

        rows = dataframe_to_rows(df1,index=False, header=False)
        for r_idx, row in enumerate(rows,9):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        rows = dataframe_to_rows(df2,index=False, header=False)
        for r_idx, row in enumerate(rows,73):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(shablon)

        #  ========== сейчас мы загрузим и сохраним данные по проверкам

        df=pd.read_sql("exec mon_vp.p_CheckMonitorVpAndCovid",con)
        part_one = df.iloc[:,range(26)]
        part_two = df.iloc[:,[0] + list(range(26,58,1)) ]

        wb= openpyxl.load_workbook(shablon)
        ws= wb['Проверка']

        rows = dataframe_to_rows(part_one,index=False, header=False)
        for r_idx, row in enumerate(rows,3):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        wb.save(shablon)

        wb= openpyxl.load_workbook(shablon)
        ws= wb['Разница']

        rows = dataframe_to_rows(part_two,index=False, header=False)
        for r_idx, row in enumerate(rows,7):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(shablon)
        try:
            shutil.copyfile(shablon, get_dir('VP_CV') + '/' + shablon.split('/')[-1])
        except PermissionError:
            raise my_except('Закройте файлик! Не могу скопировать')
        return shablon

def load_report_guber(a):
    def load_sheet(file, sheetName, ColumsName, startRows, header_):
        df = pd.read_excel(file, sheet_name= sheetName,header = None , usecols=ColumsName,  skiprows = startRows)
        df = df.set_axis(header_, axis=1, inplace=False)
        df["idRows"] =pd.to_numeric(df["idRows"]) 
        df = df.sort_values(["idRows"])
        df = df.loc[df["idRows"].notnull()]
        df = df.drop_duplicates()
        df = df.fillna(0)
        return df
    def load_file(file, sheetName, ColumsName, startRows, header_, tableName):    
        list_.clear()
        try:
            excel = load_sheet(file, sheetName, ColumsName, startRows, header_)
        except:
            pass
        else:
            list_.append(excel)
            df = pd.concat(list_)
        try:
            df.to_sql(tableName,con,schema='mon_vp',index = False,if_exists='append')
        except:
            pass
    def check_data_tab(name):
        sql=f"""
        IF (EXISTS (SELECT * FROM {name})) 
        SELECT 1 ELSE SELECT 0 """
        return pd.read_sql(sql,con).iat[0,0]
    def UpdateShablonFile(workb, nameSheet, svod, startRows):
        ws = workb[nameSheet]
        rows = dataframe_to_rows(svod,index=False, header=False)
        for r_idx, row in enumerate(rows, startRows):  
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    directory = get_dir('MG')
    path = directory + '/' + datetime.datetime.now().strftime("%Y%m%d")
    try:
        os.mkdir(path)
    except:
        pass
    df = pd.DataFrame()
    list_ = []
    header_vp = ['idRows','nameMO','indicators','vp_Received_Count_All_SPb','vp_Received_Count_All_LO','vp_Received_Count_toDay_Spb','vp_Received_Count_toDay_LO','vp_Discharged_Count_All_SPb','vp_Discharged_Count_All_LO','vp_Discharged_Count_toDay_Spb','vp_Discharged_Count_toDay_LO','vp_Died_Count_All_SPb','vp_Died_Count_All_LO','vp_Died_Count_toDay_Spb','vp_Died_Count_toDay_LO','vp_Hospital_Count_All','vp_Hospital_Count_Spb','vp_Hospital_Count_LO','vp_Hospital_Count_Ivl'] 
    header_cv = ['idRows','nameMO','indicators','cv_Diagnosis_Count_All_SPb','cv_Diagnosis_Count_All_LO','cv_Diagnosis_Count_toDay_Spb','cv_Diagnosis_Count_toDay_LO','cv_Discharged_Count_All_SPb','cv_Discharged_Count_All_LO','cv_Discharged_Count_toDay_Spb','cv_Discharged_Count_toDay_LO','cv_Died_Count_All_SPb','cv_Died_Count_All_LO','cv_Died_Count_toDay_Spb','cv_Died_Count_toDay_LO','cv_Hospital_Count_All','cv_Hospital_Count_Spb','cv_Hospital_Count_LO','cv_Hospital_Count_Ivl']
    header_ivl = ['idRows','nameMO','ivl_Invaz_Count_All','ivl_Invaz_Count_Busy','ivl_Invaz_Count_Free_All','ivl_Invaz_Count_Faulty','ivl_NeInvaz_Count_All','ivl_NeInvaz_Count_Busy','ivl_NeInvaz_Count_Free_All','ivl_NeInvaz_Count_Faulty','ivl_Pacient_Count_All','ivl_Pacient_Count_Covid']
    header_bunk = ['idRows','nameMO','bn_Count_All','bn_Count_Ill_All','bn_Count_Ill_Faulty','bn_Count_Ill_Free']

    for file in glob.glob(directory + '/из_почты/[!~$]*.xls*'):
        load_file(file, 'Cвод_ОРВИ_и_Пневм', 'A:S', 4, header_vp, 'ReportGubernator_Pnevm')
        load_file(file, 'Свод_COVID', 'A:S', 4, header_cv, 'ReportGubernator_Covid')
        load_file(file, 'Свод_ИВЛ', 'A:L', 3, header_ivl, 'ReportGubernator_Ivl')
        load_file(file, 'Свод_Койки', 'A:F', 2, header_bunk, 'ReportGubernator_Bunk')
        os.replace(file, path + '/' + os.path.basename(file))

    if check_data_tab('mon_vp.v_DebtorsReportGubernator'):
        return short_report('SELECT * FROM mon_vp.v_DebtorsReportGubernator')
    else:
        file1 = get_dir('help') + '/09_стационары_для_Справки_Губернатора.xlsx'
        file2 = get_dir('help') + '/09_стационары_для_Справки_Губернатора2.xlsx'
        df_covid = pd.read_sql('SELECT * FROM [mon_vp].[v_GrandReport_Guber_Covid]' ,con)
        df_covid = df_covid.sort_values(["idRows"])
        df1_covid = df_covid.drop('Установлены диагнозы: вчера',1).drop('Установлены диагнозы: должно быть',1).drop('Установлены диагнозы: фактически',1).drop('на стационарном лечении: вчерашние данные',1).drop('на стационарном лечении: должно быть',1).drop('на стационарном лечении: фактически',1)

        df_pnev = pd.read_sql('SELECT * FROM [mon_vp].[v_GrandReport_Guber_Pnev]' ,con)
        df_pnev = df_pnev.sort_values(["idRows"])
        df1_pnev = df_pnev.drop('Установлены диагнозы: вчера',1).drop('Установлены диагнозы: должно быть',1).drop('Установлены диагнозы: фактически',1).drop('на стационарном лечении: вчерашние данные',1).drop('на стационарном лечении: должно быть',1).drop('на стационарном лечении: фактически',1)

        df_ivl = pd.read_sql('SELECT * FROM [mon_vp].[v_GrandReport_Guber_Ivl]' ,con)
        df_ivl = df_ivl.sort_values(["idRows"])

        df_bunk = pd.read_sql('SELECT * FROM [mon_vp].[v_GrandReport_Guber_Bunk]' ,con)
        df_bunk = df_bunk.sort_values(["idRows"])

        df_sys = pd.read_sql('SELECT * FROM mon_vp.v_GrandReport' ,con)
        df1_sys = df_sys.loc[df_sys.typeMO==1].sort_values(["numSort"]).drop('typeMO',1).drop('numSort',1)
        df2_sys = df_sys.loc[df_sys.typeMO==2].sort_values(["numSort"]).drop('typeMO',1).drop('numSort',1)

        wb = openpyxl.load_workbook(file1)

        UpdateShablonFile(wb, 'Cвод_ОРВИ_и_Пневм', df1_pnev, 6)
        UpdateShablonFile(wb, 'Свод_COVID', df1_covid, 6)
        UpdateShablonFile(wb, 'Свод_ИВЛ', df_ivl, 5)
        UpdateShablonFile(wb, 'Свод_Койки', df_bunk, 4)

        if check_data_tab('mon_vp.v_DebtorsReport'):
            pass
        else:
            UpdateShablonFile(wb, 'Отчет_СЮС', df1_sys, 9)
            UpdateShablonFile(wb, 'Отчет_СЮС', df2_sys, 73)
        
        new_file1 = directory + '/09_стационары для Справки Губернатора_'+ datetime.datetime.now().strftime("%d.%m.%Y_%H_%M") + '.xlsx'
        wb.save(new_file1)

        wb1 = openpyxl.load_workbook(file2)

        UpdateShablonFile(wb1, 'Свод', df_pnev, 8)
        UpdateShablonFile(wb1, 'Свод', df_covid, 97)

        ws = wb1['Свод']
        ws['Q2'] = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%d.%m.%Y')
                
        new_file2 = directory + '/09_стационары для Справки Губернатора_'+ datetime.datetime.now().strftime("%d.%m.%Y") + '.xlsx'
        wb1.save(new_file2)
        
        shutil.copyfile(new_file1,get_dir('temp') + '/' + new_file1.split('/')[-1])
        shutil.copyfile(new_file2,get_dir('temp') + '/' + new_file2.split('/')[-1])
        return  get_dir('temp') + '/' +  new_file1.split('/')[-1]+ ';' + get_dir('temp') + '/' +  new_file2.split('/')[-1]

