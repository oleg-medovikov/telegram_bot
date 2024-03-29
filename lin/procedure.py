import datetime,glob,os,shutil,sqlalchemy,openpyxl,requests,time
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import numpy as np
from collections import namedtuple
from dateutil import relativedelta
from collections import namedtuple
from loader import get_dir
from sending import send
from sqlalchemy.orm import sessionmaker

server  = os.getenv('server')
user    = os.getenv('mysqldomain') + '\\' + os.getenv('mysqluser')
passwd  = os.getenv('mypassword')
dbase   = os.getenv('db')

eng = sqlalchemy.create_engine(f"mssql+pymssql://{user}:{passwd}@{server}/{dbase}", pool_pre_ping=True)
con = eng.connect()

class my_exception(Exception):
    pass

def check_robot(a):
    date = datetime.datetime.today().strftime("%Y_%m_%d")
    path = get_dir('path_robot') +'/'+ date + '/*'
    spisok = 'В директории Robot сейчас лежат файлы:'
    for file in glob.glob(path):
        spisok += '\n' + file.rsplit('/',1)[-1]
    return spisok

def sort_death_mg(a):
    def search_mo(street,house):
        for mo in mo_org:
            if mo.Street == street:
                if mo.House == house:
                    return mo.Name_MO
        #print('Не найдено МО ' + street + ' ' + house)
        return 0
    date_otch = (datetime.datetime.today() - datetime.timedelta(days=0)).strftime("%d.%m.%Y")
    file      = get_dir('sort_death_mg') + r'/[!~]*Умершие от Covid-19*'+ date_otch + '*.xlsx'
    excel     = glob.glob(file)
    if len(excel) == 0:
        return 'Я не нашёл файлик за сегодня!'
    cols = ['№ п/п','Возраст','Субъект','Улица смерти','Дом смерти','Краткое наименование','Место смерти']
    df = pd.read_excel(excel[0],header=1, usecols = cols )
    df = df[(df['№ п/п'].notnull() ) & ( df['№ п/п'] != 0 ) ]
    df.index = range(len(df))
    mo = namedtuple('mo',['Name_MO','Street','House'])
    sql = """
    SELECT  [Name_MO],[Street],[House]
    FROM [COVID].[Nsi].[Address_MO]
    """
    mo_org=[]
    for row in con.execute(sql):
        mo_org.append(mo(*row))
    df.loc[~df['Место смерти'].isin(['в стационаре'])  , 'Name_MO'] = 'БСМЭ\ПАБ'
    for i in df.loc[df['Место смерти'].isin(['в стационаре'])].index:
        if search_mo(df.at[i,'Улица смерти'],df.at[i,'Дом смерти']):
            df.loc[i,'Name_MO'] = search_mo(df.at[i,'Улица смерти'],df.at[i,'Дом смерти'])
        else:
            df.loc[i,'Name_MO'] = df.at[i,'Краткое наименование']

    otchet = df[df['Субъект'] != 'Ленинградская обл'].groupby(by='Name_MO',as_index=False,).count()
    for column in otchet.columns:
        if column not in ('Name_MO','№ п/п'):
            del otchet[column]
    otchet.rename(columns = {'№ п/п':'Всего СПб'}, inplace = True)
    otchet = otchet.merge(df[df['Субъект'] == 'Ленинградская обл'].groupby(by='Name_MO',as_index=False,).count(), how='outer',left_on='Name_MO',right_on='Name_MO')
    otchet.rename(columns = {'Name_MO':'Медицинская организация', '№ п/п':'Всего ЛО'}, inplace = True)
    df = df.fillna(0)
    for column in otchet.columns:
        if column not in ('Медицинская организация','Всего СПб','Всего ЛО'):
            del otchet[column]

    vozrast = [18,65,150]
    status = ['Дети до 18 СПб','Взрослые до 65 СПб','Пенсионеры после 65 СПб','Дети до 18 ЛО','Взрослые до 65 ЛО','Пенсионеры после 65 ЛО']
    for i in range(len(otchet)):
        for stat in status:
            otchet.loc[i,stat] = 0
    for k in range(len(df)):
        if df.at[k,'Субъект'] != 'Ленинградская обл':
            if int(df.at[k,'Возраст']) < vozrast[0]:
                otchet.loc[otchet['Медицинская организация'] == df.at[k,'Name_MO'], status[0] ] += 1
            else:
                if int(df.at[k,'Возраст']) < vozrast[1]:
                    otchet.loc[otchet['Медицинская организация'] == df.at[k,'Name_MO'], status[1] ] += 1
                else:
                    if int(df.at[k,'Возраст']) < vozrast[2]:
                        otchet.loc[otchet['Медицинская организация'] == df.at[k,'Name_MO'], status[2] ] += 1
        else:
            if int(df.at[k,'Возраст']) < vozrast[0]:
                otchet.loc[otchet['Медицинская организация'] == df.at[k,'Name_MO'], status[3] ] += 1
            else:
                if int(df.at[k,'Возраст']) < vozrast[1]:
                    otchet.loc[otchet['Медицинская организация'] == df.at[k,'Name_MO'], status[4] ] += 1
                else:
                    if int(df.at[k,'Возраст']) < vozrast[2]:
                        otchet.loc[otchet['Медицинская организация'] == df.at[k,'Name_MO'], status[5] ] += 1

    shablon = get_dir('help') + '/ШаблонМГ.xlsx'
    file    = get_dir('sort_death_mg') + '/Свод по возрастам '+ date_otch + '.xlsx'

    shutil.copyfile(shablon,file)
    wb= openpyxl.load_workbook(file)
    ws = wb['Свод по возрастам']
    rows = dataframe_to_rows(otchet,index=False, header=False)

    for r_idx, row in enumerate(rows,2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(file)
    

    otchet = otchet.fillna(0)
    new = pd.DataFrame()
    for i in range(len(otchet)):
        k = len(new)
        new.loc[k,'Медицинская организация'] = otchet.at[i,'Медицинская организация']
        new.loc[k,'Всего СПб'] = otchet.at[i,'Всего СПб']
        new.loc[k,'Всего ЛО']  = otchet.at[i,'Всего ЛО']

        new.loc[k+1,'Медицинская организация'] = 'Дети до 18'
        new.loc[k+1,'Всего СПб'] = otchet.at[i,'Дети до 18 СПб']
        new.loc[k+1,'Всего ЛО']  = otchet.at[i,'Дети до 18 ЛО']

        new.loc[k+2,'Медицинская организация'] = 'Взрослые до 65'
        new.loc[k+2,'Всего СПб'] = otchet.at[i,'Взрослые до 65 СПб']
        new.loc[k+2,'Всего ЛО']  = otchet.at[i,'Взрослые до 65 ЛО']

        new.loc[k+3,'Медицинская организация'] = 'Пенсионеры после 65'
        new.loc[k+3,'Всего СПб'] = otchet.at[i,'Пенсионеры после 65 СПб']
        new.loc[k+3,'Всего ЛО']  = otchet.at[i,'Пенсионеры после 65 ЛО']

    
    # Удалить нули
    #new = new.loc[(new['Всего СПб']!=0) | (new['Всего ЛО'] != 0 ) ]
    
    wb= openpyxl.load_workbook(file)
    ws = wb['Перевернутый свод']
    rows = dataframe_to_rows(new,index=False, header=False)

    for r_idx, row in enumerate(rows,2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(file)

    return 'Сгенерирован файл ' + file.split('/')[-1]

def medical_personal_sick(a):
    medPers = pd.read_sql('EXEC  med.p_StartMedicalPersonalSick',con)
    date = datetime.datetime.today().strftime("%Y_%m_%d_%H_%M")
    file = get_dir('med_sick') + '/Заболевшие медики '+ date +'.xlsx'
    with pd.ExcelWriter(file) as writer:
        medPers.to_excel(writer,sheet_name='meducal',index=False)
    return 'Создан файл по заболевшим сотрудникам'

def svod_40_COVID_19(a):
    path = get_dir('40_covid_19') + '/[!~]*.xlsx'
    list_=[]
    usecolumns = 'A,B,C,D,F,G,I,J,L,M,O,P,R,S,U,V,X,Y,AA,AB,AC,AD,AF,AG,AI,AJ,AK'
    date_otch = None
    for xlsx in glob.glob(path):
        try:
            mo = pd.read_excel(xlsx, header=2,usecols=usecolumns,sheet_name='Для заполнения')
        except:
            raise my_exception('Проблема с файлом ' + xlsx)
        mo = mo[mo[2].notnull() & (~mo[2].isin(['Пункт вакцинации'])) ]
        if date_otch is None:
            date_otch = pd.to_datetime(xlsx.split('/')[-1].split('_Основной')[0][-1-9:],format="%d.%m.%Y").date()
        if len(mo.index) > 1:
            for i in mo.index:
                if i:
                    mo.loc[i,0] = 'Пункт вакцинации'
                    mo.loc[i,1] = mo.at[i,2].split(' ')[0]
                    mo.loc[i,2] = str(mo.at[i,2].split(' ',1)[1])
                else:
                    mo.loc[i,0] = 'Медицинская организация'
            list_.append(mo)

    if len(list_):
        df=pd.concat(list_)

        cols = list(df.columns.values)
        cols = cols[-1:] + cols[:-1]
        df = df[cols]
        df.index=range(1,len(df)+1)

        itog = pd.DataFrame()
        potreb = pd.DataFrame()
        for col in df.columns:
            if col not in [0,1,2,24,26,27,28]:
                itog.loc[0,col] = df.loc[df[0] == 'Медицинская организация', col ].sum()

        for i in range(len(df)):
            k = len(potreb)
            potreb.loc[k,'Date'] = date_otch
            potreb.loc[k,'District'] = df.iat[i,1]
            potreb.loc[k,'Type'] = df.iat[i,0]
            potreb.loc[k,'Name'] = df.iat[i,2]
            potreb.loc[k,'Vsego_day_v1'] = df.iat[i,21]
            potreb.loc[k,'Vsego_day_v2'] = df.iat[i,23]
            potreb.loc[k,'Ostatok_v1'] = df.iat[i,4] - df.iat[i,20] - df.iat[i,24]
            potreb.loc[k,'Ostatok_v2'] = df.iat[i,4] - df.iat[i,22]
            potreb.loc[k,'Potrebnost_v1'] = df.iat[i,21]
            potreb.loc[k,'Potrebnost_v2'] = df.iat[i,23]

        pd.read_sql(f"delete from robo.vaccine_potrebnost where cast([Date] as date) = '{str(date_otch)}' select 1",con)
        potreb.to_sql('vaccine_potrebnost',con,schema='robo',if_exists = 'append',index=False)

        new_name = str(date_otch) + '_40_COVID_19_cvod.xlsx'
        shablon_path = get_dir('help')

        shutil.copyfile(shablon_path + '/40_COVID_19_svod.xlsx', shablon_path  + '/' + new_name)

        wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
        ws = wb['Пунты вакцинации']
        rows = dataframe_to_rows(df,index=False, header=False)
        for r_idx, row in enumerate(rows,5):  
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        wb.save( shablon_path  + '/' + new_name) 

        return(shablon_path  + '/' + new_name)
    else:
        raise my_exception('Пустая папка!')

def svod_vachine_dates(a):
    sql="""SELECT [Date],[District],[Name],[Vsego_day_v1],[Vsego_day_v2]
    FROM [COVID].[robo].[vaccine_potrebnost]
    where [Type] = 'Пункт вакцинации' and [Name] <> '0'
    and [Date] != '2021-01-30'
    order by [Date]"""

    df = pd.read_sql(sql,con)

    res = df.pivot_table(index=['District','Name'], columns=['Date'], values=['Vsego_day_v1','Vsego_day_v2'],fill_value=0, aggfunc=np.sum)
    summ = df.pivot_table(index=['District'], columns=['Date'], values=['Vsego_day_v1','Vsego_day_v2'],fill_value=0, aggfunc=np.sum)
    total = pd.DataFrame(summ.sum()).T

    res['Район'] = res.index.get_level_values(0)
    res['Пункт вакцинации'] = res.index.get_level_values(1)
    summ['Район'] = summ.index.get_level_values(0)
    summ['Пункт вакцинации'] = 'Всего'
    total['Район'] = 'Весь город'
    total['Пункт вакцинации'] = 'Все'

    itog= pd.concat([total,summ,res], ignore_index=True)
    itog = itog.set_index(['Район','Пункт вакцинации'])     

    excel = get_dir('temp') + '/vaccine_dates.xlsx'
    itog.to_excel(excel)
    return excel

def razlojit_death_week(a):
    #date_end   = '2022-01-05'
    #date_start = '2021-12-23'
    date_end   = (datetime.datetime.today() + relativedelta.relativedelta(weeks=-1,weekday=2)).date()
    date_start = date_end - datetime.timedelta(days=6) 

    sql = f"""
    select dbo.get_Gid(idPatient) as 'Gid',[Медицинская организация],[ФИО],[Дата рождения]
    ,dbo.[f_calculation_age]([Дата рождения],[Дата исхода заболевания]) as 'Возраст'
    ,[Посмертный диагноз]
    from robo.v_FedReg
      where [Исход заболевания] = 'Смерть'
      and [Дата исхода заболевания]  BETWEEN '{date_start}' AND '{date_end}'
     -- and YEAR([Дата исхода заболевания]) = YEAR(getdate())
      and ([Посмертный диагноз]  in ('U07.1','U07.2')
           or [Посмертный диагноз] like 'J1[2-8]%' ) 
      and [Субъект РФ] = 'г. Санкт-Петербург'
    """
    df = pd.read_sql(sql,con)
    columns = ['Район проживания'
               ,'Дата начала заболевания','Факт обращения за медицинской помощью на амбулаторном этапе (да/нет)'
               ,'Дата обращения за медицинской помощью  на амбулаторном этапе','Факт выполнения КТ на амбулаторном этапе (да/нет)'
               ,'Факт выполнения ПЦР-SARS-CoV-2  на амбулаторном этапе (да/нет)'
               ,'Факт получения бесплатной лекарственной терапии (БЛТ) на амбулаторном этапе (да/нет)'
               ,'Дата госпитализации','Степень тяжести состояния при госпитализации (легкая, ср.тяжести, тяжелая)'
               ,'Поступление в ОРИТ  при госпитализации (да/нет)','Смерть наступила в первые сутки с момента госпитализации (да/нет)'
               ,'Факт получения антицитокиновой терапии в стационаре (да/нет)'
    ]

    for col in columns:
        df[col] = ''

    MOs = df['Медицинская организация'].unique()

    mo_directory = pd.read_sql('SELECT [Наименование в ФР], [user] from robo.directory', con)

    report_1 = pd.DataFrame()
    for MO in MOs:
        try:
            directory = get_dir('covid') + mo_directory.loc[mo_directory['Наименование в ФР'] == MO.replace(' (стац)','').replace(' (амб.)',''), 'user'].unique()[0]
        except:
            print('Не найдена организация', MO)
        else:
            try:
                os.mkdir(directory+'Умершие за неделю')
            except:
                pass
            otchet = df[df['Медицинская организация'] == MO]
            del otchet['Gid']
            if '(стац)' in MO:
                file = directory+'Умершие за неделю' + '/умершие (стац) с '+ str(date_start) + ' по ' + str(date_end) + '.xlsx'
            elif '(амб.)' in MO:
                file = directory+'Умершие за неделю' + '/умершие (амб.) с '+ str(date_start) + ' по ' + str(date_end) + '.xlsx'
            else:
                file = directory+'Умершие за неделю' + '/умершие с '+ str(date_start) + ' по ' + str(date_end) + '.xlsx'

            print(MO+';'+file)
#            with pd.ExcelWriter(file) as writer:
#                otchet.to_excel(writer,index=False)
            shutil.copyfile(get_dir('help') + '/death_week_shablon.xlsx', file)
            wb= openpyxl.load_workbook( file)
            ws = wb['death_week']
            rows = dataframe_to_rows(otchet,index=False, header=False)
            for r_idx, row in enumerate(rows,2):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            wb.save(file) 
            k = len(report_1)
            report_1.loc[k,'Медицинская организация'] = MO
            report_1.loc[k,'файл'] = file

    report_file = get_dir('temp') + '/разложенные файлы.xlsx'
    with pd.ExcelWriter(report_file) as writer:
        report_1.to_excel(writer,sheet_name = 'файлы',index=False)
        df.to_excel(writer,sheet_name = 'свод',index=False)

    report = pd.DataFrame(df['Gid'])
    report['time'] = datetime.datetime.now()
    report['Dates'] = str(date_start) + ' - ' + str(date_end)
    report.to_sql(
        name='report_deth_week',
        con=con,
        schema='robo',
        if_exists='append',
        index=False
    )
    return report_file

def sbor_death_week_files(a):
    #date_start = '2022-11-24'
    #date_end   = '2022-11-30'
    date_end   = (datetime.datetime.today() + relativedelta.relativedelta(weeks=-1,weekday=2)).date()
    date_start = date_end - datetime.timedelta(days=6) 

    path     = get_dir('covid') + f'/EPID.COVID.*/EPID.COVID.*/Умершие за неделю/*{date_start} по {date_end}*.xlsx'
    new_path = get_dir('death_week') + f'/с {date_start} по {date_end}'

    try:
        os.mkdir(new_path)
    except:
        pass

    for file in glob.glob(path):
       shutil.copyfile(file, new_path +'/'+ file.split('/')[-1 -2] +' '+ file.split('/')[-1] )

    return f'Файлы с {date_start} по {date_end} собраны в папку'


def sbor_death_week_svod(a):
    #date_start = '2022-12-08'
    #date_end   = '2022-12-14'
    date_end   = (datetime.datetime.today() + relativedelta.relativedelta(weeks=-1,weekday=2)).date()
    date_start = date_end - datetime.timedelta(days=6) 

    path     = get_dir('death_week') + f'/с {date_start} по {date_end}/[!~]*[!вод]*.xlsx'
    send('', path)
    df = pd.DataFrame()
    list_=[]
    for excel in glob.glob(path):
        chast = pd.read_excel(excel)
        chast ['file'] = excel.rsplit('/',1)[1]
        list_.append(chast)
    df = pd.concat(list_)
    df_file = get_dir('temp') + '/df_file.xlsx'
    df.to_excel(df_file) 
    #send('',str(len(df)) + ' , ' + str(len(glob.glob(path))))
    df = df.loc[~df['Медицинская организация'].isnull()]
    df['Ndays'] = (pd.to_datetime(df['Дата госпитализации'],errors='coerce') - pd.to_datetime(df['Дата начала заболевания'],errors='coerce')).dt.days
    df = df.loc[~df['ФИО'].isnull()]
    df.index = range(len(df))
    cheak_diagnoz = df[['ФИО','Дата рождения','Посмертный диагноз']].to_sql('cheak_diagnoz',con,schema='tmp',if_exists='replace',index=True)
    
    sql = """select c.*,fr.[Посмертный диагноз] as 'Посмертный диагноз новый' from (
    select * from tmp.cheak_diagnoz ) as c
    left join (select * from robo.v_FedReg where [Исход заболевания] = 'Смерть') as fr
    on (c.[ФИО] = fr.[ФИО] and c.[Дата рождения] = fr.[Дата рождения])
    where  c.[Посмертный диагноз] <> fr.[Посмертный диагноз]
    order by [index]
    """
    
    cheak = pd.read_sql(sql,con)

    for i in range(len(cheak)):
        index = int(cheak.at[i,'index'])
        diagnoz = cheak.at[i,'Посмертный диагноз новый']
        df.loc[index,'Посмертный диагноз'] = diagnoz
    send('',f"exec robo.death_week_value'{date_start}','{date_end}'")
    svod = pd.read_sql(f"exec robo.death_week_value'{date_start}','{date_end}'", con)
    svod = svod.fillna(0)
    
    
    MOs = svod['Медицинская организация'].unique()
    
    for org in df['Медицинская организация'].unique():
        if not org in MOs:
            send('epid', 'Неправильная организация! \n' + org)

    for MO in MOs:
        svod.loc[svod['Медицинская организация'].isin([MO]),'J'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Факт обращения за медицинской помощью на амбулаторном этапе (да/нет)'].isin(['да','Да']) &\
                                                        df['Посмертный диагноз'].isin(['U07.1'])
                                                                         ])
        svod.loc[svod['Медицинская организация'].isin([MO]),'K'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Факт обращения за медицинской помощью на амбулаторном этапе (да/нет)'].isin(['да','Да']) &\
                                                        df['Посмертный диагноз'].isin(['U07.2'])
                                                                         ])
        svod.loc[svod['Медицинская организация'].isin([MO]),'L'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Факт обращения за медицинской помощью на амбулаторном этапе (да/нет)'].isin(['да','Да']) &\
                                                        df['Посмертный диагноз'].isin(['J18.9'])
                                                                         ])
        svod.loc[svod['Медицинская организация'].isin([MO]),'M'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Факт получения бесплатной лекарственной терапии (БЛТ) на амбулаторном этапе (да/нет)'].isin(['да','Да']) &\
                                                        df['Посмертный диагноз'].isin(['U07.1'])
                                                                         ])
        svod.loc[svod['Медицинская организация'].isin([MO]),'N'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Факт получения бесплатной лекарственной терапии (БЛТ) на амбулаторном этапе (да/нет)'].isin(['да','Да']) &\
                                                        df['Посмертный диагноз'].isin(['U07.2'])
                                                                         ])
        svod.loc[svod['Медицинская организация'].isin([MO]),'O'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Посмертный диагноз'].isin(['U07.1']) & (df['Ndays'] > 5)
                                                                         ])
        svod.loc[svod['Медицинская организация'].isin([MO]),'P'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Посмертный диагноз'].isin(['U07.2']) & (df['Ndays'] > 5) 
                                                                         ])
        svod.loc[svod['Медицинская организация'].isin([MO]),'Q'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Посмертный диагноз'].isin(['J18.9']) & (df['Ndays'] > 5)
                                                                         ])
        svod.loc[svod['Медицинская организация'].isin([MO]),'R'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Посмертный диагноз'].isin(['U07.1']) \
                                                        & (df['Степень тяжести состояния при госпитализации (легкая, ср.тяжести, тяжелая)'].str.contains('яжел') )
                                                                         ])
        svod.loc[svod['Медицинская организация'].isin([MO]),'S'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Посмертный диагноз'].isin(['U07.2']) \
                                                        & (df['Степень тяжести состояния при госпитализации (легкая, ср.тяжести, тяжелая)'].str.contains('яжел') )
                                                                         ])
        svod.loc[svod['Медицинская организация'].isin([MO]),'T'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Посмертный диагноз'].isin(['J18.9']) \
                                                        & (df['Степень тяжести состояния при госпитализации (легкая, ср.тяжести, тяжелая)'].str.contains('яжел') )
                                                                         ])
        svod.loc[svod['Медицинская организация'].isin([MO]),'U'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Посмертный диагноз'].isin(['U07.1']) & \
                                                        df['Поступление в ОРИТ  при госпитализации (да/нет)'].isin(['да','Да']) \
                                                        & (df['Степень тяжести состояния при госпитализации (легкая, ср.тяжести, тяжелая)'].str.contains('яжел') )
                                                                         ])
        svod.loc[svod['Медицинская организация'].isin([MO]),'V'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Посмертный диагноз'].isin(['U07.2']) & \
                                                        df['Поступление в ОРИТ  при госпитализации (да/нет)'].isin(['да','Да']) \
                                                        & (df['Степень тяжести состояния при госпитализации (легкая, ср.тяжести, тяжелая)'].str.contains('яжел') )
                                                                         ])
        svod.loc[svod['Медицинская организация'].isin([MO]),'W'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Посмертный диагноз'].isin(['J18.9']) & \
                                                        df['Поступление в ОРИТ  при госпитализации (да/нет)'].isin(['да','Да']) \
                                                        & (df['Степень тяжести состояния при госпитализации (легкая, ср.тяжести, тяжелая)'].str.contains('яжел') )
                                                                         ])
        svod.loc[svod['Медицинская организация'].isin([MO]),'X'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Посмертный диагноз'].isin(['U07.1']) & \
                                                        df['Факт выполнения КТ на амбулаторном этапе (да/нет)'].isin(['да','Да']) 
                                                                         ])    
        svod.loc[svod['Медицинская организация'].isin([MO]),'Y'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Посмертный диагноз'].isin(['U07.2']) & \
                                                        df['Факт выполнения КТ на амбулаторном этапе (да/нет)'].isin(['да','Да']) 
                                                                         ])    
        svod.loc[svod['Медицинская организация'].isin([MO]),'Z'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Посмертный диагноз'].isin(['J18.9']) & \
                                                        df['Факт выполнения КТ на амбулаторном этапе (да/нет)'].isin(['да','Да']) 
                                                                         ])
        svod.loc[svod['Медицинская организация'].isin([MO]),'AA'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Посмертный диагноз'].isin(['U07.1']) & \
                                                        df['Факт выполнения ПЦР-SARS-CoV-2  на амбулаторном этапе (да/нет)'].isin(['да','Да']) 
                                                                         ])
        svod.loc[svod['Медицинская организация'].isin([MO]),'AB'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Посмертный диагноз'].isin(['U07.2']) & \
                                                        df['Факт выполнения ПЦР-SARS-CoV-2  на амбулаторном этапе (да/нет)'].isin(['да','Да']) 
                                                                         ])
        svod.loc[svod['Медицинская организация'].isin([MO]),'AC'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Посмертный диагноз'].isin(['J18.9']) & \
                                                        df['Факт выполнения ПЦР-SARS-CoV-2  на амбулаторном этапе (да/нет)'].isin(['да','Да']) 
                                                                         ])    
        svod.loc[svod['Медицинская организация'].isin([MO]),'AP'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Посмертный диагноз'].isin(['U07.1']) & \
                                                        df['Факт получения антицитокиновой терапии в стационаре (да/нет)'].isin(['да','Да']) 
                                                                         ])
        svod.loc[svod['Медицинская организация'].isin([MO]),'AQ'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Посмертный диагноз'].isin(['U07.2']) & \
                                                        df['Факт получения антицитокиновой терапии в стационаре (да/нет)'].isin(['да','Да']) 
                                                                         ])    
        svod.loc[svod['Медицинская организация'].isin([MO]),'AR'] = len(df[df['Медицинская организация'].isin([MO]) & \
                                                        df['Посмертный диагноз'].isin(['J18.9']) & \
                                                        df['Факт получения антицитокиновой терапии в стационаре (да/нет)'].isin(['да','Да']) 
                                                                         ])
    svod.rename(columns= {
        'Умерло за истекшую неделю' : 'Умерло за истекшую неделю (учитываются пациенты, по которым запись о летальном исходе внесена в ФР в течение отчетной недели)'
        ,'Из них: Основная причина смерти COVID-19' : 'Из них (по данным ПАЗ или по данным Посмертного клинического диагноза): основная причина смерти - COVID-19  U07.1'
        ,'Из них: U07.2' : 'Из них, U07.2'
        ,'Из них: Внебольничные пневмонии' : 'Из них, Внебольничные пневмонии'
        ,'Умерших в возрасте до 60 лет' : 'Количество умерших в возрасте до 60 лет (от всех заболеваний)'
        ,'Умерших в возрасте до 60 лет (u07.1)' :  'Количество умерших в возрасте до 60 лет (U07.1)'
        ,'Умерших в возрасте до 60 лет (u07.2)' : 'Количество умерших в возрасте до 60 лет (U07.2)'
        ,'Умерших в возрасте до 60 лет (пневмонии)' : 'Количество умерших в возрасте до 60 лет (пневмонии)'
        ,'J' : 'Количество пациентов, умерших от U07.1, обращавшихся за медицинской помощью на амбулаторном этапе'
        ,'K' : 'Количество пациентов, умерших от U07.2, обращавшихся за медицинской помощью на амбулаторном этапе'
        ,'L' : 'Количество пациентов, умерших от внебольничной пневмонии, обращавшихся за медицинской помощью на амбулаторном этапе'
        ,'M' : 'Количество пациентов, умерших от U07.1, получавших бесплатное лекарственное лечение амбулаторно'
        ,'N' : 'Количество пациентов, умерших от U07.2, получавших бесплатное лекарственное лечение амбулаторно'
        ,'O' : 'Количество пациентов, умерших от U07.1, поступивших в стационар после 5  дней после начала заболевания'
        ,'P' : 'Количество пациентов, умерших от U07.2, поступивших в стационар после 5 дней после начала заболевания'
        ,'Q' : 'Количество пациентов, умерших от внебольничной пневмонии, поступивших в стационар после 5  дней после начала заболевания'
        ,'R' : 'Количество пациентов, умерших от U07.1, поступивших в стационар в тяжелом состоянии'
        ,'S' : 'Количество пациентов, умерших от U07.2, поступивших в стационар в тяжелом состоянии'
        ,'T' : 'Количество пациентов, умерших от ВБП, поступивших в стационар в тяжелом состоянии'
        ,'U' : 'Количество пациентов, умерших от U07.1, поступивших в стационар в тяжелом состоянии в ОРИТ'
        ,'V' : 'Количество пациентов, умерших от U07.2, поступивших в стационар в тяжелом состоянии в ОРИТ'
        ,'W' : 'Количество пациентов, умерших от ВБП, поступивших в стационар в тяжелом состоянии в ОРИТ'
        ,'X' : 'Количество пациентов, умерших от U07.1, которым на амбулаторном этапе выполнялась КТ'
        ,'Y' : 'Количество пациентов, умерших от U07.2, которым на амбулаторном этапе выполнялась КТ'
        ,'Z' : 'Количество пациентов, умерших от ВБП, которым на амбулаторном этапе выполнялась КТ'
        ,'AA': 'Количество пациентов, умерших от U07.1, которым на амбулаторном этапе выполнялась ПЦР-SARS-CoV-2'
        ,'AB': 'Количество пациентов, умерших от U07.2, которым на амбулаторном этапе выполнялась ПЦР-SARS-CoV-2'
        ,'AC': 'Количество пациентов, умерших от ВБП, которым на амбулаторном этапе выполнялась ПЦР-SARS-CoV-2'
        ,'Умершие от U07.1 имеющие СД': 'Количество пациентов, умерших от U07.1, имеющих СД'
        ,'Умершие от U07.2 имеющие СД': 'Количество пациентов, умерших от U07.2, имеющих СД'
        ,'Умершие от ВПБ имеющие СД': 'Количество пациентов, умерших от ВБП, имеющих СД'
        ,'Умершие от U07.1 имеющие АГ' : 'Количество пациентов, умерших от U07.1, имеющих АГ'
        ,'Умершие от U07.2 имеющие АГ' : 'Количество пациентов, умерших от U07.2, имеющих АГ'
        ,'Умершие от ВПБ имеющие АГ' : 'Количество пациентов, умерших от ВБП, имеющих АГ'
        ,'Умершие от U07.1 имеющие ИБС' : 'Количество пациентов, умерших от U07.1, имеющих ИБС'
        ,'Умершие от U07.2 имеющие ИБС' : 'Количество пациентов, умерших от U07.2, имеющих ИБС'
        ,'Умершие от ВПБ имеющие ИБС' : 'Количество пациентов, умерших от ВБП, имеющих ИБС'
        ,'Умершие от U07.1 имеющие ожирение' : 'Количество пациентов, умерших от U07.1, имеющих ожирение'
        ,'Умершие от U07.2 имеющие ожирение' : 'Количество пациентов, умерших от U07.2, имеющих ожирение'
        ,'Умершие от ВПБ имеющие ожирение' : 'Количество пациентов, умерших от ВБП, имеющих ожирение'
        ,'AP' : 'Количество пациентов, умерших от U07.1, получавших в стационаре антицитокиновую терапию (АЦТ)'
        ,'AQ' : 'Количество пациентов, умерших от U07.2, получавших в стационаре АЦТ'
        ,'AR' : 'Количество пациентов, умерших от ВБП, получавших в стационаре АЦТ'
    },inplace = True)

    sp = pd.DataFrame()

    summ =  svod['Из них (по данным ПАЗ или по данным Посмертного клинического диагноза): основная причина смерти - COVID-19  U07.1'].sum() \
                + svod['Из них, U07.2'].sum() + svod['Из них, Внебольничные пневмонии'].sum()

    def add_stroka_one(name,u071,u072,vpb):
        k =len(sp)
        sp.loc[k,'Столбец'] = name
        sp.loc[k,'Всего'] = u071 + u072 + vpb
        sp.loc[k,'U07.1'] = u071
        sp.loc[k,'U07.1 (процент)'] = round(100 * u071 /  summ , 1)
        sp.loc[k,'U07.2'] = u072
        sp.loc[k,'U07.2 (процент)'] = round(100 * u072 / summ , 1)
        sp.loc[k,'Пневмонии'] = vpb
        sp.loc[k,'Пневмонии (процент)'] = round(100 * vpb /  summ , 1)

    def add_stroka(name,u071,u072,vpb):
        k =len(sp)
        sp.loc[k,'Столбец'] = name
        sp.loc[k,'Всего'] = u071 + u072 + vpb
        sp.loc[k,'U07.1'] = u071
        sp.loc[k,'U07.1 (процент)'] = round(100 * u071 /  sp.at[0,'U07.1'] , 1)
        sp.loc[k,'U07.2'] = u072
        sp.loc[k,'U07.2 (процент)'] = round(100 * u072 / sp.at[0,'U07.2'] , 1)
        sp.loc[k,'Пневмонии'] = vpb
        sp.loc[k,'Пневмонии (процент)'] = round(100 * vpb /  sp.at[0,'Пневмонии'] , 1)  
        
    def add_stroka_free(name,u071,u072):
        k =len(sp)
        sp.loc[k,'Столбец'] = name
        sp.loc[k,'Всего'] = u071 + u072
        sp.loc[k,'U07.1'] = u071
        sp.loc[k,'U07.1 (процент)'] = round(100 * u071 /  sp.at[2,'U07.1'] , 1)
        sp.loc[k,'U07.2'] = u072
        sp.loc[k,'U07.2 (процент)'] = round(100 * u072 / sp.at[2,'U07.2'] , 1)

    add_stroka_one (
        'Умерли U07.1, U07.2, пневмонии за неделю'
        ,svod['Из них (по данным ПАЗ или по данным Посмертного клинического диагноза): основная причина смерти - COVID-19  U07.1'].sum()
        ,svod['Из них, U07.2'].sum()
        ,svod['Из них, Внебольничные пневмонии'].sum()

    )   

    add_stroka (
        'Умерших в возрасте до 60 лет'
        ,svod['Количество умерших в возрасте до 60 лет (U07.1)'].sum()
        ,svod['Количество умерших в возрасте до 60 лет (U07.2)'].sum()
        ,svod['Количество умерших в возрасте до 60 лет (пневмонии)'].sum()
    )

    add_stroka (
        'Обращались за медицинской помощью на догоспитальном этапе'
        ,svod['Количество пациентов, умерших от U07.1, обращавшихся за медицинской помощью на амбулаторном этапе'].sum()
        ,svod['Количество пациентов, умерших от U07.2, обращавшихся за медицинской помощью на амбулаторном этапе'].sum()
        ,svod['Количество пациентов, умерших от внебольничной пневмонии, обращавшихся за медицинской помощью на амбулаторном этапе'].sum()
    )

    add_stroka_free (
        'Получавшие бесплатную лекарственную терапию'
        ,svod['Количество пациентов, умерших от U07.1, получавших бесплатное лекарственное лечение амбулаторно'].sum()
        ,svod['Количество пациентов, умерших от U07.2, получавших бесплатное лекарственное лечение амбулаторно'].sum()
    )

    add_stroka (
        'Поступили через 5 дней и более от начала заболевания'
        ,svod['Количество пациентов, умерших от U07.1, поступивших в стационар после 5  дней после начала заболевания'].sum()
        ,svod['Количество пациентов, умерших от U07.2, поступивших в стационар после 5 дней после начала заболевания'].sum()
        ,svod['Количество пациентов, умерших от внебольничной пневмонии, поступивших в стационар после 5  дней после начала заболевания'].sum()
    )

    add_stroka (
        'Поступили в тяжелом состоянии'
        ,svod['Количество пациентов, умерших от U07.1, поступивших в стационар в тяжелом состоянии'].sum()
        ,svod['Количество пациентов, умерших от U07.2, поступивших в стационар в тяжелом состоянии'].sum()
        ,svod['Количество пациентов, умерших от ВБП, поступивших в стационар в тяжелом состоянии'].sum()
    )


    add_stroka (
        'Поступили в тяжелом состоянии (поступили в ОРИТ)'
        ,svod['Количество пациентов, умерших от U07.1, поступивших в стационар в тяжелом состоянии в ОРИТ'].sum()
        ,svod['Количество пациентов, умерших от U07.2, поступивших в стационар в тяжелом состоянии в ОРИТ'].sum()
        ,svod['Количество пациентов, умерших от ВБП, поступивших в стационар в тяжелом состоянии в ОРИТ'].sum()
    )

    add_stroka (
        'Выполнялась КТ на амбулаторном этапе'
        ,svod['Количество пациентов, умерших от U07.1, которым на амбулаторном этапе выполнялась КТ'].sum()
        ,svod['Количество пациентов, умерших от U07.2, которым на амбулаторном этапе выполнялась КТ'].sum()
        ,svod['Количество пациентов, умерших от ВБП, которым на амбулаторном этапе выполнялась КТ'].sum()
    )
    add_stroka (
        'Выполнялась ПЦР на амбулаторном этапе'
        ,svod['Количество пациентов, умерших от U07.1, которым на амбулаторном этапе выполнялась ПЦР-SARS-CoV-2'].sum()
        ,svod['Количество пациентов, умерших от U07.2, которым на амбулаторном этапе выполнялась ПЦР-SARS-CoV-2'].sum()
        ,svod['Количество пациентов, умерших от ВБП, которым на амбулаторном этапе выполнялась ПЦР-SARS-CoV-2'].sum()
    )

    add_stroka (
        'Получали антицитокиновую терапию'
        ,svod['Количество пациентов, умерших от U07.1, получавших в стационаре антицитокиновую терапию (АЦТ)'].sum()
        ,svod['Количество пациентов, умерших от U07.2, получавших в стационаре АЦТ'].sum()
        ,svod['Количество пациентов, умерших от ВБП, получавших в стационаре АЦТ'].sum()
    )

    add_stroka (
        'Сопутствующая ИБС'
        ,svod['Количество пациентов, умерших от U07.1, имеющих ИБС'].sum()
        ,svod['Количество пациентов, умерших от U07.2, имеющих ИБС'].sum()
        ,svod['Количество пациентов, умерших от ВБП, имеющих ИБС'].sum()
    )

    add_stroka (
        'Сопутствующий СД'
        ,svod['Количество пациентов, умерших от U07.1, имеющих СД'].sum()
        ,svod['Количество пациентов, умерших от U07.2, имеющих СД'].sum()
        ,svod['Количество пациентов, умерших от ВБП, имеющих СД'].sum()
    )

    add_stroka (
        'Сопутствующий артериальная гипертония'
        ,svod['Количество пациентов, умерших от U07.1, имеющих АГ'].sum()
        ,svod['Количество пациентов, умерших от U07.2, имеющих АГ'].sum()
        ,svod['Количество пациентов, умерших от ВБП, имеющих АГ'].sum()
    )

    add_stroka (
        'Сопутствующий ожирение'
        ,svod['Количество пациентов, умерших от U07.1, имеющих ожирение'].sum()
        ,svod['Количество пациентов, умерших от U07.2, имеющих ожирение'].sum()
        ,svod['Количество пациентов, умерших от ВБП, имеющих ожирение'].sum()
    )
# ========== Расчёт районов ==================
    
    
    empty = ''
    for i in df.loc[df['Район проживания'].isnull()].index:
        empty += 'Пустой район в ' + str(df.at[i,'Медицинская организация']) + '\n'

    if len(empty):
        send('epid', empty)
    df['Район проживания'] = df['Район проживания'].fillna('Пустое значение')
    df['Район проживания'] = df['Район проживания'].str.lower()
    try:
        df['Возраст'] = pd.to_numeric( df['Возраст'] )
    except Exception as e:
        send('', str(e))

    districts = df['Район проживания'].unique()
    zone = pd.DataFrame()

    for area in districts:
        k = len (zone)
        zone.loc[k,'Район проживания'] = area
        zone.loc[k,'Количество умерших'] = len(df.loc[df['Район проживания'] == area])
        zone.loc[k,'из них: в первые сутки'] = len(df.loc[(df['Район проживания'] == area) \
                & (df['Смерть наступила в первые сутки с момента госпитализации (да/нет)'].isin(['да']))])
        zone.loc[k,'из них: возраст старше 60'] = len(df.loc[(df['Район проживания'] == area) \
                & (df['Смерть наступила в первые сутки с момента госпитализации (да/нет)'].isin(['да'])) \
                & (df['Возраст'] >= 60) ])

    file_svod= get_dir('temp') + f'/Умершие за неделю с {date_start} по {date_end} свод.xlsx'
    with pd.ExcelWriter(file_svod) as writer:
        svod.to_excel(writer,sheet_name='Свод по МО',index=False)
        sp.to_excel(writer,sheet_name='Проценты',index=False)
        zone.to_excel(writer,sheet_name='Умершие по районам',index=False)
        cheak.to_excel(writer,sheet_name='Изменившиеся диагнозы',index=False)
    try:
        shutil.copyfile(file_svod,new_path + '/свод.xlsx')
    except:
        pass
    return file_svod +';'+ df_file

def patient_amb_stac(a):
    sql =  open('sql/covid/patient_amb_stac.sql','r').read()
    with sqlalchemy.create_engine(f"mssql+pymssql://{user}:{passwd}@{server}/{dbase}",pool_pre_ping=True).connect() as con:
       df = pd.read_sql(sql,con)
    
    date_otch = datetime.datetime.now().strftime('%d_%m_%Y')
    new_name  = date_otch + '_Пациенты на амб и стац лечении.xlsx'
    shablon_path = get_dir('help')

    shutil.copyfile(shablon_path + '/patient_amb_stac.xlsx' , shablon_path  + '/' + new_name)

    wb= openpyxl.load_workbook( shablon_path  + '/' + new_name)
    ws = wb['Лист1']
    rows = dataframe_to_rows(df,index=False, header=False)
    for r_idx, row in enumerate(rows,2):  
        for c_idx, value in enumerate(row,1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save( shablon_path  + '/' + new_name)
    
    return  shablon_path  + '/' + new_name

def svod_unique_patient(date_global):
    def search(fio,birthday):
        for i in range(len(svod_list)):
            if birthday == svod_list[i].birthday:
                if fio ==  svod_list[i].fio:
                    return  svod_list[i].number_row
        return None

    date_svod = (date_global - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
    
    #svod = pd.read_excel(get_dir('cov_list') + '/Автосвод ' + date_svod +'.xlsx', sheet_name = 'Свод',dtype = str)
    svod = pd.read_csv(get_dir('cov_list') + '/Автосвод ' + date_svod +'.csv',na_filter =False,dtype = str, delimiter=';', engine='python',encoding = 'utf-8')
    
    list_ = []
    date_rpn = date_global.strftime("%d.%m.%Y")
    file  = get_dir('covid') + '/EPID.COVID.RPN/Заболевшие covid в ФС за ' + date_rpn +'.xlsx'
    file2 = get_dir('covid') + '/EPID.COVID.RPN/Заболевшие covid в ФС за ' + date_rpn +'.xls'
    try:
        rpn = pd.read_excel(file,dtype = str)
    except:
        rpn = pd.read_excel(file2,dtype = str)
    rpn['Дата отчета'] = date_rpn

    svod.index = range(len(svod))

       
    send('epid', str(svod.columns) + str(rpn.columns))
    

    svod_1 = svod
    svod_1['дата рождения'] = pd.to_datetime(svod_1['дата рождения'],errors='coerce')
    svod_1 = svod_1[svod_1['дата рождения'].notnull()]
    svod_1.index = range(len(svod_1))

    rpn_1 = rpn
    rpn_1['Дата рождения '] = pd.to_datetime(rpn_1['Дата рождения '],errors='coerce')
    rpn_1 = rpn_1[rpn_1['Дата рождения '].notnull()]
    rpn_1.index = range(len(rpn_1))
    rpn_1['Номер строки'] = range(len(rpn_1))
    rpn_1.index = range(len(rpn_1))

    send('epid','cobral')
    svod_list = []
    rpn_list  = []
    human = namedtuple('hunan',['fio','birthday','number_row','dubl_row'])
    for i in range(len(svod_1)):
        svod_list.append(human(
                    str(svod_1.at[i,'Фио']).lower(),
                    svod_1.at[i,'дата рождения'].strftime('%d.%m.%Y'),
                    i,
                    -1
                    ))
    send('epid','unique rpn')
    for i in range(len(rpn_1)):
        rpn_list.append(human(
                    str(rpn_1.at[i,'фио']).lower(),
                    rpn_1.at[i,'Дата рождения '].strftime('%d.%m.%Y'),
                    i,
                    search(str(rpn_1.at[i,'фио']).lower(),rpn_1.at[i,'Дата рождения '].strftime('%d.%m.%Y'))
                    ))
    dubli = pd.DataFrame()

    for human in rpn_list:
        if human.dubl_row != None:
            k = len(dubli)
            dubli.loc[k,'Номер строки']         = human.number_row
            dubli.loc[k,'Фио']                  = human.fio
            dubli.loc[k,'Дата рождения']        = human.birthday
            dubli.loc[k,'Номер строки в своде'] = human.dubl_row

    send('epid','dubli cheak')
    """
    for human in rpn_list:
        if human.dubl_row != None:
            svod.loc[svod['№п/п']==human.dubl_row,'Дата занесения в базу'] = \
                str(svod.loc[svod['№п/п'] == human.dubl_row,'Дата занесения в базу'].unique()[0] ) \
                + " , " \
                +  rpn.loc[rpn['Unnamed: 0'] == human.number_row, 'Дата отчета'].unique()[0]
    """

    for human in rpn_list:
        if human.dubl_row != None:
            svod.loc[human.dubl_row,'Дата занесения в базу'] = \
                str(svod.at[human.dubl_row,'Дата занесения в базу'] ) \
                + " , " \
                +  rpn.at[human.number_row, 'Дата отчета']
    
    
    send('epid','changes dates')
    for human in rpn_list:
        if human.dubl_row != None:
            rpn = rpn.drop(human.number_row)
            #rpn = rpn[rpn['Unnamed: 0']!=human.number_row]

    send('epid','zakonchil dubli')
    
    rpn = rpn[['фио','Дата рождения ', 'м/ж', 'Учреждение зарегистрировавшее диагноз']]
    
    svod = svod[['Дата занесения в базу','Роспотребнадзор','Фио', 'дата рождения', 'адрес', 'Направил материал']]
    rpn.columns = ['Фио', 'дата рождения', 'адрес', 'Направил материал']
    rpn['Дата занесения в базу'] =  date_global.strftime("%Y-%m-%d")
    rpn['Роспотребнадзор'] = 'Роспотребнадзор'

    rpn.index = range(len(rpn))

    send('epid', 'new row = ' + str(len(rpn)))

    svod.index = range(len(svod))
    

    svod = svod.append(rpn, ignore_index=True)

    send('epid', 'file save')

    file = get_dir('cov_list') + '/Автосвод ' + date_global.strftime("%Y-%m-%d") +'.csv'
    
    #svod.to_csv(file,index=False,sep=";",encoding='cp1251',errors="replace")
    svod.to_csv(file,index=False,sep=";",encoding='utf-8',errors="replace")

    #with pd.ExcelWriter(file) as writer:
    #    svod.to_excel(writer,index=False, sheet_name='Свод')
    #    rpn.to_excel(writer,index=False, sheet_name='РПН')
    #    dubli.to_excel(writer,index=False, sheet_name='Дубли')
    return 'Свод сделан!'

def get_il_stopcorona(a):
    url = "https://xn--80aesfpebagmfblc0a.xn--p1ai/covid_data.json?do=region_stats&code=RU-SPE"

    data = requests.get(url).json()
    df = pd.DataFrame.from_dict(data)
    for i in range(len(df) -1):
        df.loc[i, 'value'] = int(df.at[i,'sick']) - int(df.at[i+1,'sick'])
    
    df = df.loc[~df.value.isnull()]
       
    report = pd.DataFrame()
    for i in range(len(df)):
        report.loc[i,'date_rows'] = pd.to_datetime(df.at[i,'date'],format='%d.%m.%Y').date()
        report.loc[i,'value_name'] = 'Новые заболевшие за сутки согласно СТОПКАРОНОВИРУС'
        report.loc[i,'value_count'] = df.at[i,'value']
    
    dates = '('
    for date in report.date_rows.unique():
        dates +="'" + str(date) + "',"
    
    dates = dates[:-1]
    dates +=')'
    
    sql = f"DELETE FROM Pds.stopcorona where date_rows in {dates}"
    
    #send ("", sql)
    with sqlalchemy.create_engine(f"mssql+pymssql://{user}:{passwd}@miacbase3/MIAC_DS", pool_pre_ping=True).connect() as c:
        Session = sessionmaker(bind=c)
        session = Session()
        session.execute(sql)
        session.commit()
        session.close()


    with sqlalchemy.create_engine(f"mssql+pymssql://{user}:{passwd}@miacbase3/MIAC_DS", pool_pre_ping=True).connect() as c:
        report.to_sql('stopcorona',c,schema='Pds',index=False,if_exists='append')
    report.to_sql('values',con,schema='robo',index=False,if_exists='append')

    return df.at[0,'value']

def vigruzka_death_5_oclock(a):
    path ='/mnt/COVID-списки/Covi/' + datetime.datetime.now().strftime("%Y_%m_%d")
    while True:
        file = glob.glob(path + '/*УМСРС*.xlsx')
        if len(file):
            print('Файл найден!')
            break
        else:
            print('Файл не найден!')
            time.sleep(60)

    names =['№ п/п', 'Номер свидетельства о смерти', 'Дата выдачи', 'Взамен', 'Категория МС', 'Фамилия', 'Имя', 'Отчество', 'Пол', 'Дата рождения',
         'Дата смерти', 'Возраст', 'Страна', 'Субъект', 'Район', 'Город', 'Населенный пункт', 'Элемент планировочной структуры',
         'Район СПБ', 'Улица', 'Дом', 'Корпус', 'Строение', 'Квартира', 'Страна смерти',  'Субъект смерти', 'Район смерти',
         'Город смерти', 'Населенный пункт смерти', 'Элемент планировочной структуры смерти', 'Район СПБ смерти', 'Улица смерти', 'Дом смерти',
         'Корпус смерти', 'Строение смерти', 'Квартира смерти', 'Место смерти', 'Код МКБ-10 а', 'Болезнь или состояние, непосред приведшее к смерти',
         'Код МКБ-10 б', 'Патол. состояние, кот. привело к указанной причине', 'Код МКБ-10 в', 'Первоначальная причина смерти', 'Код МКБ-10 г',
         'Внешняя причина при травмах и отравлениях', 'Код II-1', 'Прочие важные состояния', 'Код МКБ-10 а(д)', 'Основное заболевание плода или ребенка',
         'Код МКБ-10 б(д)', 'Другие заболевания плода или ребенка', 'Код МКБ-10 в(д)', 'Основное заболевание матери', 'Код МКБ-10 г(д)',
         'Другие заболевания матери', 'Код МКБ-10 д(д)', 'Другие обстоятельства мертворождения', 'Установил причины смерти',
         'Адрес МО', 'Краткое наименование', 'Осмотр трупа', 'Записи в мед.док.', 'Предшествующего наблюдения',
         'Вскрытие', 'Статус МС', 'Дубликат', 'Испорченное', 'Напечатано', 'в случае смерти результате ДТП'
        ]
    
    sheetname = 'COVID_оснв'

    try:
        fr = pd.read_excel(file[0],sheet_name = sheetname, header = 10, usecols=names, dtype=str)
    except Exception as e:
        return str(e)
    else:
        fr.to_sql(
                'cv_input_umsrs_now_day',
                con,
                schema='dbo',
                if_exists='replace',
                index = False 
                )

    Sesson = sessionmaker(bind=con)
    session = Session()
    session.execute('EXEC [dbo].[Insert_Table_cv_umsrs_now_day_2]')
    session.commit()
    session.close()

    sverka1 = pdread_sql('select * from dbo.v_Report_for_KZ_and_MO_Dead_Compare_At_All', con)
    sverka2 = pd.read_sql('select * from dbo.v_Report_for_KZ_and_MO_Dead_Compare_At_All_2', con)
    sverka3 = pd.read_sql('select * from v_Report_for_MO_Dead_Compare_At_All', con)

    path = '/mnt/COVID-списки/Covi/Programm_Reports_with_Covid'

    date = (datetime.datetime.today() + datetime.timedelta(days=1)).strftime("%Y.%m.%d")

    files = glob.glob(path + '/[!Копия]*.xlsx')

    for file in files:
        try:
            os.remove(file)
        except PermissionError:
            pass

    with pd.ExcelWriter(path + '/' + date + '_Сверка_с_ФР.xlsx') as writer:
        sverka1.to_excel(writer,index=False)
    with pd.ExcelWriter(path + '/' + date + '_COVID_Умершие.xlsx') as writer:
        sverka2.to_excel(writer,index=False)

    MOs = sverka3['Наименование МО'].unique()

    for mo in MOs:
        otchet = sverka3[sverka3['Наименование МО'] == mo ]
        otchet.index = range(len(otchet))
    
    with pd.ExcelWriter(path+ '/' + date + '_COVID_Умершие_' + mo.replace('"','') + '.xlsx') as writer:
        otchet.to_excel(writer,index=False)

    dirs = pd.red_sql('select user,[Наименование в ФР] from robo.directory', con)
    root = get_dir('covid')

    for mo in MOs:
        try:
            user = dirs.loc[dirs['Наименование в ФР'] == mo, 'user'].values[0]
        except IndexError:
            pass
        else:
            try:
                shutil.move(path+ '/'+ date + '_COVID_Умершие_' + mo.replace('"','') + '.xlsx',  root + user +  date + '_COVID_Умершие_' + mo.replace('"','') + '.xlsx')
            except PermissionError:
                pass

    return "Успешно выполнил"
