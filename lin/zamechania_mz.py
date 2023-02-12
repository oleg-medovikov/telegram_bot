import pandas as pd
import os,datetime,glob,sqlalchemy,shutil,openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

from loader import get_dir
from sending import send

server  = os.getenv('server')
user    = os.getenv('mysqldomain') + '\\' + os.getenv('mysqluser')
passwd  = os.getenv('mypassword')
dbase   = os.getenv('db')

eng = sqlalchemy.create_engine(f"mssql+pymssql://{user}:{passwd}@{server}/{dbase}",pool_pre_ping=True)
con = eng.connect()


class my_except(Exception):
    pass

def get_path_mo(organization):
    if organization == 'ФГБОУ ВО ПСПбГМУ им. И.П.Павлова Минздрава России (амб.)':
        papka = 'Замечания Мин. Здравоохранения'
    elif 'стац' in organization:
        organization = organization.replace(' (стац)','')
        papka = 'Замечания Мин. Здравоохранения/стационарное'
    elif 'амб' in organization:
        organization = organization.replace(' (амб.)','')
        papka = 'Замечания Мин. Здравоохранения/амбулаторное'
    else:
        papka = 'Замечания Мин. Здравоохранения'
    sql = f"select top(1) [user] from robo.directory where [Наименование в ФР] = '{organization}'"
    root = get_dir('covid')
    try:
        dir = pd.read_sql(sql,con).iat[0,0]
    except:
        return 0
    else:
        if not dir is None:
            return root + dir + papka
        else:
            return root + '/anower mo'

def put_excel_for_mo(df,name,date):
    if date is None:
        date = str(datetime.datetime.now().date())
    stat = pd.DataFrame()
    for org in df['Медицинская организация'].unique():
        k = len(stat)
        stat.loc[k,'Медицинская организация'] = org
        part = df[df['Медицинская организация'] == org ]
        part.index = range(1,len(part)+1)
        part.fillna(0, inplace = True)
        part = part.applymap(str)
        root = get_path_mo(org)
        if root:
            path_otch = root #+ 'Замечания Мин. Здравоохранения'
            try:
                os.makedirs(path_otch)
            except OSError:
                pass
            file = path_otch + '/' + date + ' ' + name + '.xlsx'
            with pd.ExcelWriter(file) as writer:
                part.to_excel(writer,sheet_name='унрз')
            stat.loc[k,'Статус'] = 'Файл положен'
            stat.loc[k,'Имя файла'] = file
        else:
            stat.loc[k,'Статус'] = 'Не найдена директория для файла'
    stat.index = range(1,len(stat) + 1)
    with pd.ExcelWriter(get_dir('temp') + '/отчёт по разложению ' + name + '.xlsx') as writer:
        stat.to_excel(writer,sheet_name='унрз') 

def put_svod(df,name,date):
    if date is None:
        date = str(datetime.datetime.now().date()) 
    path = get_dir('zam_svod') + '/' + name
    name = date + ' ' + name + '.xlsx'
    try:
        os.mkdir(path)
    except OSError:
        pass
    df.index = range(1, len(df) + 1)
    df = df.applymap(str)
    df.fillna('пусто')
    with pd.ExcelWriter(path + '/' + name ) as writer:
        df.to_excel(writer) 

def no_snils(a): 
    sql = open('sql/zam/no_snils.sql','r').read()
    df = pd.read_sql(sql,con)
    put_svod(df,'Нет СНИЛСа',None)
    put_excel_for_mo(df,'Нет СНИЛСа',None)
    return get_dir('temp') + '/' + 'отчёт по разложению Нет СНИЛСа.xlsx'

def bez_izhoda(a):
    sql = open('sql/zam/bez_ishod.sql','r').read()
    df = pd.read_sql(sql,con) 
    put_svod(df,'Без исхода 30 дней',None)
    put_excel_for_mo(df,'Без исхода 30 дней',None)
    return get_dir('temp') + '/' + 'отчёт по разложению Без исхода 30 дней.xlsx'

def bez_ambulat_level(a):
    sql = open('sql/zam/bez_ambulat_level.sql','r').read()
    df = pd.read_sql(sql,con) 
    put_svod(df,'Нет амбулаторного этапа',None)
    put_excel_for_mo(df,'Нет амбулаторного этапа',None)

    sql = open('sql/zam/bez_ambulat_level_death_covid.sql','r').read()
    df = pd.read_sql(sql,con) 
    put_svod(df,'Нет амбулаторного этапа, умер от COVID',None)
    put_excel_for_mo(df,'Нет амбулаторного этапа, умер от COVID',None)

    sql = open('sql/zam/bez_ambulat_level_death_nocovid.sql','r').read()
    df = pd.read_sql(sql,con) 
    put_svod(df,'Нет амбулаторного этапа, умер не от COVID',None)
    put_excel_for_mo(df,'Нет амбулаторного этапа, умер не от COVID',None)
    temp = get_dir('temp')
    files =   temp + '/' + 'отчёт по разложению Нет амбулаторного этапа.xlsx' + ';' \
            + temp + '/' + 'отчёт по разложению Нет амбулаторного этапа, умер не от COVID.xlsx' +';'\
            + temp + '/' + 'отчёт по разложению Нет амбулаторного этапа, умер от COVID.xlsx'  
    return files

def no_OMS(a):
    sql = open('sql/zam/no_OMS.sql','r').read()
    df = pd.read_sql(sql,con) 
    put_svod(df,'Нет данных ОМС',None)
    put_excel_for_mo(df,'Нет данных ОМС',None)
    return get_dir('temp') + '/' + 'отчёт по разложению Нет данных ОМС.xlsx'

def neveren_vid_lechenia(a):
    sql = open('sql/zam/neverni_vid_lecenia.sql','r').read()
    df = pd.read_sql(sql,con) 
    put_svod(df,'неверный вид лечения',None)
    put_excel_for_mo(df,'неверный вид лечения',None)
    return get_dir('temp') + '/' + 'отчёт по разложению неверный вид лечения.xlsx'

def no_lab(a):
    sql = open('sql/zam/no_lab.sql','r').read()
    df = pd.read_sql(sql,con) 
    put_svod(df,'нет лабораторного подтверждения',None)
    put_excel_for_mo(df,'нет лабораторного подтверждения',None)
    return get_dir('temp') + '/' + 'отчёт по разложению нет лабораторного подтверждения.xlsx'

def net_diagnoz_covid(a):
    sql = open('sql/zam/net_diagnoz_covid.sql','r').read()
    df = pd.read_sql(sql,con) 
    put_svod(df,'нет диагноза COVID',None)
    put_excel_for_mo(df,'нет диагноза COVID',None)
    return get_dir('temp') + '/' + 'отчёт по разложению нет диагноза COVID.xlsx'

def net_pad(a):
    sql = open('sql/zam/net_pad.sql','r').read()
    df = pd.read_sql(sql,con) 
    put_svod(df,'нет ПАЗ',None)
    put_excel_for_mo(df,'нет ПАЗ',None)
    return  get_dir('temp') + '/' + 'отчёт по разложению нет ПАЗ.xlsx'

def net_dnevnik(a):
    sql = open('sql/zam/net_dnevnik.sql','r').read()
    df = pd.read_sql(sql,con)
    put_svod(df,'нет дневниковых записей',None)
    put_excel_for_mo(df,'нет дневниковых записей',None)
    return get_dir('temp') + '/' + 'отчёт по разложению нет дневниковых записей.xlsx'

def zavishie_statusy(a):
    #sql = open('sql/zam/zavishie_status.sql','r').read()
    #df = pd.read_sql(sql,con)
    #put_svod(df,'зависшие статусы',None)
    #put_excel_for_mo(df,'зависшие статусы',None)
    #return get_dir('temp') + '/' + 'отчёт по разложению зависшие статусы.xlsx'
    sql = open('sql/zam/bez_ambulat_level_amb.sql','r').read()
    df = pd.read_sql(sql,con)
    put_svod(df,'зависшие статусы',None)
    put_excel_for_mo(df,'зависшие статусы',None)
    sql = open('sql/zam/bez_ambulat_level_noMO.sql','r').read()
    df = pd.read_sql(sql,con)
    put_svod(df,'зависшие статусы без МО прикрепления',None)
    put_excel_for_mo(df,'зависшие статусы без МО прикрепления',None)
    return get_dir('temp') + '/' + 'отчёт по разложению зависшие статусы.xlsx' +';'+ get_dir('temp') +'/'+ 'зависшие статусы без МО прикрепления.xlsx'


def delete_old_files(date):
    path = get_dir('covid')+ f"/EPID.COVID.*/EPID.COVID.*/Замечания Мин. Здравоохранения/{date.strftime('%Y-%m-%d')}*.xlsx"
    files = glob.glob(path)
    if len(files) == 0:
        return 'Нет файлов за это число!'
    for file in files:
        try:
            os.remove(file)
        except:
            pass
    return f"Я все поудалял за  дату {date.strftime('%Y-%m-%d')}"
   
def load_snils_comment(a):
    path = get_dir('snils_com') + '/*'
    text = ''
    for file in glob.glob(path):
        try:
            df = pd.read_excel(file,usecols=['УНРЗ','Примечания'])
        except Exception as e: text +=  file.split('/')[-1] +'\n' + str(e) 
        else:
            with con.connect() as cursor:
                try:
                    cursor.execute("""TRUNCATE TABLE tmp.snils_comment""")
                except:
                    pass
            df.to_sql('snils_comment',con,schema='tmp',if_exists='append',index=False)
            with con.connect() as cursor:
                cursor.execute("""
                update  a
                    set a.[Примечания] = b.[Примечания]
                    from [robo].[snils_comment] a inner join [tmp].[snils_comment] b
                    on a.[УНРЗ] = b.[УНРЗ]
            insert [robo].[snils_comment]
                    ([УНРЗ]
                  ,[Примечания])
                 SELECT b.[УНРЗ]
                        ,b.[Примечания]
                FROM [tmp].[snils_comment] b
                left join [robo].[snils_comment] a
                on a.[УНРЗ] = b.[УНРЗ]
                where a.[УНРЗ] is null
                """)
            text += '\n Хорошо обработан файл ' + file.split('/')[-1]
    return text

def IVL(a):
    path = get_dir('Robot') +'/'+ (datetime.datetime.now() - datetime.timedelta(days=0)).strftime("%Y_%m_%d")
    try:
        file_vp = glob.glob(path + '/Мониторинг_ВП.xlsx')[0]
    except:
        raise my_except('Не найден Мониторинг_ВП.xlsx')
    try:
        file_fr = glob.glob(path + '/[!~]*едеральный*15-00.csv')[0]
    except:
        raise my_except('Не найден трёхчасовой федеральный регистр')

    send('epid', 'Использую \n' + file_vp.rsplit('/',1)[-1] + '\n' + file_fr.rsplit('/',1)[-1])

    names = ['Дата изменения РЗ','Медицинская организация','Исход заболевания','ИВЛ','Вид лечения','Субъект РФ','Диагноз']
    fr = pd.read_csv(file_fr, usecols = names, sep=';', engine='python',encoding='utf-8')
    #fr = pd.read_excel(file_fr, usecols=names,dtype=str)
    date_otch =  str(pd.to_datetime(fr['Дата изменения РЗ'],format='%d.%m.%Y').max().date())
    del fr['Дата изменения РЗ']
    names = ['mo','vp_zan','vp_ivl','cov_zan','cov_ivl']
    vp = pd.read_excel(file_vp, usecols = "A,I,L,Y,AB",header=7, names=names)

    send('epid', 'Прочитал файлы')

    # Считаем числа в в федеральном регистре 
    S_org = ['ФГБОУ ВО СЗГМУ им. И.И. Мечникова Минздрава России','ФГБОУ ВО ПСПбГМУ им. И.П.Павлова Минздрава России',
	     'СПб ГБУЗ "Городская больница №40"','СПб ГБУЗ "Городская больница №20"','СПб ГБУЗ "Николаевская больница"']

    zan1 = fr.loc[fr['Исход заболевания'].isnull() \
            & fr['Вид лечения'].isin(['Стационарное лечение']) \
            & fr['Субъект РФ'].isin(['г. Санкт-Петербург']) \
            & fr['Медицинская организация'].isin(S_org) \
            & (fr['Диагноз'].str.contains('J12.')  | fr['Диагноз'].str.contains('J18.') | fr['Диагноз'].isin(['U07.1','U07.2']) )  ]

    zan1 ['Занятые койки'] = 1 
    zan1 = zan1.groupby('Медицинская организация',as_index=False)['Занятые койки'].sum()

    zan2 = fr.loc[fr['Исход заболевания'].isnull() \
            & fr['Субъект РФ'].isin(['г. Санкт-Петербург']) \
            & ~fr['Медицинская организация'].isin(S_org) \
            & (fr['Диагноз'].str.contains('J12.')  | fr['Диагноз'].str.contains('J18.') | fr['Диагноз'].isin(['U07.1','U07.2']) )  ]

    zan2 ['Занятые койки'] = 1 
    zan2 = zan2.groupby('Медицинская организация',as_index=False)['Занятые койки'].sum()

    zan = pd.concat([zan1,zan2], ignore_index=True)

    ivl1 = fr.loc[fr['Исход заболевания'].isnull() \
            & fr['ИВЛ'].notnull() \
            & fr['Вид лечения'].isin(['Стационарное лечение']) \
            & fr['Медицинская организация'].isin(S_org) ]

    ivl1 ['ИВЛ'] = 1
    ivl1 = ivl1.groupby('Медицинская организация',as_index=False)['ИВЛ'].sum()

    ivl2 = fr.loc[fr['Исход заболевания'].isnull() \
            & fr['ИВЛ'].notnull() \
            & ~fr['Медицинская организация'].isin(S_org) ]

    ivl2 ['ИВЛ'] = 1
    ivl2 = ivl2.groupby('Медицинская организация',as_index=False)['ИВЛ'].sum()

    ivl = pd.concat([ivl1,ivl2], ignore_index=True)

    df = zan.merge(ivl,how='outer')

    df = df.fillna(0)

    change_mo =[ [r'Больница №20 Поликлиническое отделение №42 (площадка Ленсовета)',r'СПб ГБУЗ "Городская больница №20"']
                ,[r'СПб ГБУЗ "Городская больница №40" площадка Пансионат "Заря"',r'СПб ГБУЗ "Городская больница №40"']
                ,[r'СПб ГБУЗ "Городская Александровская больница"',r'СПб ГБУЗ "Александровская больница"']
                ,[r'СПб ГБУЗ "Клиническая больница Святителя Луки"',r'СПб ГБУЗ Клиническая больница Святителя Луки']
                ,[r'СПб ГБУЗ "Городская больница Святой преподобномученицы Елизаветы"',r'СПб ГБУЗ "Елизаветинская больница"']
                ,[r'СПб ГБУЗ "Детская городская больница Святой Ольги"',r'СПб ГБУЗ "ДГБ Св. Ольги"']
                ,[r'СПб ГБУЗ "Детская городская клиническая больница №5 имени Нила Федоровича Филатова"',r'СПб ГБУЗ "ДГКБ №5 им. Н.Ф.Филатова"']
                ,[r'СПб ГКУЗ "Городская психиатрическая больница №3 им.И.И.Скворцова-Степанова"',r'СПб ГКУЗ "ГПБ №3 им. И.И.Скворцова-Степанова"']
                ,[r'СПб ГБУЗ "ДГМКЦ ВМТ им.К.А.Раухфуса"',r'СПБ ГБУЗ «ДГМКЦ ВМТ им. К.А.Раухфуса»']
                ,[r'СПб ГБУЗ "Центр по профилактике и борьбе со СПИД и инфекционными заболеваниями"',r'СПб ГБУЗ "Центр СПИД и инфекционных заболеваний"']
                ,[r'ФГБОУ ВО ПСПбГМУ им. И.П.Павлова" Минздрава России',r'ФГБОУ ВО ПСПбГМУ им. И.П.Павлова Минздрава России']
                ,[r'ФГБОУ ВО СЗГМУ им.И.И.Мечникова Минздрава России',r'ФГБОУ ВО СЗГМУ им. И.И. Мечникова Минздрава России']
                ,[r'ФГБУЗ «Клиническая больница №122 им.Л.Г.Соколова» ФМБА России',r'ФГБУ "СЗОНКЦ им.Л.Г.Соколова ФМБА России"']
                ,[r'СПб ГБУЗ "Введенская больница"',r'СПБ ГБУЗ "Введенская больница"']
                ,[r'СПб ГБУЗ "Психиатрическая больница №1 им.П.П.Кащенко"',r'СПб ГБУЗ "Санкт-Петербургская психиатрическая больница №1 им.П.П.Кащенко"']
                ,[r'ФГБОУ ВО "Санкт-Петербургский Государственный педиатрический медицинский университет Минздрава России"',r'ФГБОУ ВО СПбГПМУ Минздрава России']
                ,[r'СПб ГБУЗ "Госпиталь для ветеранов войн" площадка "ЛЕНЭКСПО"',r'СПб ГБУЗ "Госпиталь для ветеранов войн"']
                ,[r'ФГБОУ ВО СПБГПМУ МИНЗДРАВА РОССИИ',r'ФГБОУ ВО СПбГПМУ Минздрава России']
                ,[r'СПб ГБУЗ "Детская городская клиническая больница №5 имени Нила Федоровича Филатова"',r'СПб ГБУЗ "ДГКБ №5 им. Н.Ф.Филатова"']
                ,[r'СПб ГБУЗ "Городская многопрофильная больница №2"' , r'СПб ГБУЗ "ГМПБ 2"']
                ,[r'СПб ГБУЗ "Клиническая инфекционная больница им. С.П. Боткина"' , r'СПб ГБУЗ "Больница Боткина"']
                      ]
    vp = vp.fillna(0)
    # Меняем названия МО и сумируем строки 
    zamena = 'произошла замена строки '
    for i in range(len(vp)):
        vp.loc[i,'zan'] = vp.at[i,'vp_zan'] + vp.at[i,'cov_zan']
        vp.loc[i,'ivl'] = vp.at[i,'vp_ivl'] + vp.at[i,'cov_ivl']
        for bad,good in change_mo:
            if vp.loc[i,'mo'] == bad:
                vp.loc[i,'mo'] = good
                zamena += '\n ' + bad + ' на ' + good

    send('epid',zamena)
    vp = vp.groupby('mo',as_index=False)['zan','ivl'].sum()
    vp.index = range(len(vp))

   # получаем отчёт по ИВЛ
    ivl_otchet = vp.merge(df, left_on='mo',right_on='Медицинская организация',how='left')
    del ivl_otchet ['mo']
    del ivl_otchet ['Занятые койки']
    del ivl_otchet ['zan']
    ivl_otchet = ivl_otchet.fillna(0)


    for i in range(len(ivl_otchet)):
        ivl_otchet.loc[i,'Разница'] = ivl_otchet.at[i,'ИВЛ'] - ivl_otchet.at[i,'ivl']

    ivl_otchet.rename(columns = { 'ivl':'ИВЛ из ежедневного отчёта','ИВЛ':'ИВЛ из Фед Регистра'}, inplace = True)

    columnsTitles=['Медицинская организация','ИВЛ из ежедневного отчёта','ИВЛ из Фед Регистра','Разница']
    ivl_otchet = ivl_otchet[ivl_otchet['Медицинская организация'] != 0 ]
    ivl_otchet=ivl_otchet.reindex(columns=columnsTitles)

    ivl_otchet.index = range(1,len(ivl_otchet)+1)

    medorg_ivl_otchet = ivl_otchet['Медицинская организация'].unique()

    # получаем отчёт по койкам
    zan_otchet = vp.merge(df, left_on='mo',right_on='Медицинская организация',how='left')

    del zan_otchet ['mo']
    del zan_otchet ['ИВЛ']
    del zan_otchet['ivl']
    zan_otchet = zan_otchet.fillna(0)

    for i in range(len(zan_otchet)):
        zan_otchet.loc[i,'Разница'] = zan_otchet.at[i,'Занятые койки'] - zan_otchet.at[i,'zan']

    zan_otchet.rename(columns = {'zan':'Заняты койки из ежедневного отчёта','Занятые койки':'Койки из Фед Регистра'}, inplace = True)
    
    columnsTitles=['Медицинская организация','Заняты койки из ежедневного отчёта','Койки из Фед Регистра','Разница']
    zan_otchet=zan_otchet.reindex(columns=columnsTitles)
    zan_otchet = zan_otchet[zan_otchet['Медицинская организация'] != 0 ]
    zan_otchet.index=range(1,len(zan_otchet)+1)

    medorg_zan_otchet = zan_otchet['Медицинская организация'].unique()

    path_otc ='/mnt/COVID-списки/Замечания Мин. Здравоохранения (Своды)/сверка ИВЛ и занятые койки'

    with pd.ExcelWriter(path_otc+ '/' + str(date_otch) + ' пациенты на ИВЛ новый.xlsx') as writer:
        ivl_otchet.to_excel(writer,sheet_name='ИВЛ')
        zan_otchet.to_excel(writer,sheet_name='занятые койки')
    
    put_excel_for_mo(ivl_otchet,'Пациенты на ИВЛ', date_otch)
    put_excel_for_mo(zan_otchet,'Занятые койки', date_otch)
    
    return get_dir('temp') + '/' + 'отчёт по разложению Пациенты на ИВЛ.xlsx' +';'+ get_dir('temp') + '/' + 'отчёт по разложению Занятые койки.xlsx'

def zamechania_mz(a):
    date = pd.read_sql("SELECT max([Дата изменения РЗ]) as 'дата отчета' from robo.v_FedReg",con)
    sql = open('sql/zam/kolichestvo.sql', 'r').read()
    
    df = pd.read_sql(sql,con)
    del df ['Принадлежность']
    df['дата отчета'] = date.iloc[0,0]

    names = [['net_diagnoz_covid.sql', 'Не стоит диагноз ковид' ],
             ['no_snils.sql', 'Без СНИЛСа'],
             ['no_OMS.sql', 'Нет сведений ОМС'],
             ['bez_ishod.sql', 'Без исхода заболевания больше 45 дней'],
             ['no_lab.sql', 'без лабораторного потверждения'],
             ['net_dnevnik.sql','Нет дневниковых записей'],
             ['net_pad.sql', 'Нет ПАД'],
             ['neverni_vid_lecenia.sql','Неверный вид лечения'],
             ['bez_ambulat_level.sql', 'Нет амбулаторного этапа'],
             ['bez_ambulat_level_amb.sql', 'Пациенты зависшие по МО'],
             ['bez_ambulat_level_noMO.sql', 'Пациенты зависшие без МО']]
     
    for file,name in names:
        sql = open('sql/zam/' + file, 'r').read()
        part = pd.read_sql(sql,con)
        part = part.groupby(by=["Медицинская организация"],as_index=False).size()
        part.rename(columns={"size": name}, inplace=True)

        df = df.merge(part, how = "left" , left_on = 'Медицинская организация', right_on = 'Медицинская организация' )

    sql = f"delete from [robo].[cv_Zamechania_fr] where [дата отчета] ='{date.iat[0,0]}'"
    con.execute(sql)
    df.fillna(0, inplace=True)
    df.to_sql('cv_Zamechania_fr',con,schema='robo',if_exists='append',index=False)
    sql = """
    delete from [robo].[cv_Zamechania_fr] where [Медицинская организация] in
            ('ФКУЗ МСЧ № 78 ФСИН России', 'ФГБУ «НМИЦ ТО им. Р.Р. Вредена» Минздрава России',
            'ФГБУ «КДЦ с поликлиникой»', 'ФКУ "Санкт-Петербургская ПБСТИН" Минздрава России',
            'ООО "МЕДУЧРЕЖДЕНИЕ №2"', 'ФГБУ ДНКЦИБ ФМБА России') 
    """
    con.execute(sql)
    return 1
def zamechania_mz_file(a):
    date = pd.read_sql("SELECT max([Дата изменения РЗ]) as 'дата отчета' from robo.v_FedReg",con)
    sql = open('sql/zam/kolichestvo.sql', 'r').read()
    
    df = pd.read_sql(sql,con)
    #df['дата отчета'] = date.iloc[0,0]

    names = [['no_snils.sql', 'Без СНИЛСа'],
             ['no_OMS.sql', 'Нет сведений ОМС'],
             ['bez_ishod.sql', 'Без исхода заболевания больше 30 дней'],
             ['net_dnevnik.sql','Нет дневниковых записей'],
             ['net_pad.sql', 'Нет ПАД'],
             ['neverni_vid_lecenia.sql','Неверный вид лечения'],
             ['bez_ambulat_level.sql', 'Нет амбулаторного этапа'],
             ['bez_ambulat_level_amb.sql', 'Нет амбулаторного этапа (Амб.)'],
             ['bez_ambulat_level_noMO.sql', 'Нет амбулаторного этапа (Без МО)']]
    
    for file,name in names:
        sql = open('sql/zam/' + file, 'r').read()
        part = pd.read_sql(sql,con)
        part = part.groupby(by=["Медицинская организация"],as_index=False).size()
        part.rename(columns={"size": name}, inplace=True)

        df = df.merge(part, how = "left" , left_on = 'Медицинская организация', right_on = 'Медицинская организация' )

    #sql = f"delete from [robo].[cv_Zamechania_fr] where [дата отчета] ='{date.iat[0,0]}'"
    #con.execute(sql)
    df.fillna(0, inplace=True)
    del df['Уникальных пациентов']
    df = df.loc[~(df['Медицинская организация'] == 'МО другого региона') ]
    df.index = range(len(df))    
    
    file = get_dir("temp") + "/Замечания_за_" + str(date.iat[0,0]) + ".xlsx"
    shutil.copyfile(get_dir('help') + '/Zamechania.xlsx',  file)
    
    wb= openpyxl.load_workbook( file)

    ws = wb['main']   
    part = df.copy()
    del part ['Тип организации']
    del part ['Принадлежность']
    rows = dataframe_to_rows(part ,index=False, header=False)
    for r_idx, row in enumerate(rows,5):
        for c_idx, value in enumerate(row, 7):
            ws.cell(row=r_idx, column=c_idx, value=value)
   
    ws = wb['amb']   
    part = df.loc[df['Тип организации'] == 'Амбулаторная' ].copy()
    del part ['Тип организации']
    del part ['Принадлежность']
    rows = dataframe_to_rows(part ,index=False, header=False)
    for r_idx, row in enumerate(rows,5):
        for c_idx, value in enumerate(row, 7):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['amb_dist']   
    part = df.loc[(df['Тип организации'] == 'Амбулаторная' ) & ~(df['Принадлежность'].isin(['комитет здравоохранения','частные','федеральные']) ) ].copy()
    del part ['Тип организации']
    rows = dataframe_to_rows(part ,index=False, header=False)
    for r_idx, row in enumerate(rows,5):
        for c_idx, value in enumerate(row, 7):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['amb_kz']   
    part = df.loc[(df['Тип организации'] == 'Амбулаторная' ) & ( df['Принадлежность'].isin(['комитет здравоохранения']) ) ].copy()
    del part ['Тип организации']
    del part ['Принадлежность']
    rows = dataframe_to_rows(part ,index=False, header=False)
    for r_idx, row in enumerate(rows,5):
        for c_idx, value in enumerate(row, 7):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['amb_fed']   
    part = df.loc[(df['Тип организации'] == 'Амбулаторная' ) & ( df['Принадлежность'].isin(['федеральные']) ) ].copy()
    del part ['Тип организации']
    del part ['Принадлежность']
    rows = dataframe_to_rows(part ,index=False, header=False)
    for r_idx, row in enumerate(rows,5):
        for c_idx, value in enumerate(row, 7):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['amb_ch']   
    part = df.loc[(df['Тип организации'] == 'Амбулаторная' ) & ( df['Принадлежность'].isin(['частные']) ) ].copy()
    del part ['Тип организации']
    del part ['Принадлежность']
    rows = dataframe_to_rows(part ,index=False, header=False)
    for r_idx, row in enumerate(rows,5):
        for c_idx, value in enumerate(row, 7):
            ws.cell(row=r_idx, column=c_idx, value=value)


   
    ws = wb['stat']   
    part = df.loc[df['Тип организации'] == 'Стационарная' ].copy()
    del part ['Тип организации']
    del part ['Принадлежность']
    rows = dataframe_to_rows(part ,index=False, header=False)
    for r_idx, row in enumerate(rows,5):
        for c_idx, value in enumerate(row, 7):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['stat_dist']   
    part = df.loc[(df['Тип организации'] == 'Стационарная' ) & ~(df['Принадлежность'].isin(['комитет здравоохранения','частные','федеральные']) ) ].copy()
    del part ['Тип организации']
    rows = dataframe_to_rows(part ,index=False, header=False)
    for r_idx, row in enumerate(rows,5):
        for c_idx, value in enumerate(row, 7):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['stat_kz']   
    part = df.loc[(df['Тип организации'] == 'Стационарная' ) & ( df['Принадлежность'].isin(['комитет здравоохранения']) )  ].copy()
    del part ['Тип организации']
    del part ['Принадлежность']
    rows = dataframe_to_rows(part ,index=False, header=False)
    for r_idx, row in enumerate(rows,5):
        for c_idx, value in enumerate(row, 7):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['stat_fed']   
    part = df.loc[(df['Тип организации'] == 'Стационарная' ) &  (df['Принадлежность'].isin(['федеральные']) ) ].copy()
    del part ['Тип организации']
    del part ['Принадлежность']
    rows = dataframe_to_rows(part ,index=False, header=False)
    for r_idx, row in enumerate(rows,5):
        for c_idx, value in enumerate(row, 7):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['stat_ch']   
    part = df.loc[(df['Тип организации'] == 'Стационарная' ) & ( df['Принадлежность'].isin(['частные']) ) ].copy()
    del part ['Тип организации']
    del part ['Принадлежность']
    rows = dataframe_to_rows(part ,index=False, header=False)
    for r_idx, row in enumerate(rows,5):
        for c_idx, value in enumerate(row, 7):
            ws.cell(row=r_idx, column=c_idx, value=value)



    wb.save( file )
 

    return file




