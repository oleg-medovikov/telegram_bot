import datetime,glob,os,shutil,openpyxl,pyodbc
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from loader import get_dir,sql_execute
from sqlalchemy import *
from sqlalchemy.orm import sessionmaker
from collections import namedtuple

con = create_engine(os.getenv('sql_engine'),convert_unicode=False)
conn=pyodbc.connect(os.getenv('sql_conn'))

cursor = conn.cursor()

def check_robot():
    date = datetime.datetime.today().strftime("%Y_%m_%d")
    path =os.getenv('path_robot') +'\\'+ date + '\\' +'*'
    spisok = 'В директории Robot сейчас лежат файлы:'
    for file in glob.glob(path):
        spisok += '\n' + file.split('\\')[-1]
    return spisok

def sort_death_mg():
    def search_mo(street,house):
        for mo in mo_org:
            if mo.Street == street:
                if mo.House == house:
                    return mo.Name_MO
        #print('Не найдено МО ' + street + ' ' + house)
        return 0
    date_otch = (datetime.datetime.today() - datetime.timedelta(days=0)).strftime("%d.%m.%Y")
    file =get_dir('sort_death_mg') + r'\[!~]*Умершие от Covid-19*'+ date_otch + '*.xlsx'
    excel = glob.glob(file)
    if len(excel) == 0:
        return 'Я не нашел файлик за сегодня!'
    df = pd.read_excel(excel[0],header=1, usecols = 'A,L,O,AJ,AK,BL,AO' )
    df = df[df['№ п/п'].notnull() & df['№ п/п'] != 0 ]
    df.index = range(len(df))
    mo = namedtuple('mo',['Name_MO','Street','House'])
    sql = """
    SELECT  [Name_MO],[Street],[House]
    FROM [COVID].[Nsi].[Address_MO]
    """
    mo_org=[]
    for row in cursor.execute(sql):
        mo_org.append(mo(*row))
    df.loc[~df['Место смерти'].isin(['в стационаре'])  , 'Name_MO'] = 'БСМЭ\ПАБ'
    for i in df.loc[df['Место смерти'].isin(['в стационаре'])].index:
        if search_mo(df.at[i,'Улица смерти'],df.at[i,'Дом смерти']):
            df.loc[i,'Name_MO'] = search_mo(df.at[i,'Улица смерти'],df.at[i,'Дом смерти'])
        else:
            df.loc[i,'Name_MO'] = df.at[i,'Краткое наименование']

    otchet = df[df['Субъект'] != 'Ленинградская обл'].groupby(by='Name_MO',as_index=False,).count()
    for column in otchet.columns:
        if column not in ('Name_MO','№ п/п'):
            del otchet[column]
    otchet.rename(columns = {'№ п/п':'Всего СПб'}, inplace = True)
    otchet = otchet.merge(df[df['Субъект'] == 'Ленинградская обл'].groupby(by='Name_MO',as_index=False,).count(), how='outer',left_on='Name_MO',right_on='Name_MO')
    otchet.rename(columns = {'Name_MO':'Медицинская организация', '№ п/п':'Всего ЛО'}, inplace = True)

    for column in otchet.columns:
        if column not in ('Медицинская организация','Всего СПб','Всего ЛО'):
            del otchet[column]

    vozrast = [18,65,150]
    status = ['Дети до 18 СПб','Взрослые до 65 СПб','Пенсионеры после 65 СПб','Дети до 18 ЛО','Взрослые до 65 ЛО','Пенсионеры после 65 ЛО']
    for i in range(len(otchet)):
        for stat in status:
            otchet.loc[i,stat] = 0
    for k in range(len(df)):
        if df.at[k,'Субъект'] != 'Ленинградская обл':
            if int(df.at[k,'Возраст']) < vozrast[0]:
                otchet.loc[otchet['Медицинская организация'] == df.at[k,'Name_MO'], status[0] ] += 1
            else:
                if int(df.at[k,'Возраст']) < vozrast[1]:
                    otchet.loc[otchet['Медицинская организация'] == df.at[k,'Name_MO'], status[1] ] += 1
                else:
                    if int(df.at[k,'Возраст']) < vozrast[2]:
                        otchet.loc[otchet['Медицинская организация'] == df.at[k,'Name_MO'], status[2] ] += 1
        else:
            if int(df.at[k,'Возраст']) < vozrast[0]:
                otchet.loc[otchet['Медицинская организация'] == df.at[k,'Name_MO'], status[3] ] += 1
            else:
                if int(df.at[k,'Возраст']) < vozrast[1]:
                    otchet.loc[otchet['Медицинская организация'] == df.at[k,'Name_MO'], status[4] ] += 1
                else:
                    if int(df.at[k,'Возраст']) < vozrast[2]:
                        otchet.loc[otchet['Медицинская организация'] == df.at[k,'Name_MO'], status[5] ] += 1

    shablon = get_dir('sort_death_mg') + r'\Шаблон.xlsx'
    file = get_dir('sort_death_mg') + r'\Свод по возрастам '+ date_otch + '.xlsx'
    shutil.copyfile(shablon,file)
    wb= openpyxl.load_workbook(file)
    ws = wb['Свод по возрастам']
    rows = dataframe_to_rows(otchet,index=False, header=False)

    for r_idx, row in enumerate(rows,2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(file)
    return 'Сгенерирован файл' + file.split('\\')[-1]

def svod_40_COVID_19():
    path = get_dir('40_covid_19') + r'\[!~]*.xls'
    list_=[]
    usecolumns = 'A,B,C,D,F,G,I,J,L,M,O,P,R,S,U,V,X,Y,AA,AB,AC,AD,AF,AG,AI,AJ,AK'
    date_otch = None
    for xls in glob.glob(path):
        mo = pd.read_excel(xls, header=2,usecols=usecolumns,sheet_name='Для заполнения')
        mo = mo[mo[2].notnull() & (~mo[2].isin(['Пункт вакцинации'])) ]
        if date_otch is None:
            date_otch = pd.to_datetime(xls.split('\\')[-1].split('_Основной')[0][-1-9:],format="%d.%m.%Y").date()
        if len(mo.index) > 1:
            for i in mo.index:
                if i:
                    mo.loc[i,0] = 'Пункт вакцинации'
                    mo.loc[i,1] = mo.at[i,2].split(' ')[0]
                    mo.loc[i,2] = str(mo.at[i,2].split(' ',1)[1])
                else:
                    mo.loc[i,0] = 'Медицинская организация'

            list_.append(mo)

    if len(list_):
        df=pd.concat(list_)

        cols = list(df.columns.values)
        cols = cols[-1:] + cols[:-1]
        df =df[cols]
        df.index=range(1,len(df)+1)

        itog = pd.DataFrame()
        potreb = pd.DataFrame()
        for col in df.columns:
            if col not in [0,1,2,24,26,27,28]:
                itog.loc[0,col] = df.loc[df[0] == 'Медицинская организация', col ].sum()

        for i in range(len(df)):
            k = len(potreb)
            potreb.loc[k,'Date'] = date_otch
            potreb.loc[k,'District'] = df.iat[i,1]
            potreb.loc[k,'Type'] = df.iat[i,0]
            potreb.loc[k,'Name'] = df.iat[i,2]
            potreb.loc[k,'Vsego_day_v1'] = df.iat[i,21]
            potreb.loc[k,'Vsego_day_v2'] = df.iat[i,23]
            potreb.loc[k,'Ostatok_v1'] = df.iat[i,4] - df.iat[i,20] - df.iat[i,24]
            potreb.loc[k,'Ostatok_v2'] = df.iat[i,4] - df.iat[i,22]
            potreb.loc[k,'Potrebnost_v1'] = df.iat[i,21]
            potreb.loc[k,'Potrebnost_v2'] = df.iat[i,23]

        sql_execute(f"delete from robo.vaccine_potrebnost where cast([Date] as date) = '{str(date_otch)}' select 1")
        potreb.to_sql('vaccine_potrebnost',con,schema='robo',if_exists = 'append',index=False)
        new_name = str(date_otch) + '_40_COVID_19_cvod.xlsx'

        shablon_path = os.getenv('path_help')
        shutil.copyfile(shablon_path + r'\40_COVID_19_svod.xlsx', shablon_path  + '\\' + new_name)

        wb= openpyxl.load_workbook( shablon_path  + '\\' + new_name)
        ws = wb['Пунты вакцинации']
        rows = dataframe_to_rows(df,index=False, header=False)
        for r_idx, row in enumerate(rows,5):  
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        wb.save( shablon_path  + '\\' + new_name) 

#        wb= openpyxl.load_workbook( shablon_path  + '\\' + new_name)
#        ws = wb['ИТОГО']
#        rows = dataframe_to_rows(itog,index=False, header=False)
#        for r_idx, row in enumerate(rows,4):  
#            for c_idx, value in enumerate(row, 1):
#                ws.cell(row=r_idx, column=c_idx, value=value)
#        wb.save( shablon_path  + '\\' + new_name) 
        return(shablon_path  + '\\' + new_name)
    else:
        return None 
